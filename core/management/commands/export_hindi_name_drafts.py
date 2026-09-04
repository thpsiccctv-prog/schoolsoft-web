import csv
from pathlib import Path

from django.core.management.base import BaseCommand

from core.models import Student


DEFAULT_OUT_DIR = r"E:\THPSIC-INTER-COLLEGE\05-reports\hindi-name-drafts"
WARNING_TEXT = "YEH DRAFT HAI - Physical Scholar Register se verify karne ke baad hi upload karein."


class Command(BaseCommand):
    help = "Export Class IX and XI Hindi name draft workbooks for office verification. Does not modify DB."

    def add_arguments(self, parser):
        parser.add_argument("--out-dir", default=DEFAULT_OUT_DIR, help="Output folder for draft Excel/CSV files.")
        parser.add_argument("--classes", default="IX,XI", help="Comma-separated class prefixes, default IX,XI.")

    def handle(self, *args, **options):
        out_dir = Path(options["out_dir"])
        out_dir.mkdir(parents=True, exist_ok=True)

        # Reuse the already-tested local transliteration dictionary/rules. It only creates suggestions.
        from generate_hindi_master_v3 import transliterate_name

        exports = []
        for prefix in [p.strip().upper() for p in options["classes"].split(",") if p.strip()]:
            rows = self._rows_for_prefix(prefix, transliterate_name)
            if not rows:
                self.stdout.write(self.style.WARNING(f"No active students found for class prefix {prefix}."))
                continue

            label = "CLASS9" if prefix == "IX" else "CLASS11" if prefix == "XI" else prefix.replace(" ", "_")
            xlsx_path = out_dir / f"{label}_HINDI_NAMES_DRAFT.xlsx"
            csv_path = out_dir / f"{label}_HINDI_NAMES_DRAFT.csv"
            self._write_csv(csv_path, rows)
            self._write_xlsx(xlsx_path, rows, title=f"{label} Hindi Names Draft")
            exports.append((prefix, len(rows), xlsx_path, csv_path))

        report_path = out_dir / "HINDI_NAME_DRAFTS_README.txt"
        report_path.write_text(
            "\n".join(
                [
                    WARNING_TEXT,
                    "",
                    "Workflow:",
                    "1. Clerk opens the *_DRAFT.xlsx file.",
                    "2. Yellow verified_* columns are only suggestions; verify every name from the physical scholar register.",
                    "3. Correct spellings directly in yellow verified_* columns.",
                    "4. Leave blank if unsure. Do not guess for board records.",
                    "5. DB upload/apply is a separate future step after office verification.",
                    "",
                    "Generated files:",
                    *[f"- {prefix}: {count} students | {xlsx} | {csv}" for prefix, count, xlsx, csv in exports],
                ]
            ),
            encoding="utf-8",
        )

        self.stdout.write(self.style.SUCCESS("Hindi name draft export complete. DB not modified."))
        for prefix, count, xlsx_path, csv_path in exports:
            self.stdout.write(f"- {prefix}: {count} students")
            self.stdout.write(f"  Excel: {xlsx_path}")
            self.stdout.write(f"  CSV:   {csv_path}")
        self.stdout.write(f"- Readme: {report_path}")

    def _rows_for_prefix(self, prefix, transliterate_name):
        qs = (
            Student.objects.filter(is_active=True, current_class__name__istartswith=prefix)
            .select_related("current_class", "current_section")
            .order_by("current_class__name", "current_section__name", "legacy_sid", "admission_no")
        )
        rows = []
        for student in qs:
            s_hint, s_conf, s_notes = transliterate_name(student.full_name or "", context_type="student")
            f_hint, f_conf, f_notes = transliterate_name(student.father_name or "", context_type="father")
            m_hint, m_conf, m_notes = transliterate_name(student.mother_name or "", context_type="mother")
            overall = self._overall_flag([s_conf, f_conf, m_conf])
            rows.append(
                {
                    "legacy_sid": student.legacy_sid or "",
                    "admission_no": student.admission_no or "",
                    "class": student.current_class.name if student.current_class else "",
                    "section": student.current_section.name if student.current_section else "",
                    "student_name_english": student.full_name or "",
                    "father_name_english": student.father_name or "",
                    "mother_name_english": student.mother_name or "",
                    "current_db_student_hindi": student.full_name_hindi or "",
                    "current_db_father_hindi": student.father_name_hindi or "",
                    "current_db_mother_hindi": student.mother_name_hindi or "",
                    "suggested_student_hindi": s_hint,
                    "student_confidence": s_conf,
                    "suggested_father_hindi": f_hint,
                    "father_confidence": f_conf,
                    "suggested_mother_hindi": m_hint,
                    "mother_confidence": m_conf,
                    "verified_student_hindi": student.full_name_hindi or s_hint,
                    "verified_father_hindi": student.father_name_hindi or f_hint,
                    "verified_mother_hindi": student.mother_name_hindi or m_hint,
                    "overall_review_flag": overall,
                    "notes": "VERIFY FROM PHYSICAL REGISTER; " + " | ".join(
                        note for note in [s_notes, f_notes, m_notes] if note
                    ),
                }
            )
        return rows

    def _overall_flag(self, confidences):
        if any(c == "LOW" for c in confidences):
            return "LOW"
        if any(c == "MEDIUM" for c in confidences):
            return "MEDIUM"
        return "HIGH"

    def _write_csv(self, path, rows):
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)

    def _write_xlsx(self, path, rows, title):
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Draft"

        warning_fill = PatternFill("solid", fgColor="FFF2CC")
        header_fill = PatternFill("solid", fgColor="0F766E")
        verified_fill = PatternFill("solid", fgColor="FFF2CC")
        current_fill = PatternFill("solid", fgColor="E5E7EB")
        low_fill = PatternFill("solid", fgColor="FCA5A5")
        medium_fill = PatternFill("solid", fgColor="FDE68A")
        high_fill = PatternFill("solid", fgColor="BBF7D0")
        thin = Side(style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(rows[0]))
        ws.cell(1, 1).value = WARNING_TEXT
        ws.cell(1, 1).font = Font(bold=True, color="9A3412", size=13)
        ws.cell(1, 1).fill = warning_fill
        ws.cell(1, 1).alignment = Alignment(horizontal="center")

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(rows[0]))
        ws.cell(2, 1).value = f"{title} - suggestions only, DB not modified"
        ws.cell(2, 1).font = Font(bold=True, color="0F172A", size=12)
        ws.cell(2, 1).alignment = Alignment(horizontal="center")

        headers = list(rows[0].keys())
        header_row = 3
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(header_row, col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row_idx, row in enumerate(rows, start=4):
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row_idx, col_idx)
                cell.value = row[header]
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if header.startswith("verified_") or header.startswith("suggested_"):
                    cell.fill = verified_fill
                    cell.font = Font(name="Nirmala UI", size=10)
                elif header.startswith("current_db_"):
                    cell.fill = current_fill
                    cell.font = Font(name="Nirmala UI", size=10)
                elif header == "overall_review_flag":
                    cell.fill = {"LOW": low_fill, "MEDIUM": medium_fill, "HIGH": high_fill}.get(row[header], warning_fill)
                    cell.font = Font(bold=True)

        ws.freeze_panes = "A4"
        ws.auto_filter.ref = ws.dimensions

        widths = {
            "student_name_english": 24,
            "father_name_english": 24,
            "mother_name_english": 24,
            "current_db_student_hindi": 24,
            "current_db_father_hindi": 24,
            "current_db_mother_hindi": 24,
            "suggested_student_hindi": 24,
            "suggested_father_hindi": 24,
            "suggested_mother_hindi": 24,
            "verified_student_hindi": 26,
            "verified_father_hindi": 26,
            "verified_mother_hindi": 26,
            "notes": 48,
        }
        for col_idx, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 14)

        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 22
        wb.save(path)
