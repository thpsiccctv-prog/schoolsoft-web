import csv
import os
import shutil
import sqlite3
import subprocess
import tempfile
from datetime import date, datetime, time
from pathlib import Path
from decimal import Decimal
from urllib.parse import urlencode

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db import connection, transaction
from django.db.models import Count, F, Max, ProtectedError, Q, Sum
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from .access import (
    is_online_deployment,
    user_can_access,
    user_can_manage_users,
    user_is_readonly,
)
from .board_registration_exports import build_rows, export_columns, export_filename
from .fee_engine import (
    PACKAGE_FEE_HEAD_NAME,
    calculate_student_due,
    calculate_structure_receipt_amount,
    package_receipt_default_amount,
    student_is_zero_fee,
    student_uses_fee_package,
)
from .forms import DisciplineRecordForm, FamilyForm, FeeReceiptEditForm, FeeReceiptEntryForm, FeeReceiptLineEntryForm, InventoryIssueForm, InventoryItemForm, SalaryPaymentForm, StudentForm, TransferCertificateForm
from .models import (
    ACADEMIC_MONTHS,
    AcademicSession,
    DisciplineRecord,
    ExamMark,
    ExamTerm,
    Family,
    FeeHead,
    FeeReceipt,
    FeeReceiptAuditLog,
    FeeStructure,
    House,
    InventoryIssue,
    InventoryItem,
    SalaryPayment,
    SalaryPaymentAuditLog,
    SchoolClass,
    SchoolProfile,
    Section,
    Staff,
    Student,
    StudentConcession,
    StudentTransport,
    TransferCertificate,
    TransportBus,
    TransportRoute,
    Voucher,
)
from .whatsapp import build_wa_link, family_due_message
from .pdf import (
    build_admission_form_pdf,
    build_character_certificate_pdf,
    build_discipline_summary_pdf,
    build_due_report_pdf,
    build_due_slip_pdf,
    build_due_up_to_month_report_pdf,
    build_fee_receipt_pdf,
    build_fee_receipt_pdf_2up,
    build_id_card_batch_pdf,
    build_id_card_pdf,
    build_staff_id_card_pdf,
    build_staff_id_card_batch_pdf,
    build_marksheet_pdf,
    build_salary_payslip_pdf,
    build_scholar_register_book_pdf,
    build_scholar_register_index_pdf,
    build_scholar_register_pdf,
    build_transfer_certificate_pdf,
)

APP_TITLE = os.environ.get("SCHOOLSOFT_APP_TITLE", "THPSIC SchoolSoft")
APP_DATA_DIR_NAME = os.environ.get("SCHOOLSOFT_APP_DATA_DIR_NAME", "THPSIC-InterCollege-SchoolSoft")
DEFAULT_BACKUP_ROOT = "E:/THPSIC-INTER-COLLEGE/04-backups/daily-db"


def get_active_school_profile():
    return SchoolProfile.objects.filter(is_active=True).first()


def apply_receipt_student_snapshot(receipt):
    student = receipt.student
    receipt.student_name_snapshot = student.full_name or ""
    receipt.father_name_snapshot = student.father_name or ""
    receipt.class_snapshot = student.current_class.name if student.current_class else ""
    receipt.section_snapshot = student.current_section.name if student.current_section else ""



def _normalise_academic_month(value):
    value = str(value or "").strip().upper()
    if not value:
        return ""
    if value in ACADEMIC_MONTHS:
        return value
    aliases = {
        "APRIL": "APR",
        "JUNE": "JUN",
        "JULY": "JUL",
        "SEPT": "SEP",
        "SEPTEMBER": "SEP",
        "MARCH": "MAR",
    }
    return aliases.get(value, value[:3])


def _receipt_due_status(receipt):
    if receipt.is_cancelled or not receipt.student_id or not receipt.session_id:
        return None
    target_month = _normalise_academic_month(receipt.to_month or receipt.from_month) or ACADEMIC_MONTHS[-1]
    if target_month not in ACADEMIC_MONTHS:
        target_month = ACADEMIC_MONTHS[-1]

    try:
        target_result = calculate_student_due(
            student=receipt.student,
            session=receipt.session,
            through_month=target_month,
        )
    except (ValidationError, AttributeError) as exc:
        return {
            "available": False,
            "target_month": target_month,
            "error": str(exc),
        }

    receipt_month = _normalise_academic_month(receipt.from_month) or target_month
    if receipt.receipt_date and receipt.session.starts_on and receipt.session.ends_on:
        # Paid-through is judged from the receipt-date month onward. A July
        # receipt for April dues should not be called April-unpaid only because
        # it was entered after the April cutoff.
        receipt_month = _normalise_academic_month(receipt.receipt_date.strftime("%b")) or receipt_month
    if receipt_month not in ACADEMIC_MONTHS:
        receipt_month = target_month

    start_index = min(ACADEMIC_MONTHS.index(receipt_month), ACADEMIC_MONTHS.index(target_month))
    target_index = ACADEMIC_MONTHS.index(target_month)
    clear_through = ""
    next_due_month = ""
    for month in ACADEMIC_MONTHS[start_index : target_index + 1]:
        try:
            month_result = calculate_student_due(
                student=receipt.student,
                session=receipt.session,
                through_month=month,
            )
        except ValidationError:
            break
        if month_result.due_amount <= Decimal("0.00"):
            clear_through = month
        elif not next_due_month:
            next_due_month = month

    return {
        "available": True,
        "target_month": target_month,
        "result": target_result,
        "clear_through": clear_through,
        "next_due_month": next_due_month,
        "has_due": target_result.due_amount > Decimal("0.00"),
        "has_credit": target_result.credit_amount > Decimal("0.00"),
    }


def _student_due_status_payload(student, session, target_month):
    month_results = []
    clear_through = ""
    next_due_month = ""
    target_index = ACADEMIC_MONTHS.index(target_month)

    for month in ACADEMIC_MONTHS[: target_index + 1]:
        month_result = calculate_student_due(
            student=student,
            session=session,
            through_month=month,
        )
        is_clear = month_result.due_amount <= Decimal("0.00")
        if is_clear:
            clear_through = month
        elif not next_due_month:
            next_due_month = month
        month_results.append(
            {
                "month": month,
                "gross_demand": str(month_result.gross_demand),
                "received_amount": str(month_result.received_amount),
                "due_amount": str(month_result.due_amount),
                "credit_amount": str(month_result.credit_amount),
                "status": "clear" if is_clear else "due",
            }
        )

    target_due_result = calculate_student_due(
        student=student,
        session=session,
        through_month=target_month,
    )
    target_result = month_results[-1]
    latest_receipt = (
        FeeReceipt.objects.filter(
            student=student,
            session=session,
            is_cancelled=False,
        )
        .order_by("-receipt_date", "-id")
        .first()
    )
    last_payment = None
    if latest_receipt:
        month_range = " - "
        if latest_receipt.from_month or latest_receipt.to_month:
            month_range = f"{latest_receipt.from_month or '-'} to {latest_receipt.to_month or '-'}"
        last_payment = {
            "receipt_no": latest_receipt.receipt_no,
            "date": latest_receipt.receipt_date.strftime("%d/%m/%Y"),
            "amount": str(latest_receipt.received_amount),
            "month_range": month_range,
            "is_legacy": latest_receipt.carried_forward,
        }

    return {
        "available": True,
        "target_month": target_month,
        "scheduled_fee_demand": str(target_due_result.scheduled_fee_demand),
        "opening_balance_amount": str(target_due_result.opening_balance_amount),
        "gross_demand": target_result["gross_demand"],
        "received_amount": target_result["received_amount"],
        "due_amount": target_result["due_amount"],
        "credit_amount": target_result["credit_amount"],
        "has_due": Decimal(target_result["due_amount"]) > Decimal("0.00"),
        "has_credit": Decimal(target_result["credit_amount"]) > Decimal("0.00"),
        "clear_through": clear_through,
        "next_due_month": next_due_month,
        "last_payment": last_payment,
        "month_results": month_results,
    }


def _format_file_mtime(path):
    try:
        return datetime.fromtimestamp(Path(path).stat().st_mtime).strftime("%d/%m/%Y %I:%M %p")
    except OSError:
        return None




def _format_sync_complete_value(raw_value):
    raw_value = (raw_value or "").strip()
    for fmt in ("%d/%m/%Y %H:%M:%S.%f", "%d/%m/%Y %H:%M:%S", "%d-%m-%Y %H:%M:%S.%f", "%d-%m-%Y %H:%M:%S"):
        try:
            return datetime.strptime(raw_value, fmt).strftime("%d/%m/%Y %I:%M %p")
        except ValueError:
            pass
    return raw_value

def _sync_complete_time_from_lines(lines):
    for line in reversed(lines):
        if line.startswith("Sync complete at "):
            return _format_sync_complete_value(line.replace("Sync complete at ", "", 1).strip())
    return None


def _sync_marker_paths():
    paths = []
    local_appdata = os.environ.get("LOCALAPPDATA")
    if local_appdata:
        paths.append(Path(local_appdata) / APP_DATA_DIR_NAME / "sync-success.marker")
    paths.append(Path(settings.BASE_DIR) / "sync-backups" / "sync-success.marker")
    return paths


def _read_last_sync_success(sync_log):
    marker_paths = _sync_marker_paths()
    for marker in marker_paths:
        try:
            marker_time = _sync_complete_time_from_lines(marker.read_text(encoding="utf-8", errors="ignore").splitlines())
        except OSError:
            marker_time = None
        if marker_time:
            return marker_time, str(marker)

    try:
        log_time = _sync_complete_time_from_lines(sync_log.read_text(encoding="utf-8", errors="ignore").splitlines())
    except OSError:
        log_time = None
    if log_time:
        return log_time, str(sync_log)
    return None, str(marker_paths[0] if marker_paths else sync_log)

def _latest_backup_info():
    backup_root = Path(os.environ.get("SCHOOLSOFT_BACKUP_ROOT", DEFAULT_BACKUP_ROOT))
    try:
        backups = [item for item in backup_root.iterdir() if item.is_dir()]
    except OSError:
        backups = []
    if not backups:
        return {
            "label": "Last Backup",
            "value": "No backup found",
            "detail": str(backup_root),
            "tone": "danger",
            "warning": "Database backup nahi mila!",
        }

    latest = max(backups, key=lambda item: item.stat().st_mtime)
    backup_time = datetime.fromtimestamp(latest.stat().st_mtime)
    elapsed = datetime.now() - backup_time
    elapsed_days = max(0, elapsed.days)
    tone = "ok"
    warning = None
    if elapsed_days >= 3:
        tone = "danger"
        warning = f"Backup {elapsed_days} din se nahi hua!"
    elif elapsed_days >= 1:
        tone = "warn"
        warning = f"Backup {elapsed_days} din se nahi hua"

    return {
        "label": "Last Backup",
        "value": _format_file_mtime(latest) or latest.name,
        "detail": str(latest),
        "tone": tone,
        "warning": warning,
    }

def _backup_root():
    return Path(os.environ.get("SCHOOLSOFT_BACKUP_ROOT", DEFAULT_BACKUP_ROOT))


def _new_backup_dir(root):
    stamp = timezone.localtime().strftime("%Y%m%d-%H%M%S")
    candidate = root / stamp
    if not candidate.exists():
        return candidate
    for index in range(1, 100):
        candidate = root / f"{stamp}-{index:02d}"
        if not candidate.exists():
            return candidate
    raise OSError("Backup folder name available nahi mila.")


def _active_sqlite_db_path():
    if connection.vendor != "sqlite":
        raise ValueError("Backup Now sirf Desktop/SQLite database ke liye available hai.")
    explicit_path = os.environ.get("SCHOOLSOFT_SQLITE_PATH")
    if explicit_path:
        return Path(explicit_path)
    live_desktop_path = Path(os.environ.get("LOCALAPPDATA", "")) / APP_DATA_DIR_NAME / "db.sqlite3"
    if live_desktop_path.exists():
        return live_desktop_path
    db_name = connection.settings_dict.get("NAME")
    if not db_name:
        raise ValueError("Active database path nahi mila.")
    return Path(db_name)

def _restore_note_text(db_path, backup_dir):
    return (
        f"{APP_TITLE} Backup Restore Note\n"
        "===============================\n\n"
        f"Backup folder: {backup_dir}\n"
        f"Source DB: {db_path}\n\n"
        "Restore rule:\n"
        f"1. {APP_TITLE} EXE ko poori tarah band karein.\n"
        "2. Current live DB ka alag dated backup banaye bina overwrite na karein.\n"
        f"3. Is folder ke db.sqlite3 ko %LOCALAPPDATA%\\{APP_DATA_DIR_NAME}\\db.sqlite3 par copy karein.\n"
        f"4. Agar media folder hai to use %LOCALAPPDATA%\\{APP_DATA_DIR_NAME}\\media me restore karein.\n"
        "5. Restore ke baad dashboard counts aur cash book verify karein.\n"
    )


def _perform_backup_now():
    db_path = _active_sqlite_db_path()
    backup_root = _backup_root()
    backup_dir = _new_backup_dir(backup_root)

    backup_root.mkdir(parents=True, exist_ok=True)
    backup_dir.mkdir()

    with connection.cursor() as cursor:
        cursor.execute("PRAGMA wal_checkpoint(TRUNCATE);")

    backup_db = backup_dir / "db.sqlite3"
    if db_path.exists():
        shutil.copy2(db_path, backup_db)
    else:
        connection.ensure_connection()
        backup_connection = sqlite3.connect(backup_db)
        try:
            connection.connection.backup(backup_connection)
        finally:
            backup_connection.close()

    media_root = Path(settings.MEDIA_ROOT)
    if media_root.exists() and media_root.is_dir():
        shutil.copytree(media_root, backup_dir / "media")

    (backup_dir / "RESTORE-NOTE.txt").write_text(
        _restore_note_text(db_path, backup_dir),
        encoding="utf-8",
    )
    return backup_dir

def _dashboard_system_status():
    sync_log = Path(settings.BASE_DIR) / "sync-backups" / "sync-last.log"
    sync_time, sync_detail = _read_last_sync_success(sync_log)
    sync_status = {
        "label": "Last Online Sync",
        "value": sync_time or "No successful sync found",
        "detail": sync_detail,
        "tone": "ok" if sync_time else "warn",
    }
    return [_latest_backup_info(), sync_status]



def _sync_batch_candidates():
    base_dir = Path(settings.BASE_DIR).resolve()
    candidates = [base_dir / "sync-desktop-to-online.bat"]
    for parent in [base_dir, *base_dir.parents]:
        candidates.append(parent / "01-source-code" / "schoolsoft_web" / "sync-desktop-to-online.bat")
    return candidates


def _find_sync_batch_file():
    for candidate in _sync_batch_candidates():
        if candidate.exists():
            return candidate
    return None


def online_sync_start(request):
    if request.method != "POST":
        return redirect("core:dashboard")

    sync_bat = _find_sync_batch_file()
    if not sync_bat:
        messages.error(request, "Online sync script nahi mila. Source folder check kijiye.")
        return redirect("core:dashboard")

    try:
        subprocess.Popen(
            ["cmd.exe", "/k", "call", sync_bat.name, "/auto"],
            cwd=str(sync_bat.parent),
            creationflags=getattr(subprocess, "CREATE_NEW_CONSOLE", 0),
        )
    except OSError as exc:
        messages.error(request, f"Online sync start nahi hua: {exc}")
    else:
        messages.success(
            request,
            "Online sync window start ho gaya. Sync complete hone tak data entry na karein.",
        )
    return redirect("core:dashboard")

def backup_start(request):
    if request.method != "POST":
        return redirect("core:dashboard")

    try:
        backup_dir = _perform_backup_now()
    except (OSError, ValueError) as exc:
        messages.error(request, f"Backup nahi bana: {exc}")
    else:
        messages.success(request, f"Backup complete: {backup_dir}")
    return redirect("core:dashboard")

def dashboard(request):
    user = request.user
    today = timezone.localdate()
    active_session = AcademicSession.objects.filter(is_active=True).order_by("-starts_on", "name").first()
    readonly = user_is_readonly(user)

    # Each KPI/tile's underlying query only runs if the user can actually see
    # it - this avoids both a misleading dashboard (cards linking to a 403)
    # and unnecessary work/exposure of numbers the user isn't permitted to view.
    kpis = []
    if user_can_access(user, "fee_collection"):
        today_received = FeeReceipt.objects.filter(receipt_date=today, is_cancelled=False).aggregate(
            total=Sum("received_amount")
        )["total"] or Decimal("0.00")

        # Previous-collection-day context (Stage 2): find the most recent earlier
        # date that actually had money collected (not just "yesterday" - Sundays/
        # holidays have zero receipts, so this naturally skips them without
        # needing a Holiday/working-day calendar model). Deliberately no
        # zyada/kam delta yet - comparing a partial "today so far" total against
        # a previous complete day would be misleading most mornings.
        prev_date = (
            FeeReceipt.objects.filter(receipt_date__lt=today, is_cancelled=False, received_amount__gt=0)
            .order_by("-receipt_date")
            .values_list("receipt_date", flat=True)
            .first()
        )
        prev_total = None
        prev_days_ago = None
        if prev_date:
            prev_days_ago = (today - prev_date).days
            if prev_days_ago <= 13:
                prev_total = FeeReceipt.objects.filter(
                    receipt_date=prev_date, is_cancelled=False
                ).aggregate(total=Sum("received_amount"))["total"] or Decimal("0.00")

        trend_pct = None
        if prev_total and prev_total > 0:
            trend_pct = round(((today_received - prev_total) / prev_total) * 100)

        kpis.append({
            "label": "Today's Collection",
            "value": today_received,
            "tone": "success",
            "icon": "collection",
            "currency": True,
            "empty_caption": "Abhi tak koi collection nahi",
            "prev_date": prev_date,
            "prev_total": prev_total,
            "prev_days_ago": prev_days_ago,
            "trend_pct": trend_pct,
        })

        current_month_start = today.replace(day=1)
        current_month_collected = FeeReceipt.objects.filter(
            receipt_date__gte=current_month_start,
            receipt_date__lte=today,
            is_cancelled=False
        ).aggregate(total=Sum("received_amount"))["total"] or Decimal("0.00")

        kpis.append({
            "label": "This Month Collection",
            "value": current_month_collected,
            "tone": "success",
            "icon": "collection",
            "currency": True,
            "caption": "Current calendar month",
        })

        receipts_today = FeeReceipt.objects.filter(
            receipt_date=today,
            is_cancelled=False
        ).count()

        kpis.append({
            "label": "Receipts Issued Today",
            "value": receipts_today,
            "tone": "neutral",
            "icon": "book",
            "currency": False,
            "caption": "Total receipts generated today",
        })

    # "Today's Expense" is cash-basis, matching the Cash Book: Daily Expense
    # vouchers + today's cash salary payments. Salary IS kharcha for the
    # office - showing only diesel Rs 1,000 while a Rs 5,000 salary slip went
    # out the same day makes the card misleading. Each half is gated on its
    # own module permission so a user only ever sees numbers they could reach
    # via the corresponding pages (accounts -> vouchers, staff -> salary).
    can_see_vouchers = user_can_access(user, "accounts")
    can_see_salary = user_can_access(user, "staff")
    if can_see_vouchers or can_see_salary:
        today_expenses = Decimal("0.00")
        if can_see_vouchers:
            today_expenses += Voucher.objects.filter(
                voucher_date=today,
                is_cancelled=False,
                voucher_type=Voucher.VoucherType.CASH_PAYMENT,
            ).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
        if can_see_salary:
            # Net pay (gross - deductions - advance recovery) and CASH mode
            # only - the same amount/filters the Cash Book uses for salary
            # rows, so the card and the Cash Book never disagree. Non-cash
            # salary (bank transfer/cheque) is deliberately excluded.
            today_expenses += SalaryPayment.objects.filter(
                payment_date=today,
                is_cancelled=False,
                payment_mode=SalaryPayment.PaymentMode.CASH,
            ).aggregate(total=Sum("amount_paid"))["total"] or Decimal("0.00")
        kpis.append({
            "label": "Today's Expense",
            "value": today_expenses,
            "tone": "outflow",
            "icon": "expense",
            "currency": True,
            "caption": "Salary payments included" if can_see_salary else None,
        })

    if user_can_access(user, "dues"):
        total_dues = Decimal("0.00")
        if active_session:
            total_dues = FeeReceipt.objects.filter(
                is_cancelled=False, carried_forward=False, session=active_session, student__is_active=True
            ).aggregate(total=Sum("legacy_due_amount"))["total"] or Decimal("0.00")
        kpis.append({
            "label": "Receipt Dues (Legacy)",
            "value": total_dues,
            "tone": "attention",
            "icon": "dues",
            "currency": True,
            "caption": "Old receipt-based balance",
        })

    student_total = 0
    active_students = 0
    if user_can_access(user, "students"):
        student_counts = Student.objects.aggregate(
            total=Count("pk"),
            active=Count("pk", filter=Q(is_active=True)),
        )
        student_total = student_counts["total"]
        active_students = student_counts["active"]
        kpis.append({
            "label": "Active Students",
            "value": active_students,
            "tone": "neutral",
            "icon": "students",
            "currency": False,
        })

    tiles = []
    if user_can_access(user, "fee_collection") and not readonly:
        tiles.append({
            "url": reverse("core:receipt_create"), "color": "positive",
            "label": "Fee Collection", "sub": "Daily counter receipt desk",
            "icon": "receipt", "group": "finance"
        })
    if user_can_access(user, "students"):
        tiles.append({
            "url": reverse("core:student_list"), "color": "neutral",
            "label": "Students", "sub": "total", "sub_value": student_total,
            "count": active_students, "icon": "students", "group": "academics_admin"
        })
    if user_can_access(user, "receipts"):
        receipt_count = FeeReceipt.objects.filter(session=active_session, is_cancelled=False).count() if active_session else 0
        tiles.append({
            "url": reverse("core:receipt_list"), "color": "neutral",
            "label": "Receipts", "sub": "Register and PDFs",
            "count": receipt_count, "icon": "book", "group": "finance"
        })
    if user_can_access(user, "dues"):
        tiles.append({
            "url": reverse("core:due_up_to_month_report"), "color": "attention",
            "label": "Dues", "sub": "Due up to selected month", "icon": "clock", "group": "finance"
        })
    if user_can_access(user, "marks"):
        tiles.append({
            "url": reverse("core:marks_report"), "color": "neutral",
            "label": "Marks", "sub": "Results and marksheets", "icon": "cap", "group": "academics_admin"
        })
    if user_can_access(user, "collection"):
        tiles.append({
            "url": reverse("core:collection_report"), "color": "positive",
            "label": "Collection", "sub": "Daily collection report", "icon": "chart", "group": "finance"
        })
    if user_can_access(user, "fee_setup"):
        tiles.append({
            "url": reverse("core:fee_structure_report"), "color": "neutral",
            "label": "Fee Setup", "sub": "Heads and structure",
            "count": FeeHead.objects.count(), "icon": "grid", "group": "finance"
        })
    if user_can_access(user, "staff"):
        tiles.append({
            "url": reverse("core:staff_list"), "color": "neutral",
            "label": "Staff", "sub": "Teacher records", "icon": "people", "group": "academics_admin"
        })
    if user_can_access(user, "transport"):
        tiles.append({
            "url": reverse("core:transport_list"), "color": "neutral",
            "label": "Transport", "sub": "Routes and buses", "icon": "bus", "group": "academics_admin"
        })
    if user_can_access(user, "school_profile"):
        tiles.append({
            "url": reverse("core:school_profile_detail"), "color": "neutral",
            "label": "School Profile", "sub": "Identity and settings", "icon": "building", "group": "academics_admin"
        })

    for tile in tiles:
        tile.setdefault("count", None)
        tile.setdefault("sub_value", None)

    has_finance_tiles = any(t.get("group") == "finance" for t in tiles)
    has_academics_tiles = any(t.get("group") == "academics_admin" for t in tiles)

    context = {
        "today": today,
        "active_session": active_session,
        "dashboard_kpis": kpis,
        "tiles": tiles,
        "has_finance_tiles": has_finance_tiles,
        "has_academics_tiles": has_academics_tiles,
        "system_status": _dashboard_system_status(),
        "can_new_receipt": user_can_access(user, "fee_collection") and not readonly,
        "can_due_report": user_can_access(user, "dues"),
        "can_online_sync": (
            user_can_manage_users(user)
            and not is_online_deployment()
            and os.environ.get("SCHOOLSOFT_ONLINE_SYNC_ENABLED", "0").strip().lower() in {"1", "true", "yes", "on"}
        ),
        "can_backup": user_can_manage_users(user) and connection.vendor == "sqlite",
    }
    return render(request, "core/dashboard.html", context)


def _get_filtered_students(request):
    query = request.GET.get("q", "").strip()
    class_id = request.GET.get("class", "").strip()
    section_id = request.GET.get("section", "").strip()
    status = request.GET.get("status", "active").strip() or "active"

    students = Student.objects.select_related("current_class", "current_section").order_by(
        "current_class__display_order",
        "current_section__name",
        "roll_no",
        "full_name",
    )

    if status == "active":
        students = students.filter(is_active=True)
    elif status == "inactive":
        students = students.filter(is_active=False)

    if query:
        students = students.filter(
            Q(full_name__icontains=query)
            | Q(father_name__icontains=query)
            | Q(mother_name__icontains=query)
            | Q(admission_no__icontains=query)
            | Q(legacy_sid__icontains=query)
            | Q(mobile_primary__icontains=query)
        )

    if class_id:
        students = students.filter(current_class_id=class_id)

    if section_id:
        students = students.filter(current_section_id=section_id)

    return students, query, class_id, section_id, status


def student_list(request):
    students, query, class_id, section_id, status = _get_filtered_students(request)

    all_students = Student.objects.all()
    total_all_students = all_students.count()
    total_active_students = all_students.filter(is_active=True).count()
    total_inactive_students = total_all_students - total_active_students

    try:
        per_page = int(request.GET.get("per_page", 50))
    except ValueError:
        per_page = 50
        
    paginator = Paginator(students, per_page)
    page = paginator.get_page(request.GET.get("page"))

    context = {
        "page": page,
        "per_page": per_page,
        "total_students": students.count(),
        "query": query,
        "selected_class": class_id,
        "selected_section": section_id,
        "selected_status": status,
        "classes": SchoolClass.objects.order_by("display_order", "name"),
        "sections": Section.objects.select_related("school_class").order_by(
            "school_class__display_order",
            "name",
        ),
        "total_students": students.count(),
        "total_all_students": total_all_students,
        "total_active_students": total_active_students,
        "total_inactive_students": total_inactive_students,
    }
    return render(request, "core/student_list.html", context)


def student_detail(request, pk):
    student = get_object_or_404(
        Student.objects.select_related("current_class", "current_section"),
        pk=pk,
    )
    due_total = FeeReceipt.objects.filter(
        student=student, is_cancelled=False, carried_forward=False, legacy_due_amount__gt=0
    ).aggregate(total=Sum("legacy_due_amount"))["total"] or 0
    
    active_session = AcademicSession.objects.filter(is_active=True).order_by("-starts_on").first()
    concessions = StudentConcession.objects.filter(student=student, session=active_session).order_by("-id") if active_session else []
    
    school_profile = get_active_school_profile()
    return render(
        request,
        "core/student_detail.html",
        {
            "student": student,
            "due_total": due_total,
            "school_name": school_profile.name if school_profile else "",
            "concessions": concessions,
        },
    )


def school_profile_detail(request):
    profile = get_active_school_profile()
    all_profiles = SchoolProfile.objects.order_by("-is_active", "name")
    return render(
        request,
        "core/school_profile_detail.html",
        {
            "profile": profile,
            "all_profiles": all_profiles,
        },
    )


def fee_structure_report(request):
    selected_session_id = request.GET.get("session", "").strip()
    selected_class_id = request.GET.get("class", "").strip()
    row_mode = request.GET.get("rows", "nonzero").strip() or "nonzero"
    sessions = AcademicSession.objects.order_by("-is_active", "-starts_on", "name")
    selected_session = None

    if selected_session_id:
        selected_session = sessions.filter(pk=selected_session_id).first()

    if selected_session is None:
        selected_session = sessions.filter(is_active=True, fee_structures__isnull=False).distinct().first()

    if selected_session is None:
        selected_session = sessions.filter(fee_structures__isnull=False).distinct().first()

    structures = FeeStructure.objects.select_related("session", "school_class", "fee_head").order_by(
        "school_class__display_order",
        "school_class__name",
        "fee_head__name",
    )

    if selected_session:
        structures = structures.filter(session=selected_session)

    if selected_class_id:
        structures = structures.filter(school_class_id=selected_class_id)

    if row_mode != "all":
        structures = structures.filter(amount__gt=Decimal("0.00"))

    totals = structures.aggregate(
        total_amount=Sum("amount"),
        total_rows=Count("id"),
    )
    class_totals = (
        structures.values(
            "school_class_id",
            "school_class__name",
            "school_class__display_order",
        )
        .annotate(total_amount=Sum("amount"), fee_heads=Count("id"))
        .order_by("school_class__display_order", "school_class__name")
    )

    return render(
        request,
        "core/fee_structure_report.html",
        {
            "structures": structures,
            "class_totals": class_totals,
            "sessions": sessions,
            "classes": SchoolClass.objects.order_by("display_order", "name"),
            "selected_session": selected_session,
            "selected_session_id": str(selected_session.id) if selected_session else selected_session_id,
            "selected_class_id": selected_class_id,
            "row_mode": row_mode,
            "totals": totals,
        },
    )


def marks_report(request):
    selected_term_id = request.GET.get("term", "").strip()
    selected_class_id = request.GET.get("class", "").strip()
    query = request.GET.get("q", "").strip()

    terms = (
        ExamTerm.objects.select_related("session")
        .filter(tests__marks__isnull=False)
        .distinct()
        .order_by("-session__name", "display_order", "name")
    )
    selected_term = terms.filter(pk=selected_term_id).first() if selected_term_id else None
    if selected_term is None:
        selected_term = terms.first()

    marks = ExamMark.objects.select_related(
        "student",
        "student__current_class",
        "student__current_section",
        "exam_test",
        "exam_test__subject",
        "exam_test__term",
    )

    if selected_term:
        marks = marks.filter(exam_test__term=selected_term)
    else:
        marks = marks.none()

    if selected_class_id:
        marks = marks.filter(exam_test__school_class_id=selected_class_id)

    if query:
        marks = marks.filter(
            Q(student__full_name__icontains=query)
            | Q(student__father_name__icontains=query)
            | Q(student__legacy_sid__icontains=query)
            | Q(student__admission_no__icontains=query)
        )

    summary = marks.aggregate(
        mark_rows=Count("id"),
        subjects=Count("exam_test__subject", distinct=True),
        tests=Count("exam_test", distinct=True),
    )
    grouped_rows = marks.values(
        "student_id",
        "student__legacy_sid",
        "student__admission_no",
        "student__full_name",
        "student__father_name",
        "student__current_class__name",
        "student__current_class__display_order",
        "student__current_section__name",
    ).annotate(
        subjects=Count("id"),
        absent=Count("id", filter=Q(is_absent=True)),
        total_obtained=Sum("marks_obtained"),
        total_max=Sum("exam_test__max_marks"),
    ).order_by(
        "student__current_class__display_order",
        "student__current_section__name",
        "student__full_name",
    )

    rows = []
    for row in grouped_rows:
        total_obtained = row["total_obtained"] or Decimal("0.00")
        total_max = row["total_max"] or Decimal("0.00")
        percentage = None
        if total_max:
            percentage = (total_obtained / total_max * Decimal("100")).quantize(Decimal("0.01"))

        rows.append(
            {
                **row,
                "total_obtained": total_obtained,
                "total_max": total_max,
                "percentage": percentage,
            }
        )

    paginator = Paginator(rows, 50)
    page = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "core/marks_report.html",
        {
            "page": page,
            "terms": terms,
            "classes": SchoolClass.objects.order_by("display_order", "name"),
            "selected_term": selected_term,
            "selected_term_id": str(selected_term.id) if selected_term else selected_term_id,
            "selected_class_id": selected_class_id,
            "query": query,
            "summary": summary,
            "total_students": len(rows),
        },
    )


def receipt_list(request):
    query = request.GET.get("q", "").strip()
    class_id = request.GET.get("class", "").strip()
    payment_mode = request.GET.get("payment_mode", "").strip()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()
    session_id = request.GET.get("session", "").strip()
    
    active_session = AcademicSession.objects.filter(is_active=True).first()
    if not session_id and not query:
        session_id = str(active_session.id) if active_session else ""
        
    receipts = FeeReceipt.objects.select_related(
        "student",
        "student__current_class",
        "student__current_section",
        "session",
    ).order_by("-receipt_date", "-legacy_receipt_no", "-id")

    if session_id and session_id != "all":
        receipts = receipts.filter(session_id=session_id)

    if query:
        receipts = receipts.filter(
            Q(receipt_no__icontains=query)
            | Q(legacy_receipt_no__icontains=query)
            | Q(student__full_name__icontains=query)
            | Q(student__legacy_sid__icontains=query)
            | Q(student__admission_no__icontains=query)
        )

    if class_id:
        receipts = receipts.filter(student__current_class_id=class_id)

    if payment_mode:
        receipts = receipts.filter(payment_mode=payment_mode)

    if date_from:
        receipts = receipts.filter(receipt_date__gte=date_from)

    if date_to:
        receipts = receipts.filter(receipt_date__lte=date_to)

    totals = receipts.filter(is_cancelled=False).aggregate(
        received=Sum("received_amount"),
        net=Sum("legacy_net_total"),
        # carried_forward receipts stay visible in the register (so staff can
        # see the history) but their due is now represented on whatever newer
        # receipt absorbed it via Previous Due - excluded here to avoid the
        # same balance being counted twice in this total.
        due=Sum("legacy_due_amount", filter=Q(carried_forward=False)),
    )
    cancelled_count = receipts.filter(is_cancelled=True).count()
    paginator = Paginator(receipts, 50)
    page = paginator.get_page(request.GET.get("page"))

    context = {
        "page": page,
        "query": query,
        "selected_class": class_id,
        "selected_payment_mode": payment_mode,
        "selected_session": session_id,
        "date_from": date_from,
        "date_to": date_to,
        "classes": SchoolClass.objects.order_by("display_order", "name"),
        "sessions": AcademicSession.objects.order_by("-starts_on", "name"),
        "payment_modes": FeeReceipt.PaymentMode.choices,
        "totals": totals,
        "total_receipts": receipts.count(),
        "cancelled_count": cancelled_count,
    }
    return render(request, "core/receipt_list.html", context)



_DUE_RESULT_MONEY_FIELDS = (
    "scheduled_fee_demand",
    "transport_demand",
    "opening_balance_amount",
    "late_fee_amount",
    "gross_demand",
    "received_amount",
    "concession_amount",
    "due_amount",
    "credit_amount",
)
_ACADEMIC_MONTH_NUMBERS = {
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
}


def _default_due_month(session):
    today = timezone.localdate()
    if session and session.starts_on and today < session.starts_on:
        return ACADEMIC_MONTHS[0]
    if session and session.ends_on and today > session.ends_on:
        return ACADEMIC_MONTHS[-1]
    return next(
        (month for month, number in _ACADEMIC_MONTH_NUMBERS.items() if number == today.month),
        ACADEMIC_MONTHS[0],
    )


def _due_month_label(session, through_month):
    month_number = _ACADEMIC_MONTH_NUMBERS.get(through_month)
    if not month_number:
        return through_month
    if session and session.starts_on and session.ends_on:
        start_month = session.starts_on.replace(day=1)
        end_month = session.ends_on.replace(day=1)
        for year in range(session.starts_on.year, session.ends_on.year + 1):
            candidate = date(year, month_number, 1)
            if start_month <= candidate <= end_month:
                return candidate.strftime("%B %Y")
    return through_month


def _due_up_to_month_data(request):
    sessions = AcademicSession.objects.order_by("-starts_on", "name")
    active_session = sessions.filter(is_active=True).first() or sessions.first()
    requested_session = request.GET.get("session", "").strip()
    filter_errors = []

    session = active_session
    if requested_session:
        session = sessions.filter(pk=requested_session).first() if requested_session.isdigit() else None
        if session is None:
            filter_errors.append("Selected academic session was not found.")

    selected_month = request.GET.get("month", "").strip().upper() or _default_due_month(session)
    if selected_month not in ACADEMIC_MONTHS:
        filter_errors.append("Selected target month is invalid; the default month is shown instead.")
        selected_month = _default_due_month(session)

    selected_class = request.GET.get("class", "").strip()
    selected_section = request.GET.get("section", "").strip()
    requested_status = request.GET.get("status", "active").strip() or "active"
    selected_status = "active"
    selected_balance = request.GET.get("balance", "all").strip() or "all"
    query = request.GET.get("q", "").strip()

    if requested_status != "active":
        filter_errors.append(
            "Due report is restricted to active students; archived records were excluded."
        )
    if selected_balance not in {"all", "due", "settled", "credit"}:
        filter_errors.append("Selected balance filter is invalid; All Balances is being used.")
        selected_balance = "all"

    invalid_id_filter = False
    if selected_class and not selected_class.isdigit():
        filter_errors.append("Selected class is invalid.")
        invalid_id_filter = True
    if selected_section and not selected_section.isdigit():
        filter_errors.append("Selected section is invalid.")
        invalid_id_filter = True

    students = Student.objects.filter(is_active=True).select_related(
        "current_class", "current_section"
    ).order_by(
        "current_class__display_order",
        "current_section__name",
        "roll_no",
        "full_name",
    )
    if selected_class and not invalid_id_filter:
        students = students.filter(current_class_id=selected_class)
    if selected_section and not invalid_id_filter:
        students = students.filter(current_section_id=selected_section)
    if query:
        students = students.filter(
            Q(full_name__icontains=query)
            | Q(father_name__icontains=query)
            | Q(admission_no__icontains=query)
            | Q(legacy_sid__icontains=query)
            | Q(mobile_primary__icontains=query)
        )

    rows = []
    skipped = []
    calculated_count = 0
    totals = {field: Decimal("0.00") for field in _DUE_RESULT_MONEY_FIELDS}
    candidates = [] if session is None or invalid_id_filter else list(students)

    for student in candidates:
        try:
            result = calculate_student_due(
                student=student,
                session=session,
                through_month=selected_month,
            )
        except ValidationError as exc:
            skipped.append({"student": student, "reason": "; ".join(exc.messages)})
            continue

        calculated_count += 1
        if selected_balance == "due" and result.due_amount <= Decimal("0.00"):
            continue
        if selected_balance == "settled" and (
            result.due_amount > Decimal("0.00") or result.credit_amount > Decimal("0.00")
        ):
            continue
        if selected_balance == "credit" and result.credit_amount <= Decimal("0.00"):
            continue

        class_label = student.current_class.name if student.current_class else ""
        if student.current_section:
            class_label = f"{class_label}-{student.current_section.name}" if class_label else student.current_section.name
        rows.append(
            {
                "student": student,
                "result": result,
                "class_label": class_label,
                "student_type": "New" if result.is_new_student else "Old",
            }
        )
        for field in _DUE_RESULT_MONEY_FIELDS:
            totals[field] += getattr(result, field)

    totals["students"] = len(rows)
    totals["due_students"] = sum(1 for row in rows if row["result"].due_amount > Decimal("0.00"))
    totals["credit_students"] = sum(1 for row in rows if row["result"].credit_amount > Decimal("0.00"))

    export_params = {
        "session": session.pk if session else "",
        "class": selected_class,
        "section": selected_section,
        "month": selected_month,
        "status": selected_status,
        "balance": selected_balance,
        "q": query,
    }
    export_query = urlencode({key: value for key, value in export_params.items() if value != ""})

    return {
        "rows": rows,
        "totals": totals,
        "query": query,
        "sessions": sessions,
        "classes": SchoolClass.objects.order_by("display_order", "name"),
        "sections": Section.objects.select_related("school_class").order_by(
            "school_class__display_order", "name"
        ),
        "months": ACADEMIC_MONTHS,
        "selected_session": str(session.pk) if session else requested_session,
        "selected_session_object": session,
        "selected_class": selected_class,
        "selected_section": selected_section,
        "selected_month": selected_month,
        "selected_status": selected_status,
        "selected_balance": selected_balance,
        "target_label": _due_month_label(session, selected_month),
        "filter_errors": filter_errors,
        "skipped": skipped,
        "candidate_count": len(candidates),
        "calculated_count": calculated_count,
        "total_students": len(rows),
        "export_query": export_query,
        "school_name": (get_active_school_profile().name if get_active_school_profile() else ""),
    }



def defaulter_list(request):
    sessions = AcademicSession.objects.order_by("-starts_on", "name")
    active_session = sessions.filter(is_active=True).first() or sessions.first()
    if not active_session:
        return render(request, "core/defaulter_list.html", {"error": "No active academic session found."})

    default_m = _default_due_month(active_session)
    default_idx = ACADEMIC_MONTHS.index(default_m) if default_m in ACADEMIC_MONTHS else 0
    available_months = ACADEMIC_MONTHS[:default_idx + 1]

    selected_month = request.GET.get("month", "").strip().upper() or default_m
    if selected_month not in available_months:
        selected_month = default_m

    selected_class = request.GET.get("class", "").strip()

    invalid_id_filter = False
    if selected_class and not selected_class.isdigit():
        invalid_id_filter = True

    students = Student.objects.filter(is_active=True).select_related(
        "current_class", "current_section"
    ).order_by(
        "current_class__display_order",
        "current_section__name",
        "roll_no",
        "full_name",
    )
    if selected_class and not invalid_id_filter:
        students = students.filter(current_class_id=selected_class)

    candidates = [] if invalid_id_filter else list(students)

    # Current paid excludes carried-forward imports because opening balances are
    # already net of those old receipts. Keep the latest historical receipt
    # visible as an information-only last payment so old deposits do not look
    # missing in the defaulters list.
    candidate_ids = [s.id for s in candidates]
    last_receipt_map = {}
    last_receipts = FeeReceipt.objects.filter(
        student_id__in=candidate_ids,
        is_cancelled=False,
        session=active_session,
    ).order_by("student_id", "-receipt_date", "-id").only(
        "student_id",
        "receipt_date",
        "received_amount",
        "carried_forward",
    )
    for receipt in last_receipts:
        last_receipt_map.setdefault(receipt.student_id, receipt)

    ytd_payments = FeeReceipt.objects.filter(
        student_id__in=candidate_ids,
        is_cancelled=False,
        session=active_session,
    ).values("student_id").annotate(total_paid=Sum("received_amount"))
    ytd_payment_map = {item["student_id"]: item["total_paid"] for item in ytd_payments}

    defaulters = []
    total_due = Decimal("0.00")
    
    target_idx = available_months.index(selected_month) if selected_month in available_months else 0

    for student in candidates:
        try:
            result = calculate_student_due(
                student=student,
                session=active_session,
                through_month=selected_month,
            )
        except ValidationError:
            continue

        if result.due_amount > Decimal("0.00"):
            class_name = student.current_class.name if student.current_class else "Unknown"
            if student.current_section:
                class_name = f"{class_name}-{student.current_section.name}"

            defaulters.append({
                "student": student,
                "class_name": class_name,
                "due_amount": result.due_amount,
                "opening_due": result.opening_balance_amount,
                "ytd_paid": ytd_payment_map.get(student.id) or Decimal("0.00"),
                "last_payment_date": (
                    last_receipt_map[student.id].receipt_date if student.id in last_receipt_map else None
                ),
                "last_payment_amount": (
                    last_receipt_map[student.id].received_amount if student.id in last_receipt_map else Decimal("0.00")
                ),
                "last_payment_is_legacy": (
                    last_receipt_map[student.id].carried_forward if student.id in last_receipt_map else False
                ),
            })
            total_due += result.due_amount

    # Grouping
    grouped_defaulters = {}
    for d in defaulters:
        cname = d["class_name"]
        if cname not in grouped_defaulters:
            grouped_defaulters[cname] = []
        grouped_defaulters[cname].append(d)

    context = {
        "school_name": get_active_school_profile().name if get_active_school_profile() else "",
        "classes": SchoolClass.objects.order_by("display_order", "name"),
        "available_months": available_months,
        "selected_month": selected_month,
        "selected_class": selected_class,
        "defaulters": defaulters,
        "grouped_defaulters": grouped_defaulters,
        "total_defaulters": len(defaulters),
        "total_due": total_due,
        "target_label": _due_month_label(active_session, selected_month),
    }
    return render(request, "core/defaulter_list.html", context)


def due_up_to_month_report(request):
    context = _due_up_to_month_data(request)
    context["page"] = Paginator(context["rows"], 100).get_page(request.GET.get("page"))
    return render(request, "core/due_up_to_month_report.html", context)


def due_up_to_month_report_pdf(request):
    context = _due_up_to_month_data(request)
    pdf_bytes = build_due_up_to_month_report_pdf(
        context["rows"],
        context["totals"],
        get_active_school_profile(),
        context["selected_session_object"],
        context["selected_month"],
        context["target_label"],
    )
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="due-up-to-{context["selected_month"].lower()}.pdf"'
    )
    return response


def due_slip_pdf(request):
    context = _due_up_to_month_data(request)
    pdf_bytes = build_due_slip_pdf(
        context["rows"],
        get_active_school_profile(),
        context["selected_session_object"],
        context["selected_month"],
        context["target_label"],
    )
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="due-slips-{context["selected_month"].lower()}.pdf"'
    )
    return response


def due_report(request):
    rows, totals = get_due_report_rows(request)
    paginator = Paginator(rows, 50)
    page = paginator.get_page(request.GET.get("page"))
    school_profile = get_active_school_profile()

    return render(
        request,
        "core/due_report.html",
        {
            "page": page,
            "query": request.GET.get("q", "").strip(),
            "selected_class": request.GET.get("class", "").strip(),
            "selected_session": getattr(request, "_due_selected_session", ""),
            "selected_status": getattr(request, "_due_selected_status", "active"),
            "classes": SchoolClass.objects.order_by("display_order", "name"),
            "sessions": AcademicSession.objects.order_by("-starts_on", "name"),
            "total_students": rows.count(),
            "totals": totals,
            "school_name": school_profile.name if school_profile else "",
        },
    )


def due_report_pdf(request):
    rows, totals = get_due_report_rows(request)
    pdf_bytes = build_due_report_pdf(list(rows), totals, get_active_school_profile())
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="due-report.pdf"'
    return response


def get_due_report_rows(request):
    query = request.GET.get("q", "").strip()
    class_id = request.GET.get("class", "").strip()
    session_id = request.GET.get("session", "").strip()
    status = request.GET.get("status", "active").strip()
    
    active_session = AcademicSession.objects.filter(is_active=True).first()
    if not session_id and not query:
        session_id = str(active_session.id) if active_session else ""
        
    request._due_selected_session = session_id
    request._due_selected_status = status

    receipts = FeeReceipt.objects.select_related(
        "student",
        "student__current_class",
        "student__current_section",
    ).filter(legacy_due_amount__gt=0, is_cancelled=False, carried_forward=False)

    if session_id and session_id != "all":
        receipts = receipts.filter(session_id=session_id)
        
    if status == "active":
        receipts = receipts.filter(student__is_active=True)
    elif status == "inactive":
        receipts = receipts.filter(student__is_active=False)

    if query:
        receipts = receipts.filter(
            Q(student__full_name__icontains=query)
            | Q(student__father_name__icontains=query)
            | Q(student__legacy_sid__icontains=query)
            | Q(student__admission_no__icontains=query)
        )

    if class_id:
        receipts = receipts.filter(student__current_class_id=class_id)

    rows = receipts.values(
        "student_id",
        "student__legacy_sid",
        "student__full_name",
        "student__father_name",
        "student__current_class__name",
        "student__current_class__display_order",
        "student__current_section__name",
        "student__mobile_primary",
        "student__admission_no",
    ).annotate(
        due_amount=Sum("legacy_due_amount"),
        paid_amount=Sum("received_amount"),
        net_amount=Sum("legacy_net_total"),
    ).order_by(
        "student__current_class__display_order",
        "student__current_section__name",
        "student__full_name",
    )

    totals = rows.aggregate(
        due=Sum("due_amount"),
        paid=Sum("paid_amount"),
        net=Sum("net_amount"),
    )
    return rows, totals


def receipt_create(request):
    recent_receipts = FeeReceipt.objects.select_related(
        "student",
        "student__current_class",
        "student__current_section",
    ).order_by("-receipt_date", "-id")[:8]
    today_totals = FeeReceipt.objects.filter(receipt_date=timezone.localdate()).aggregate(
        received=Sum("received_amount"),
        due=Sum("legacy_due_amount"),
        count=Count("id"),
    )

    if request.method == "POST":
        receipt_form = FeeReceiptEntryForm(request.POST)
        line_form = FeeReceiptLineEntryForm(request.POST)

        if receipt_form.is_valid() and line_form.is_valid():
            line_total = line_form.cleaned_data["line_total"]
            with transaction.atomic():
                receipt = receipt_form.save(commit=False)
                receipt.receipt_no = next_manual_receipt_no()
                receipt.legacy_fee_total = line_total
                receipt.legacy_net_total = (
                    line_total + receipt.late_fee_amount - receipt.concession_amount
                )
                receipt.legacy_due_amount = max(
                    receipt.legacy_net_total - receipt.received_amount,
                    Decimal("0.00"),
                )
                apply_receipt_student_snapshot(receipt)
                receipt.save()

                for fee_head, amount in line_form.amounts():
                    receipt.lines.create(fee_head=fee_head, amount=amount)

            if request.POST.get("action") == "save_print":
                return redirect(f"{reverse('core:receipt_detail', kwargs={'pk': receipt.pk})}?autoprint=1")
            return redirect("core:receipt_detail", pk=receipt.pk)
    else:
        student_id = request.GET.get("student")
        initial_data = {}
        if student_id:
            initial_data["student"] = student_id
        receipt_form = FeeReceiptEntryForm(initial=initial_data)
        line_form = FeeReceiptLineEntryForm()

    return render(
        request,
        "core/receipt_form.html",
        {
            "receipt_form": receipt_form,
            "line_form": line_form,
            "recent_receipts": recent_receipts,
            "today_totals": today_totals,
            "default_due_month": _default_due_month(AcademicSession.objects.filter(is_active=True).first()),
        },
    )


def receipt_detail(request, pk):
    receipt = get_object_or_404(
        FeeReceipt.objects.select_related(
            "student",
            "student__current_class",
            "student__current_section",
            "session",
        ).prefetch_related("lines", "lines__fee_head"),
        pk=pk,
    )
    return render(
        request,
        "core/receipt_detail.html",
        {
            "receipt": receipt,
            "school_profile": get_active_school_profile(),
            "receipt_due_status": _receipt_due_status(receipt),
        },
    )


def receipt_cancel(request, pk):
    receipt = get_object_or_404(
        FeeReceipt.objects.select_related("student", "session"),
        pk=pk,
    )

    if receipt.is_cancelled:
        messages.info(request, f"Receipt {receipt.receipt_no} is already cancelled.")
        return redirect("core:receipt_detail", pk=receipt.pk)

    if request.method == "POST":
        reason = request.POST.get("cancel_reason", "").strip()
        if not reason:
            messages.error(request, "A cancellation reason is required.")
            return render(request, "core/receipt_cancel_confirm.html", {"receipt": receipt})

        receipt.is_cancelled = True
        receipt.cancelled_at = timezone.now()
        receipt.cancelled_by = request.user if request.user.is_authenticated else None
        receipt.cancel_reason = reason
        receipt.save(update_fields=["is_cancelled", "cancelled_at", "cancelled_by", "cancel_reason"])
        messages.success(request, f"Receipt {receipt.receipt_no} has been cancelled/voided.")
        return redirect("core:receipt_detail", pk=receipt.pk)

    return render(request, "core/receipt_cancel_confirm.html", {"receipt": receipt})


def receipt_edit(request, pk):
    receipt = get_object_or_404(
        FeeReceipt.objects.select_related("student", "session").prefetch_related("lines", "lines__fee_head"),
        pk=pk,
    )

    if receipt.is_cancelled:
        messages.error(request, f"Receipt {receipt.receipt_no} is cancelled and cannot be edited.")
        return redirect("core:receipt_detail", pk=receipt.pk)

    if request.method == "POST":
        receipt_form = FeeReceiptEditForm(request.POST, instance=receipt)
        line_form = FeeReceiptLineEntryForm(
            request.POST,
            require_positive_total=receipt.previous_due_amount <= Decimal("0.00"),
        )

        if receipt_form.is_valid() and line_form.is_valid():
            line_total = line_form.cleaned_data["line_total"]
            if line_total <= Decimal("0.00") and receipt.previous_due_amount <= Decimal("0.00"):
                line_form.add_error(
                    None, "Enter at least one fee head amount before saving."
                )
                return render(
                    request,
                    "core/receipt_edit.html",
                    {"receipt_form": receipt_form, "line_form": line_form, "receipt": receipt},
                )
            with transaction.atomic():
                original_receipt = FeeReceipt.objects.get(pk=receipt.pk)
                # Capture before snapshot
                before_snapshot = {
                    "receipt_no": original_receipt.receipt_no,
                    "student": original_receipt.display_student_name,
                    "father_name": original_receipt.display_father_name,
                    "class": original_receipt.display_class_section,
                    "session": original_receipt.session.name,
                    "receipt_date": str(original_receipt.receipt_date),
                    "from_month": original_receipt.from_month,
                    "to_month": original_receipt.to_month,
                    "payment_mode": original_receipt.payment_mode,
                    "previous_due_amount": str(original_receipt.previous_due_amount),
                    "concession_amount": str(original_receipt.concession_amount),
                    "late_fee_amount": str(original_receipt.late_fee_amount),
                    "received_amount": str(original_receipt.received_amount),
                    "legacy_fee_total": str(original_receipt.legacy_fee_total),
                    "legacy_net_total": str(original_receipt.legacy_net_total),
                    "legacy_due_amount": str(original_receipt.legacy_due_amount),
                    "remarks": original_receipt.remarks,
                    "lines": {line.fee_head.name: str(line.amount) for line in original_receipt.lines.all()}
                }

                updated_receipt = receipt_form.save(commit=False)
                
                # Delete old lines and create new ones
                updated_receipt.lines.all().delete()
                
                new_lines = []
                for fee_head, amount in line_form.amounts():
                    new_lines.append(updated_receipt.lines.create(fee_head=fee_head, amount=amount))
                
                # Recalculate totals
                line_total = line_form.cleaned_data["line_total"]
                updated_receipt.legacy_fee_total = line_total
                updated_receipt.legacy_net_total = (
                    line_total + updated_receipt.previous_due_amount
                    + updated_receipt.late_fee_amount - updated_receipt.concession_amount
                )
                updated_receipt.legacy_due_amount = max(
                    updated_receipt.legacy_net_total - updated_receipt.received_amount,
                    Decimal("0.00"),
                )
                apply_receipt_student_snapshot(updated_receipt)
                
                # Update audit fields
                updated_receipt.is_edited = True
                updated_receipt.edited_at = timezone.now()
                updated_receipt.edited_by = request.user if request.user.is_authenticated else None
                updated_receipt.edit_count += 1
                updated_receipt.save()

                # Capture after snapshot
                after_snapshot = {
                    "receipt_no": updated_receipt.receipt_no,
                    "student": updated_receipt.display_student_name,
                    "father_name": updated_receipt.display_father_name,
                    "class": updated_receipt.display_class_section,
                    "session": updated_receipt.session.name,
                    "receipt_date": str(updated_receipt.receipt_date),
                    "from_month": updated_receipt.from_month,
                    "to_month": updated_receipt.to_month,
                    "payment_mode": updated_receipt.payment_mode,
                    "previous_due_amount": str(updated_receipt.previous_due_amount),
                    "concession_amount": str(updated_receipt.concession_amount),
                    "late_fee_amount": str(updated_receipt.late_fee_amount),
                    "received_amount": str(updated_receipt.received_amount),
                    "legacy_fee_total": str(updated_receipt.legacy_fee_total),
                    "legacy_net_total": str(updated_receipt.legacy_net_total),
                    "legacy_due_amount": str(updated_receipt.legacy_due_amount),
                    "remarks": updated_receipt.remarks,
                    "lines": {line.fee_head.name: str(line.amount) for line in new_lines}
                }
                
                # Generate compact changes JSON
                changes = {}
                for key in before_snapshot:
                    if key != "lines":
                        if before_snapshot[key] != after_snapshot[key]:
                            changes[key] = {"before": before_snapshot[key], "after": after_snapshot[key]}
                
                lines_changes = {}
                all_heads = set(before_snapshot["lines"].keys()) | set(after_snapshot["lines"].keys())
                for head in all_heads:
                    b_amt = before_snapshot["lines"].get(head, "0.00")
                    a_amt = after_snapshot["lines"].get(head, "0.00")
                    if b_amt != a_amt:
                        lines_changes[head] = {"before": b_amt, "after": a_amt}
                
                if lines_changes:
                    changes["lines"] = lines_changes

                FeeReceiptAuditLog.objects.create(
                    receipt=updated_receipt,
                    action=FeeReceiptAuditLog.ActionChoices.EDITED,
                    changed_by=request.user if request.user.is_authenticated else None,
                    reason=updated_receipt.edit_reason,
                    before_snapshot=before_snapshot,
                    after_snapshot=after_snapshot,
                    changes=changes
                )

            messages.success(request, f"Receipt {receipt.receipt_no} has been corrected.")
            return redirect("core:receipt_detail", pk=receipt.pk)
    else:
        receipt_form = FeeReceiptEditForm(instance=receipt)
        
        # Populate lines
        initial_lines = {}
        for line in receipt.lines.all():
            initial_lines[f"fee_head_{line.fee_head_id}"] = line.amount
            
        line_form = FeeReceiptLineEntryForm(initial=initial_lines)

    return render(
        request,
        "core/receipt_edit.html",
        {
            "receipt": receipt,
            "receipt_form": receipt_form,
            "line_form": line_form,
        },
    )


def receipt_pdf(request, pk):
    receipt = get_object_or_404(
        FeeReceipt.objects.select_related(
            "student",
            "student__current_class",
            "student__current_section",
            "session",
        ).prefetch_related("lines", "lines__fee_head"),
        pk=pk,
    )
    pdf_bytes = build_fee_receipt_pdf(receipt, get_active_school_profile())
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{receipt.receipt_no}.pdf"'
    return response


def receipt_pdf_2up(request, pk):
    receipt = get_object_or_404(
        FeeReceipt.objects.select_related(
            "student",
            "student__current_class",
            "student__current_section",
            "session",
        ).prefetch_related("lines", "lines__fee_head"),
        pk=pk,
    )
    pdf_bytes = build_fee_receipt_pdf_2up(receipt, get_active_school_profile())
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="receipt_{receipt.receipt_no}_2up.pdf"'
    return response


def _local_print_allowed(request):
    host = request.get_host().split(":", 1)[0].strip("[]").lower()
    return os.name == "nt" and (host == "localhost" or host == "::1" or host.startswith("127."))


def _edge_executable_path():
    candidates = [
        Path(os.environ.get("PROGRAMFILES(X86)", "")) / "Microsoft" / "Edge" / "Application" / "msedge.exe",
        Path(os.environ.get("PROGRAMFILES", "")) / "Microsoft" / "Edge" / "Application" / "msedge.exe",
        Path(os.environ.get("LOCALAPPDATA", "")) / "Microsoft" / "Edge" / "Application" / "msedge.exe",
    ]
    for candidate in candidates:
        if str(candidate) and candidate.exists():
            return candidate
    return None


def _send_pdf_to_printer(pdf_path):
    try:
        os.startfile(str(pdf_path), "print")
        return "default"
    except OSError:
        edge_path = _edge_executable_path()
        if not edge_path:
            raise
        subprocess.Popen(
            [
                str(edge_path),
                "--kiosk-printing",
                "--new-window",
                pdf_path.as_uri(),
            ],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        return "edge"


def receipt_print(request, pk):
    receipt = get_object_or_404(
        FeeReceipt.objects.select_related(
            "student",
            "student__current_class",
            "student__current_section",
            "session",
        ).prefetch_related("lines", "lines__fee_head"),
        pk=pk,
    )
    if not _local_print_allowed(request):
        return redirect("core:receipt_pdf", pk=receipt.pk)

    pdf_bytes = build_fee_receipt_pdf(receipt, get_active_school_profile())
    safe_receipt_no = "".join(
        char if char.isalnum() or char in ("-", "_") else "_"
        for char in receipt.receipt_no
    ) or f"receipt-{receipt.pk}"
    print_dir = Path(os.environ.get("LOCALAPPDATA") or tempfile.gettempdir()) / APP_DATA_DIR_NAME / "print-jobs"
    print_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = print_dir / f"{safe_receipt_no}-{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    pdf_path.write_bytes(pdf_bytes)

    try:
        print_method = _send_pdf_to_printer(pdf_path)
    except OSError as exc:
        messages.error(request, f"Print start nahi ho paya: {exc}")
    else:
        if print_method == "edge":
            messages.success(request, "A5 landscape receipt Edge print mode me bhej diya gaya.")
        else:
            messages.success(request, "A5 landscape receipt print command bhej diya gaya.")
    return redirect("core:receipt_detail", pk=receipt.pk)


def student_fee_defaults(request, pk):
    student = get_object_or_404(Student.objects.select_related("current_class", "current_section"), pk=pk)
    session_id = request.GET.get("session")
    active_session = AcademicSession.objects.filter(is_active=True).order_by("-starts_on").first()
    selected_session = active_session
    if session_id:
        selected_session = AcademicSession.objects.filter(pk=session_id).first() or active_session

    structures = FeeStructure.objects.select_related("fee_head").filter(
        school_class=student.current_class,
        is_active=True,
        fee_head__is_active=True,
        fee_head__is_transport=False,
    )
    if selected_session:
        structures = structures.filter(session=selected_session)
    else:
        structures = structures.none()
    if student_is_zero_fee(student) or student_uses_fee_package(student):
        structures = structures.none()

    due_status = None
    target_month = _normalise_academic_month(request.GET.get("month")) or _default_due_month(selected_session)
    if target_month not in ACADEMIC_MONTHS:
        target_month = _default_due_month(selected_session)
    from_month = _normalise_academic_month(request.GET.get("from_month")) or target_month
    if from_month not in ACADEMIC_MONTHS:
        from_month = target_month

    amounts = {}
    if selected_session:
        for structure in structures:
            try:
                amount = calculate_structure_receipt_amount(
                    structure=structure,
                    student=student,
                    session=selected_session,
                    from_month=from_month,
                    to_month=target_month,
                )
            except ValidationError:
                continue
            if amount > Decimal("0.00"):
                amounts[f"fee_head_{structure.fee_head_id}"] = str(amount)

        if student_uses_fee_package(student):
            package_head = FeeHead.objects.filter(name=PACKAGE_FEE_HEAD_NAME, is_active=True).first()
            package_amount = package_receipt_default_amount(student=student, session=selected_session)
            if package_head and package_amount > Decimal("0.00"):
                amounts[f"fee_head_{package_head.id}"] = str(package_amount)

    active_concession = None
    if selected_session:
        concession = student.concessions.filter(session=selected_session, is_active=True).first()
        if concession:
            if concession.amount_type == 'percent':
                amount_display = f"{concession.amount}%"
            elif concession.amount_type == 'fixed':
                amount_display = f"₹{concession.amount}/month"
            else:
                amount_display = "100% Free"
            from_lbl = concession.from_month or "APR"
            to_lbl = concession.to_month or "MAR"
            month_range = f"{from_lbl} → {to_lbl}" if (concession.from_month or concession.to_month) else "Full session"
            active_concession = {
                "type": concession.get_concession_type_display(),
                "amount": amount_display,
                "month_range": month_range,
                "reason": concession.reason,
                "approved_by": concession.approved_by_name,
            }

    balance_head = FeeHead.objects.filter(name="Balance Fee", is_active=True).first()

    if selected_session:
        try:
            due_status = _student_due_status_payload(
                student=student,
                session=selected_session,
                target_month=target_month,
            )
        except (ValidationError, AttributeError) as exc:
            due_status = {
                "available": False,
                "target_month": target_month,
                "error": str(exc),
            }

    return JsonResponse(
        {
            "student": student.full_name,
            "class": student.current_class.name if student.current_class else "",
            "section": student.current_section.name if student.current_section else "",
            "student_identity": [
                part
                for part in [
                    f"SID {student.legacy_sid}" if student.legacy_sid else "",
                    "-".join(
                        part
                        for part in [
                            student.current_class.name if student.current_class else "",
                            student.current_section.name if student.current_section else "",
                        ]
                        if part
                    ),
                    f"Adm {student.admission_no}" if student.admission_no else "",
                    f"Father {student.father_name}" if student.father_name else "",
                    f"Mobile {student.mobile_primary}" if student.mobile_primary else "",
                ]
                if part
            ],
            "amounts": amounts,
            "balance_fee_field": f"fee_head_{balance_head.id}" if balance_head else "",
            "due_status": due_status,
            "active_concession": active_concession,
        }
    )

def next_manual_receipt_no():
    timestamp = timezone.localtime().strftime("%Y%m%d%H%M%S")
    base = f"MR-{timestamp}"
    receipt_no = base
    suffix = 1

    while FeeReceipt.objects.filter(receipt_no=receipt_no).exists():
        suffix += 1
        receipt_no = f"{base}-{suffix}"

    return receipt_no

def get_collection_report_rows(request):
    from datetime import datetime, time
    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')
    
    receipts = FeeReceipt.objects.select_related(
        'student',
        'student__current_class',
        'student__current_section',
    ).filter(is_cancelled=False).order_by('receipt_date', 'receipt_no')

    date_from = None
    date_to = None
    
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            receipts = receipts.filter(receipt_date__gte=date_from)
        except ValueError:
            pass
            
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            receipts = receipts.filter(receipt_date__lte=date_to)
        except ValueError:
            pass

    if not date_from_str and not date_to_str:
        # Default to today if no dates provided
        today = timezone.localdate()
        date_from_str = today.strftime('%Y-%m-%d')
        date_to_str = date_from_str
        receipts = receipts.filter(receipt_date=today)
        
    totals = receipts.aggregate(
        total_received=Sum('received_amount'),
        total_concession=Sum('concession_amount'),
        total_legacy_net=Sum('legacy_net_total')
    )
    
    return receipts, totals, date_from_str, date_to_str


def collection_report(request):
    receipts, totals, date_from_str, date_to_str = get_collection_report_rows(request)
    
    return render(request, 'core/collection_report.html', {
        'receipts': receipts,
        'totals': totals,
        'date_from_str': date_from_str,
        'date_to_str': date_to_str,
        'selected_session': getattr(request, "_col_selected_session", ""),
        'sessions': AcademicSession.objects.order_by("-starts_on", "name"),
        'total_received': totals['total_received'] or Decimal('0.00'),
    })


def collection_report_pdf(request):
    from .pdf import build_collection_report_pdf
    receipts, totals, date_from_str, date_to_str = get_collection_report_rows(request)
    pdf_bytes = build_collection_report_pdf(list(receipts), totals, date_from_str, date_to_str, get_active_school_profile())
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="collection-report.pdf"'
    return response


def admission_form_pdf(request, pk):
    student = get_object_or_404(Student.objects.select_related("current_class", "current_section"), pk=pk)
    pdf_bytes = build_admission_form_pdf(student, get_active_school_profile())
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="admission-form-{student.admission_no or student.pk}.pdf"'
    return response


def character_certificate_pdf(request, pk):
    student = get_object_or_404(Student.objects.select_related("current_class", "current_section"), pk=pk)
    pdf_bytes = build_character_certificate_pdf(student, get_active_school_profile())
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="character-certificate-{student.admission_no or student.pk}.pdf"'
    return response


def id_card_pdf(request, pk):
    student = get_object_or_404(Student.objects.select_related("current_class", "current_section", "house"), pk=pk)
    pdf_bytes = build_id_card_pdf(student, get_active_school_profile())
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="id-card-{student.admission_no or student.pk}.pdf"'
    return response


def id_card_batch_pdf(request):
    students, query, class_id, section_id, status = _get_filtered_students(request)
    students = students.select_related("current_class", "current_section", "house")
    pdf_bytes = build_id_card_batch_pdf(list(students), get_active_school_profile())
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="id-cards.pdf"'
    return response


def staff_id_card_pdf(request, pk):
    staff = get_object_or_404(Staff, pk=pk)
    pdf_bytes = build_staff_id_card_pdf(staff, get_active_school_profile())
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="id-card-staff-{staff.legacy_emp_code or staff.pk}.pdf"'
    return response


def staff_id_card_batch_pdf(request):
    staff_qs = Staff.objects.filter(is_active=True).order_by("full_name")
    pdf_bytes = build_staff_id_card_batch_pdf(list(staff_qs), get_active_school_profile())
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="id-cards-staff.pdf"'
    return response


def next_tc_number():
    timestamp = timezone.localtime().strftime("%Y%m%d%H%M%S")
    base = f"TC-{timestamp}"
    tc_number = base
    suffix = 1

    while TransferCertificate.objects.filter(tc_number=tc_number).exists():
        suffix += 1
        tc_number = f"{base}-{suffix}"

    return tc_number


def tc_detail(request, pk):
    student = get_object_or_404(Student.objects.select_related("current_class", "current_section"), pk=pk)
    tc = getattr(student, "transfer_certificate", None)

    if request.method == "POST":
        instance = tc if tc else TransferCertificate(student=student)
        form = TransferCertificateForm(request.POST, instance=instance)
        if form.is_valid():
            tc_obj = form.save(commit=False)
            tc_obj.student = student
            tc_obj.book_no = student.scholar_register_no
            tc_obj.sr_no = student.admission_no or str(student.legacy_sid or "")
            if not tc_obj.tc_number:
                tc_obj.tc_number = next_tc_number()
            if not tc_obj.last_class_studied_id and student.current_class_id:
                tc_obj.last_class_studied_id = student.current_class_id
            if not tc_obj.last_section and student.current_section:
                tc_obj.last_section = student.current_section.name
            tc_obj.save()
            return redirect("core:tc_detail", pk=student.pk)
    else:
        initial = {}
        if not tc:
            initial = {
                "last_class_studied": student.current_class_id,
                "last_section": student.current_section.name if student.current_section else "",
            }
        form = TransferCertificateForm(instance=tc, initial=initial)

    return render(
        request,
        "core/tc_form.html",
        {
            "student": student,
            "tc": tc,
            "form": form,
        },
    )


def tc_pdf(request, pk):
    student = get_object_or_404(Student, pk=pk)
    tc = get_object_or_404(TransferCertificate.objects.select_related("student", "last_class_studied"), student=student)
    pdf_bytes = build_transfer_certificate_pdf(tc, get_active_school_profile())
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{tc.tc_number}.pdf"'
    return response


def scholar_register_pdf(request, pk):
    # Unlike the TC (which only exists once a student has left), the Scholar
    # Register is the office's permanent record from admission onward - so
    # this is available for any student, active or left.
    student = get_object_or_404(
        Student.objects.select_related("current_class", "current_section", "transfer_certificate", "transfer_certificate__last_class_studied"),
        pk=pk,
    )
    pdf_bytes = build_scholar_register_pdf(student, get_active_school_profile())
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="scholar-register-{student.admission_no or student.pk}.pdf"'
    return response


def _resolve_book_range(request):
    """Reads book/from_no/to_no GET params (same names/semantics as the
    Print Register page's range controls) and resolves them into a concrete
    (book_no, from_no, to_no) integer range. Returns (None, None, None) if
    no usable range was supplied - the full register book only makes sense
    for a fixed, contiguous number range."""
    book_no_raw = request.GET.get("book", "").strip()
    from_raw = request.GET.get("from_no", "").strip()
    to_raw = request.GET.get("to_no", "").strip()

    book_no = int(book_no_raw) if book_no_raw.isdigit() and int(book_no_raw) > 0 else None
    if book_no and not from_raw:
        from_raw = str((book_no - 1) * 100 + 1)
    if book_no and not to_raw:
        to_raw = str(book_no * 100)

    from_no = int(from_raw) if from_raw.isdigit() else None
    to_no = int(to_raw) if to_raw.isdigit() else None
    if from_no is None or to_no is None or from_no > to_no:
        return None, None, None
    return book_no, from_no, to_no


def _scholar_register_book_entries(from_no, to_no):
    """One entry per SID/Admission number in [from_no, to_no], in order.
    Entry is (sid, student_or_None) - None means no student record exists
    for that number (a 'Not Allotted' slot). Includes students of every
    status (active, inactive, TC-issued) by the owner's explicit decision -
    a register is a permanent ledger, students aren't removed from it when
    they leave."""
    students_by_sid = {
        s.legacy_sid: s
        for s in Student.objects.filter(legacy_sid__gte=from_no, legacy_sid__lte=to_no).select_related(
            "current_class", "current_section", "transfer_certificate", "transfer_certificate__last_class_studied"
        )
    }
    return [(n, students_by_sid.get(n)) for n in range(from_no, to_no + 1)]


def scholar_register_book_pdf(request):
    book_no, from_no, to_no = _resolve_book_range(request)
    if from_no is None:
        messages.error(request, "Full Register Book print karne ke liye Book No. ya From/To SID range dein.")
        return redirect("core:student_register")
    entries = _scholar_register_book_entries(from_no, to_no)
    pdf_bytes = build_scholar_register_book_pdf(entries, book_no, from_no, to_no, get_active_school_profile())
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    label = f"book-{book_no}" if book_no else f"{from_no}-{to_no}"
    response["Content-Disposition"] = f'inline; filename="scholar-register-{label}.pdf"'
    return response


def scholar_register_index_pdf(request):
    book_no, from_no, to_no = _resolve_book_range(request)
    if from_no is None:
        messages.error(request, "Scholar Register Index print karne ke liye Book No. ya From/To SID range dein.")
        return redirect("core:student_register")
    entries = _scholar_register_book_entries(from_no, to_no)
    pdf_bytes = build_scholar_register_index_pdf(entries, book_no, from_no, to_no, get_active_school_profile())
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    label = f"book-{book_no}" if book_no else f"{from_no}-{to_no}"
    response["Content-Disposition"] = f'inline; filename="scholar-register-index-{label}.pdf"'
    return response


def marksheet_select(request, pk):
    student = get_object_or_404(Student.objects.select_related("current_class", "current_section"), pk=pk)
    terms = ExamTerm.objects.select_related("session").filter(
        tests__marks__student=student
    ).distinct().order_by("-session__name", "display_order")
    return render(request, "core/marksheet_select.html", {"student": student, "terms": terms})


def check_duplicate_receipt(request):
    student_id = request.GET.get("student")
    session_id = request.GET.get("session")
    from_m = request.GET.get("from_month", "").upper()
    to_m = request.GET.get("to_month", "").upper()

    if not all([student_id, session_id, from_m, to_m]):
        return JsonResponse({"error": "Missing parameters"}, status=400)

    academic_months = ["APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC", "JAN", "FEB", "MAR"]
    try:
        start_idx = academic_months.index(from_m)
        end_idx = academic_months.index(to_m)
        if start_idx > end_idx:
            # Handle reverse range or wrap around if needed, but normally it shouldn't happen.
            start_idx, end_idx = end_idx, start_idx
    except ValueError:
        return JsonResponse({"warning": False})

    requested_range = set(range(start_idx, end_idx + 1))
    
    existing = FeeReceipt.objects.filter(
        student_id=student_id,
        session_id=session_id,
        is_cancelled=False,
    ).values("receipt_no", "from_month", "to_month")

    for receipt in existing:
        r_from = receipt["from_month"].upper()
        r_to = receipt["to_month"].upper()
        try:
            r_start = academic_months.index(r_from)
            r_end = academic_months.index(r_to)
            if r_start > r_end:
                r_start, r_end = r_end, r_start
            existing_range = set(range(r_start, r_end + 1))
            
            if requested_range.intersection(existing_range):
                return JsonResponse({
                    "warning": True, 
                    "message": f"A receipt ({receipt['receipt_no']}) already covers {r_from} to {r_to} in this session."
                })
        except ValueError:
            continue

    return JsonResponse({"warning": False})


def marksheet_pdf(request, pk, term_id):
    student = get_object_or_404(Student.objects.select_related("current_class", "current_section"), pk=pk)
    term = get_object_or_404(ExamTerm.objects.select_related("session"), pk=term_id)
    exam_marks = (
        ExamMark.objects.select_related("exam_test", "exam_test__subject")
        .filter(student=student, exam_test__term=term)
        .order_by("exam_test__subject__display_order", "exam_test__subject__name")
    )
    if not exam_marks.exists():
        return HttpResponse("No marks found for this student and term.", status=404)

    pdf_bytes = build_marksheet_pdf(student, term, list(exam_marks), get_active_school_profile())
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="marksheet-{student.admission_no or student.pk}-{term.name}.pdf"'
    return response


def discipline_list(request, pk):
    student = get_object_or_404(Student, pk=pk)
    records = student.discipline_records.select_related("reported_by").all()
    school_profile = get_active_school_profile()
    return render(
        request,
        "core/discipline_list.html",
        {
            "student": student,
            "records": records,
            "school_name": school_profile.name if school_profile else "",
        },
    )


def discipline_create(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == "POST":
        form = DisciplineRecordForm(request.POST)
        if form.is_valid():
            record = form.save(commit=False)
            record.student = student
            record.reported_by = request.user
            record.save()
            messages.success(request, "Discipline record added.")
            return redirect("core:discipline_list", pk=student.pk)
    else:
        form = DisciplineRecordForm()
    return render(request, "core/discipline_form.html", {"student": student, "form": form})


def discipline_pdf(request, pk):
    student = get_object_or_404(Student, pk=pk)
    records = list(student.discipline_records.select_related("reported_by").all())
    pdf_bytes = build_discipline_summary_pdf(student, records, get_active_school_profile())
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="discipline-{student.admission_no or student.pk}.pdf"'
    return response


def inventory_item_list(request):
    items = InventoryItem.objects.all()
    return render(request, "core/inventory_item_list.html", {"items": items})


def inventory_item_create(request):
    if request.method == "POST":
        form = InventoryItemForm(request.POST)
        if form.is_valid():
            item = form.save()
            messages.success(request, f"Item '{item.name}' added.")
            return redirect("core:inventory_item_list")
    else:
        form = InventoryItemForm()
    return render(request, "core/inventory_item_form.html", {"form": form})


def inventory_item_toggle_active(request, pk):
    item = get_object_or_404(InventoryItem, pk=pk)
    if request.method == "POST":
        item.is_active = not item.is_active
        item.save(update_fields=["is_active"])
        messages.success(request, f"'{item.name}' is now {'active' if item.is_active else 'inactive'}.")
    return redirect("core:inventory_item_list")


def inventory_report(request):
    issues = InventoryIssue.objects.select_related(
        "student", "student__current_class", "student__current_section", "item"
    ).all()

    query = request.GET.get("q", "").strip()
    class_id = request.GET.get("class", "").strip()
    item_id = request.GET.get("item", "").strip()
    date_from = request.GET.get("from_date", "").strip()
    date_to = request.GET.get("to_date", "").strip()

    if query:
        issues = issues.filter(
            Q(student__full_name__icontains=query)
            | Q(student__admission_no__icontains=query)
            | Q(student__legacy_sid__icontains=query)
        )
    if class_id:
        issues = issues.filter(student__current_class_id=class_id)
    if item_id:
        issues = issues.filter(item_id=item_id)
    if date_from:
        issues = issues.filter(issue_date__gte=date_from)
    if date_to:
        issues = issues.filter(issue_date__lte=date_to)

    issues = issues.order_by("-issue_date", "-id")
    totals = issues.aggregate(total_quantity=Sum("quantity"), total_charged=Sum("amount_charged"))

    paginator = Paginator(issues, 50)
    page = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "core/inventory_report.html",
        {
            "page": page,
            "query": query,
            "selected_class": class_id,
            "selected_item": item_id,
            "date_from": date_from,
            "date_to": date_to,
            "classes": SchoolClass.objects.order_by("display_order", "name"),
            "items": InventoryItem.objects.filter(is_active=True),
            "totals": totals,
            "total_records": issues.count(),
        },
    )


def inventory_issue_list(request, pk):
    student = get_object_or_404(Student, pk=pk)
    issues = student.inventory_issues.select_related("item", "issued_by").all()
    total_charged = issues.aggregate(total=Sum("amount_charged"))["total"] or 0
    return render(
        request,
        "core/inventory_issue_list.html",
        {"student": student, "issues": issues, "total_charged": total_charged},
    )


def inventory_issue_create(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == "POST":
        form = InventoryIssueForm(request.POST)
        if form.is_valid():
            issue = form.save(commit=False)
            issue.student = student
            issue.unit_price = issue.item.unit_price
            issue.issued_by = request.user
            issue.save()
            messages.success(request, f"Issued {issue.item.name} to {student.full_name}.")
            return redirect("core:inventory_issue_list", pk=student.pk)
    else:
        form = InventoryIssueForm()
    return render(request, "core/inventory_issue_form.html", {"student": student, "form": form})


def _class_label(student):
    if student.current_class and student.current_section:
        return f"{student.current_class}-{student.current_section.name}"
    return str(student.current_class) if student.current_class else ""


def _student_due_total(student):
    return FeeReceipt.objects.filter(
        student=student, is_cancelled=False, carried_forward=False, legacy_due_amount__gt=0
    ).aggregate(total=Sum("legacy_due_amount"))["total"] or Decimal("0")


def family_list(request):
    query = request.GET.get("q", "").strip()
    families = Family.objects.annotate(member_count=Count("members", distinct=True))
    if query:
        families = families.filter(Q(name__icontains=query) | Q(primary_mobile__icontains=query))
    families = families.order_by("name")
    return render(request, "core/family_list.html", {"families": families, "query": query})


def family_create(request):
    if request.method == "POST":
        form = FamilyForm(request.POST)
        if form.is_valid():
            family = form.save()
            messages.success(request, f"Family '{family.name}' created.")
            return redirect("core:family_detail", pk=family.pk)
    else:
        form = FamilyForm()
    return render(request, "core/family_form.html", {"form": form})


def family_detail(request, pk):
    family = get_object_or_404(Family, pk=pk)
    members = family.members.select_related("current_class", "current_section").order_by("full_name")

    member_dues = []
    total_due = Decimal("0")
    for member in members:
        due = _student_due_total(member)
        member_dues.append({"student": member, "due": due, "class_label": _class_label(member)})
        total_due += due

    query = request.GET.get("q", "").strip()
    search_results = []
    if query:
        search_results = (
            Student.objects.filter(
                Q(full_name__icontains=query)
                | Q(admission_no__icontains=query)
                | Q(legacy_sid__icontains=query)
            )
            .exclude(family=family)
            .select_related("current_class", "current_section")[:20]
        )

    school_profile = get_active_school_profile()
    school_name = school_profile.name if school_profile else ""

    members_tuples = [(md["student"].full_name, md["class_label"], md["due"]) for md in member_dues]
    family_message = family_due_message(family.name, members_tuples, total_due, school_name)
    family_wa_url = build_wa_link(family.primary_mobile, family_message)

    return render(
        request,
        "core/family_detail.html",
        {
            "family": family,
            "member_dues": member_dues,
            "total_due": total_due,
            "query": query,
            "search_results": search_results,
            "school_name": school_name,
            "family_wa_url": family_wa_url,
        },
    )


def family_add_student(request, pk):
    family = get_object_or_404(Family, pk=pk)
    if request.method == "POST":
        student = get_object_or_404(Student, pk=request.POST.get("student_id"))
        student.family = family
        student.save(update_fields=["family"])
        messages.success(request, f"{student.full_name} added to '{family.name}'.")
    return redirect("core:family_detail", pk=family.pk)


def family_remove_student(request, pk, student_id):
    family = get_object_or_404(Family, pk=pk)
    if request.method == "POST":
        student = get_object_or_404(Student, pk=student_id, family=family)
        student.family = None
        student.save(update_fields=["family"])
        messages.success(request, f"{student.full_name} removed from '{family.name}'.")
    return redirect("core:family_detail", pk=family.pk)


def family_suggestions(request):
    unlinked = (
        Student.objects.filter(family__isnull=True, is_active=True)
        .exclude(father_name="")
        .exclude(mobile_primary="")
        .select_related("current_class", "current_section")
    )

    groups = {}
    for student in unlinked:
        key = (student.father_name.strip().lower(), student.mobile_primary.strip())
        groups.setdefault(key, []).append(student)

    suggestions = [
        {
            "father_name": students[0].father_name.strip(),
            "mobile": key[1],
            "students": students,
        }
        for key, students in groups.items()
        if len(students) >= 2
    ]
    suggestions.sort(key=lambda g: g["father_name"])

    return render(request, "core/family_suggestions.html", {"suggestions": suggestions})


def family_create_from_suggestion(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        mobile = request.POST.get("mobile", "").strip()
        student_ids = request.POST.getlist("student_ids")
        if name and student_ids:
            family = Family.objects.create(name=name, primary_mobile=mobile)
            Student.objects.filter(id__in=student_ids, family__isnull=True).update(family=family)
            messages.success(request, f"Family '{family.name}' created with {len(student_ids)} student(s) linked.")
            return redirect("core:family_detail", pk=family.pk)
        messages.error(request, "Select at least one student and provide a family name.")
    return redirect("core:family_suggestions")


def staff_list(request):
    query = request.GET.get("q", "").strip()
    staff_type = request.GET.get("staff_type", "").strip()
    status = request.GET.get("status", "active").strip()
    staff_qs = Staff.objects.order_by("full_name")

    if query:
        staff_qs = staff_qs.filter(
            Q(full_name__icontains=query)
            | Q(designation__icontains=query)
            | Q(phone__icontains=query)
            | Q(legacy_emp_code__icontains=query)
        )

    if staff_type:
        staff_qs = staff_qs.filter(staff_type=staff_type)

    if status == "active":
        staff_qs = staff_qs.filter(is_active=True)
    elif status == "left":
        staff_qs = staff_qs.filter(is_active=False)

    paginator = Paginator(staff_qs, 50)
    page = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "core/staff_list.html",
        {
            "page": page,
            "query": query,
            "selected_staff_type": staff_type,
            "selected_status": status,
            "staff_types": Staff.StaffType.choices,
            "total_staff": staff_qs.count(),
            "active_staff": Staff.objects.filter(is_active=True).count(),
            "salary_payments_count": SalaryPayment.objects.count(),
        },
    )


def staff_detail(request, pk):
    staff = get_object_or_404(Staff, pk=pk)
    salary_payments = staff.salary_payments.order_by("-pay_month", "-id")[:12]
    return render(
        request,
        "core/staff_detail.html",
        {
            "staff": staff,
            "salary_payments": salary_payments,
        },
    )


def transport_list(request):
    query = request.GET.get("q", "").strip()
    route_id = request.GET.get("route", "").strip()
    status = request.GET.get("status", "active").strip()

    assignments = StudentTransport.objects.select_related(
        "student",
        "student__current_class",
        "student__current_section",
        "route",
        "bus",
    ).order_by(
        "student__current_class__display_order",
        "student__current_section__name",
        "student__full_name",
    )

    if query:
        assignments = assignments.filter(
            Q(student__full_name__icontains=query)
            | Q(student__father_name__icontains=query)
            | Q(student__legacy_sid__icontains=query)
            | Q(legacy_route_name__icontains=query)
            | Q(legacy_bus_label__icontains=query)
            | Q(stop_name__icontains=query)
        )

    if route_id:
        assignments = assignments.filter(route_id=route_id)

    if status == "active":
        assignments = assignments.filter(is_active=True)
    elif status == "inactive":
        assignments = assignments.filter(is_active=False)
    elif status == "enabled":
        assignments = assignments.filter(is_transport_enabled=True)
    elif status == "disabled":
        assignments = assignments.filter(is_transport_enabled=False)

    paginator = Paginator(assignments, 50)
    page = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "core/transport_list.html",
        {
            "page": page,
            "query": query,
            "selected_route": route_id,
            "selected_status": status,
            "routes": TransportRoute.objects.order_by("name"),
            "route_summary": TransportRoute.objects.annotate(
                assignment_count=Count("student_assignments")
            ).order_by("name")[:12],
            "total_assignments": assignments.count(),
            "active_assignments": StudentTransport.objects.filter(is_active=True).count(),
            "enabled_assignments": StudentTransport.objects.filter(is_transport_enabled=True).count(),
            "total_routes": TransportRoute.objects.count(),
            "total_buses": TransportBus.objects.count(),
        },
    )


def next_slip_no():
    timestamp = timezone.localtime().strftime("%Y%m%d%H%M%S")
    base = f"SAL-{timestamp}"
    slip_no = base
    suffix = 1

    while SalaryPayment.objects.filter(slip_no=slip_no).exists():
        suffix += 1
        slip_no = f"{base}-{suffix}"

    return slip_no


def salary_payment_create(request):
    if request.method == "POST":
        form = SalaryPaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.slip_no = next_slip_no()
            payment.save()
            SalaryPaymentAuditLog.objects.create(
                payment=payment,
                action=SalaryPaymentAuditLog.ActionChoices.CREATED,
                changed_by=request.user if request.user.is_authenticated else None,
                reason="Initial generation"
            )
            # Build post-save status message for the clerk
            monthly = payment.staff.basic_pay + payment.staff.da + payment.staff.other_allowances
            total_paid = SalaryPayment.objects.filter(
                staff=payment.staff, pay_month=payment.pay_month, is_cancelled=False
            ).aggregate(t=Sum("amount_paid"))["t"] or Decimal("0.00")
            remaining = monthly - total_paid
            month_label = payment.pay_month.strftime("%B %Y")
            if monthly <= 0:
                salary_msg = f"{month_label} ki payment â‚¹{payment.amount_paid:,.0f} save ho gayi."
            elif remaining <= 0:
                salary_msg = f"{month_label} ki salary â‚¹{monthly:,.0f} poori tarah paid ho gayi. âœ…"
            elif remaining < 0:
                advance_amt = abs(remaining)
                salary_msg = f"{month_label} ki salary â‚¹{monthly:,.0f} se â‚¹{advance_amt:,.0f} advance diya gaya. ðŸ”µ"
            else:
                salary_msg = (
                    f"{month_label} ki salary â‚¹{monthly:,.0f} mein se â‚¹{total_paid:,.0f} paid. "
                    f"â‚¹{remaining:,.0f} abhi bhi baaki hai. ðŸŸ¡"
                )
            request.session["salary_msg"] = salary_msg
            return redirect("core:salary_payment_detail", pk=payment.pk)
    else:
        form = SalaryPaymentForm()

    active_staff = Staff.objects.filter(is_active=True)
    staff_defaults = {}
    pending_advances = {}

    for staff in active_staff:
        monthly_salary = float(staff.basic_pay + staff.da + staff.other_allowances)
        staff_defaults[staff.id] = {
            "basic_pay": str(staff.basic_pay),
            "da": str(staff.da),
            "other_allowances": str(staff.other_allowances),
            "monthly_salary": monthly_salary,
        }

        adv_given = Voucher.objects.filter(
            staff=staff,
            debit_account__group__name__iexact="Advance Given",
            is_cancelled=False
        ).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

        adv_recovered = SalaryPayment.objects.filter(
            staff=staff,
            is_cancelled=False
        ).aggregate(total=Sum("advance_recovery"))["total"] or Decimal("0.00")

        pending_advances[staff.id] = float(adv_given - adv_recovered)

    return render(
        request,
        "core/salary_payment_form.html",
        {
            "form": form,
            "staff_defaults": staff_defaults,
            "pending_advances": pending_advances,
        },
    )


def salary_status_api(request):
    staff_id = request.GET.get("staff_id")
    month = request.GET.get("month")
    if not staff_id or not month:
        return JsonResponse({"error": "Missing params"}, status=400)

    try:
        from datetime import datetime
        # Parse DD/MM/YYYY to a standard YYYY-MM-DD date object
        for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%Y-%m"]:
                try:
                    parsed_month = datetime.strptime(month, fmt).date()
                    break
                except ValueError:
                    continue

        staff = Staff.objects.get(pk=staff_id)
        monthly = staff.basic_pay + staff.da + staff.other_allowances
        total_paid = SalaryPayment.objects.filter(
            staff=staff,
            pay_month__year=parsed_month.year,
            pay_month__month=parsed_month.month,
            is_cancelled=False,
        ).aggregate(t=Sum("amount_paid"))["t"] or Decimal("0.00")

        try:
            with open(r"C:\Users\THPSIC THINKCLINTE2\.gemini\antigravity\scratch\api_debug.txt", "a") as f:
                f.write(f"API CALL: staff_id={staff_id}, month={month}, parsed_month={parsed_month}, total_paid={total_paid}\n")
        except Exception:
            pass

        response = JsonResponse({
            "monthly_salary": float(monthly),
            "total_paid": float(total_paid),
            "remaining": float(monthly - total_paid),
        })
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response
    except Exception:
        return JsonResponse({"error": "Invalid data"}, status=400)


def salary_payslip_pdf(request, pk):
    payment = get_object_or_404(SalaryPayment.objects.select_related("staff"), pk=pk)
    pdf_bytes = build_salary_payslip_pdf(payment, get_active_school_profile())
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{payment.slip_no}.pdf"'
    return response


def _handle_student_transport(form, student):
    transport_required = form.cleaned_data.get("transport_required")
    active_session = AcademicSession.objects.filter(is_active=True).order_by("-starts_on").first()
    if not active_session:
        return
        
    if transport_required:
        route = form.cleaned_data.get("transport_route")
        stop_name = form.cleaned_data.get("stop_name")
        
        transport = StudentTransport.objects.filter(student=student, session=active_session).first()
        if not transport:
            transport = StudentTransport(student=student, session=active_session)
            
        transport.route = route
        transport.stop_name = stop_name
        transport.monthly_amount = route.monthly_charge if route else 0
        transport.start_month = "APR"
        transport.end_month = "MAR"
        transport.is_transport_enabled = True
        transport.billing_confirmed = True
        transport.is_active = True
        transport.save()
    else:
        active_transports = StudentTransport.objects.filter(student=student, session=active_session, is_active=True)
        for t in active_transports:
            t.is_active = False
            t.save()


def _handle_student_concession(post_data, student):
    """
    Save or update a StudentConcession from the student create/edit form.
    If concession_type is blank, do nothing (no concession to assign).
    """
    concession_type = post_data.get("concession_type", "").strip()
    if not concession_type:
        return  # No concession chosen — leave existing record untouched.

    active_session = AcademicSession.objects.filter(is_active=True).order_by("-starts_on").first()
    if not active_session:
        return

    is_full_free = concession_type == "full_free"

    if is_full_free:
        # Full Free: no amount needed, amount_type = 'full'.
        amount = None
        amount_type = "full"
    else:
        raw = post_data.get("concession_amount", "").strip()
        try:
            amount = Decimal(raw)
            if amount < 0:
                return  # Invalid amount — skip silently.
        except (ValueError, TypeError, Exception):
            return  # Can't parse — skip.
        amount_type = "fixed"

    StudentConcession.objects.update_or_create(
        student=student,
        session=active_session,
        defaults={
            "concession_type": concession_type,
            "amount_type": amount_type,
            "from_month": post_data.get("concession_from_month", "").strip(),
            "to_month": post_data.get("concession_to_month", "").strip(),
            "amount": amount,
            "reason": post_data.get("concession_reason", "").strip(),
            "approved_by_name": post_data.get("concession_approved_by", "").strip(),  # correct field name
            "is_active": post_data.get("concession_is_active") == "on",
        },
    )


def student_create(request):
    if request.method == "POST":
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            student = form.save()
            _handle_student_transport(form, student)
            _handle_student_concession(request.POST, student)
            action = request.POST.get("action")
            if action == "save_new":
                return redirect("core:student_create")
            return redirect("core:student_detail", pk=student.pk)
    else:
        form = StudentForm()
        
    recent_students = Student.objects.order_by("-id")[:5]
    context = {
        "form": form,
        "recent_students": recent_students,
        "is_edit": False,
    }
    return render(request, "core/student_form.html", context)


def student_update(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == "POST":
        form = StudentForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            form.save()
            _handle_student_transport(form, student)
            _handle_student_concession(request.POST, student)
            return redirect("core:student_detail", pk=student.pk)
    else:
        form = StudentForm(instance=student)

    # Pass existing active concession so the template can pre-fill fields.
    active_session = AcademicSession.objects.filter(is_active=True).order_by("-starts_on").first()
    existing_concession = None
    if active_session:
        existing_concession = StudentConcession.objects.filter(
            student=student, session=active_session
        ).first()

    recent_students = Student.objects.order_by("-id")[:5]
    context = {
        "form": form,
        "student": student,
        "recent_students": recent_students,
        "is_edit": True,
        "existing_concession": existing_concession,
    }
    return render(request, "core/student_form.html", context)


def check_student_duplicate(request):
    """Async endpoint to check for duplicate students."""
    field = request.GET.get("field")
    value = request.GET.get("value", "").strip()
    exclude_id = request.GET.get("exclude_id")
    
    if not value or not field:
        return JsonResponse({"duplicate": False})
        
    query = Q()
    if field == "legacy_sid":
        query = Q(legacy_sid=value)
    elif field == "admission_no":
        query = Q(admission_no=value)
    elif field == "mobile":
        query = Q(mobile_primary=value) | Q(mobile_secondary=value)
    else:
        return JsonResponse({"duplicate": False})
        
    qs = Student.objects.filter(query)
    if exclude_id and exclude_id.isdigit():
        qs = qs.exclude(pk=exclude_id)
        
    matches = qs.values_list("full_name", "current_class__name")[:3]
    
    if matches:
        names = [f"{m[0]} ({m[1] or 'No class'})" for m in matches]
        return JsonResponse({
            "duplicate": True, 
            "message": f"Found existing student(s) with this {field}: " + ", ".join(names)
        })
        
    return JsonResponse({"duplicate": False})

def student_register(request):
    """
    Renders a printable student register with the exact same filtering as student_list.
    """
    students, query, class_id, section_id, status = _get_filtered_students(request)

    book_no = request.GET.get("book", "").strip()
    from_no = request.GET.get("from_no", "").strip()
    to_no = request.GET.get("to_no", "").strip()
    if book_no.isdigit() and int(book_no) > 0:
        book_number = int(book_no)
        from_no = from_no or str(((book_number - 1) * 100) + 1)
        to_no = to_no or str(book_number * 100)

    if from_no.isdigit():
        students = students.filter(legacy_sid__gte=int(from_no))
    if to_no.isdigit():
        students = students.filter(legacy_sid__lte=int(to_no))

    # Determine filter description for header
    filter_description = []
    if class_id:
        cls = SchoolClass.objects.filter(pk=class_id).first()
        if cls:
            filter_description.append(f"Class: {cls.name}")
    if section_id:
        sec = Section.objects.filter(pk=section_id).first()
        if sec:
            filter_description.append(f"Section: {sec.name}")
    filter_description.append(f"Status: {status.title()}")
    if query:
        filter_description.append(f"Search: '{query}'")
    if from_no or to_no:
        filter_description.append(f"SID Range: {from_no or 'start'} to {to_no or 'end'}")

    context = {
        "students": students,
        "filter_description": " | ".join(filter_description),
        "total_students": students.count(),
        "print_date": timezone.now(),
        "query": query,
        "selected_class": class_id,
        "selected_section": section_id,
        "selected_status": status,
        "book_no": book_no,
        "from_no": from_no,
        "to_no": to_no,
    }
    return render(request, "core/student_register_report.html", context)

def student_export_csv(request):
    """
    Exports the currently filtered students to a CSV file.
    """
    students, query, class_id, section_id, status = _get_filtered_students(request)
    
    # Generate filename
    date_str = timezone.now().strftime("%Y%m%d")
    status_str = status.title()
    filename = f"Students_{status_str}_{date_str}.csv"
    
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    
    # Write UTF-8 BOM for Excel compatibility with Hindi/Unicode characters
    response.write('\ufeff')
    
    writer = csv.writer(response)
    # Header row
    writer.writerow([
        "Registration No",
        "Admission No",
        "SID",
        "Admission Date",
        "Student Name",
        "Class",
        "Section",
        "DOB",
        "Gender",
        "Father's Name",
        "Mother's Name",
        "Category",
        "Religion",
        "Aadhaar Number",
        "Mobile Primary",
        "Mobile Secondary",
        "Address",
        "Status"
    ])
    
    for st in students:
        class_name = st.current_class.name if st.current_class else ""
        section_name = st.current_section.name if st.current_section else ""
        adm_date = st.admission_date.strftime("%d/%m/%Y") if st.admission_date else ""
        dob = st.date_of_birth.strftime("%d/%m/%Y") if st.date_of_birth else ""
        status_text = "Active" if st.is_active else "Inactive"
        
        address = st.address_local or st.address_permanent or ""
        
        writer.writerow([
            st.registration_no,
            st.admission_no,
            st.legacy_sid,
            adm_date,
            st.full_name,
            class_name,
            section_name,
            dob,
            st.get_gender_display(),
            st.father_name,
            st.mother_name,
            st.category,
            st.religion,
            st.aadhaar_no,
            st.mobile_primary,
            st.mobile_secondary,
            address.replace("\n", " ").replace("\r", ""),
            status_text
        ])
        
    return response


def board_registration_export_csv(request, kind):
    """
    Exports UP Board registration-ready CSV files with exact template headings.
    """
    if kind not in {"class9", "class11-upboard", "class11-others"}:
        raise Http404("Unknown board registration export type")

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{export_filename(kind)}"'
    response.write("\ufeff")

    writer = csv.writer(response)
    writer.writerow(export_columns(kind))
    writer.writerows(build_rows(kind))
    return response

def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    
    if request.method == "POST":
        try:
            student.delete()
            messages.success(request, f"Student {student.full_name} deleted successfully.")
        except ProtectedError:
            messages.error(request, f"Cannot delete {student.full_name} because they have fee receipts or marks. Please mark them as Inactive instead.")
        return redirect("core:student_list")
        
    return render(request, "core/student_confirm_delete.html", {"student": student})

# ---------------------------------------------------------------------------
# Accounts / Cash Book (Phase 1)
# ---------------------------------------------------------------------------

def _generate_voucher_no(session, voucher_type):
    from django.db import connection, transaction
    from .models import VoucherCounter
    year_str = session.name
    if not year_str and session.starts_on and session.ends_on:
        year_str = f"{session.starts_on.year}-{str(session.ends_on.year)[-2:]}"
    year_str = year_str or "SESSION"

    with transaction.atomic():
        counter, _ = VoucherCounter.objects.select_for_update().get_or_create(
            session=session, voucher_type=voucher_type
        )
        counter.last_number += 1
        counter.save()
        return f"{voucher_type}-{year_str}-{counter.last_number:04d}"


def _get_cash_account():
    from .models import LedgerAccount
    return LedgerAccount.objects.filter(is_cash_or_bank=True, name__icontains="cash").first()


@login_required
@permission_required('core.access_accounts', raise_exception=True)
def expense_create(request):
    from .forms import VoucherForm
    from .models import AcademicSession, Voucher, VoucherAuditLog, AccountGroup, Staff
    from django.utils import timezone
    from django.db import connection, transaction
    
    session = AcademicSession.objects.filter(is_active=True).first()
    if not session:
        messages.error(request, "No active academic session found.")
        return redirect("core:dashboard")

    if request.method == "POST":
        form = VoucherForm(request.POST, voucher_kind="expense")
        if form.is_valid():
            voucher = form.save(commit=False)
            voucher.session = session
            voucher.voucher_type = Voucher.VoucherType.CASH_PAYMENT
            voucher.created_by = request.user
            
            # For expense/advance payment, credit should be Cash/Bank and debit should be Expense,
            # Advance Given, or a Liability being paid down (e.g. repaying a personal loan someone
            # advanced to the school - a real case: "Pragati ka paisa wapas kar diya gya" in the
            # legacy ledger. Repaying a loan reduces what the school owes, which is a Liability
            # debit, not an Expense - the old check rejected this with no way to record it).
            if not voucher.credit_account.is_cash_or_bank:
                messages.error(request, "For expenses/advances, the credit account must be a Cash or Bank account.")
            elif not (
                voucher.debit_account.group.group_type == AccountGroup.GroupType.EXPENSE
                or voucher.debit_account.group.group_type == AccountGroup.GroupType.LIABILITY
                or voucher.debit_account.group.name.lower() == "advance given"
            ):
                messages.error(request, "For expenses/advances, the debit account must be an Expense, Liability, or Advance Given account.")
            else:
                with transaction.atomic():
                    voucher.voucher_no = _generate_voucher_no(session, Voucher.VoucherType.CASH_PAYMENT)
                    voucher.save()
                    
                    VoucherAuditLog.objects.create(
                        voucher=voucher,
                        action=VoucherAuditLog.Action.CREATED,
                        changed_by=request.user,
                    )
                
                messages.success(request, f"Expense voucher {voucher.voucher_no} created successfully.")
                if "save_and_add_another" in request.POST:
                    return redirect("core:expense_create")
                return redirect("core:voucher_detail", pk=voucher.pk)
    else:
        initial = {"credit_account": _get_cash_account()}
        form = VoucherForm(initial=initial, voucher_kind="expense")

    return render(request, "core/expense_form.html", {"form": form})

@login_required
@permission_required('core.access_accounts', raise_exception=True)
def receipt_other_create(request):
    from .forms import VoucherForm
    from .models import AcademicSession, Voucher, VoucherAuditLog, AccountGroup
    from django.utils import timezone
    from django.db import connection, transaction
    
    session = AcademicSession.objects.filter(is_active=True).first()
    if not session:
        messages.error(request, "No active academic session found.")
        return redirect("core:dashboard")

    if request.method == "POST":
        form = VoucherForm(request.POST, voucher_kind="receipt")
        if form.is_valid():
            voucher = form.save(commit=False)
            voucher.session = session
            voucher.voucher_type = Voucher.VoucherType.CASH_RECEIPT
            voucher.created_by = request.user
            
            # For receipts, debit should be Cash/Bank, credit should be Income OR a Liability
            # increasing (e.g. someone personally advancing cash to the school - a real case:
            # "PRAGATI PERSONAL/ADVANCE A/C ... SCHOOL KE KHARCH KE LIYE LOAN LIYA GAYA" in the
            # legacy ledger). Receiving a loan is real cash in, but it's not Income - it's a
            # Liability the school now owes back. The old check rejected this with no way to
            # record it, which would have blocked the owner's own decision to move this kind of
            # entry into the new app going forward.
            if not voucher.debit_account.is_cash_or_bank:
                messages.error(request, "For receipts, the debit account must be a Cash or Bank account.")
            elif voucher.credit_account.group.group_type not in (
                AccountGroup.GroupType.INCOME,
                AccountGroup.GroupType.LIABILITY,
            ):
                messages.error(request, "For receipts, the credit account must be an Income or Liability account.")
            else:
                with transaction.atomic():
                    voucher.voucher_no = _generate_voucher_no(session, Voucher.VoucherType.CASH_RECEIPT)
                    voucher.save()
                    
                    VoucherAuditLog.objects.create(
                        voucher=voucher,
                        action=VoucherAuditLog.Action.CREATED,
                        changed_by=request.user,
                    )
                
                messages.success(request, f"Receipt voucher {voucher.voucher_no} created successfully.")
                if "save_and_add_another" in request.POST:
                    return redirect("core:receipt_other_create")
                return redirect("core:voucher_detail", pk=voucher.pk)
    else:
        initial = {"debit_account": _get_cash_account()}
        form = VoucherForm(initial=initial, voucher_kind="receipt")
        
    return render(request, "core/receipt_other_form.html", {"form": form})


@login_required
@permission_required('core.access_accounts', raise_exception=True)
def ledger_list(request):
    from .models import LedgerAccount, AccountGroup
    ledgers = LedgerAccount.objects.select_related("group").all()
    return render(request, "core/ledger_list.html", {"ledgers": ledgers})


@login_required
@permission_required('core.access_accounts', raise_exception=True)
def person_list(request):
    from .models import Person, Voucher
    from django.db.models import Sum

    persons = Person.objects.prefetch_related("ledgers").all()
    
    for p in persons:
        net_balance = 0
        for l in p.ledgers.all():
            debits = Voucher.objects.filter(debit_account=l, is_cancelled=False).aggregate(s=Sum('amount'))['s'] or 0
            credits = Voucher.objects.filter(credit_account=l, is_cancelled=False).aggregate(s=Sum('amount'))['s'] or 0
            net_balance += (l.opening_balance + debits - credits)
        p.net_balance = net_balance
        p.balance_status = "DR" if net_balance > 0 else "CR"
        p.abs_balance = abs(net_balance)

    return render(request, "core/person_list.html", {"persons": persons})


@login_required
@permission_required('core.access_accounts', raise_exception=True)
def person_detail(request, pk):
    from .models import Person, Voucher
    from django.shortcuts import get_object_or_404
    from django.db.models import Sum, Q

    person = get_object_or_404(Person, pk=pk)
    
    net_balance = 0
    ledgers = person.ledgers.all()
    for l in ledgers:
        debits = Voucher.objects.filter(debit_account=l, is_cancelled=False).aggregate(s=Sum('amount'))['s'] or 0
        credits = Voucher.objects.filter(credit_account=l, is_cancelled=False).aggregate(s=Sum('amount'))['s'] or 0
        l.current_balance = l.opening_balance + debits - credits
        l.balance_status = "DR" if l.current_balance > 0 else "CR"
        l.abs_balance = abs(l.current_balance)
        net_balance += l.current_balance
    
    person.net_balance = net_balance
    person.balance_status = "DR" if net_balance > 0 else "CR"
    person.abs_balance = abs(net_balance)

    vouchers = Voucher.objects.filter(
        Q(debit_account__person=person) | Q(credit_account__person=person),
        is_cancelled=False
    ).select_related('debit_account', 'credit_account').order_by('voucher_date', 'id')

    return render(request, "core/person_detail.html", {
        "person": person,
        "ledgers": ledgers,
        "vouchers": vouchers,
    })



@login_required
@permission_required('core.access_accounts', raise_exception=True)
def ledger_create(request):
    from .forms import LedgerAccountForm
    if request.method == "POST":
        form = LedgerAccountForm(request.POST)
        if form.is_valid():
            ledger = form.save()
            messages.success(request, f"Ledger '{ledger.name}' created.")
            return redirect("core:ledger_list")
    else:
        form = LedgerAccountForm()
    return render(request, "core/ledger_form.html", {"form": form, "is_edit": False})


@login_required
@permission_required('core.access_accounts', raise_exception=True)
def ledger_edit(request, pk):
    from .models import LedgerAccount
    from .forms import LedgerAccountForm
    ledger = get_object_or_404(LedgerAccount, pk=pk)
    if request.method == "POST":
        form = LedgerAccountForm(request.POST, instance=ledger)
        if form.is_valid():
            form.save()
            messages.success(request, f"Ledger '{ledger.name}' updated.")
            return redirect("core:ledger_list")
    else:
        form = LedgerAccountForm(instance=ledger)
    return render(request, "core/ledger_form.html", {"form": form, "is_edit": True})


@login_required
@permission_required('core.access_accounts', raise_exception=True)
def voucher_list(request):
    from .models import Voucher
    vouchers = Voucher.objects.select_related("debit_account", "credit_account").all()
    
    q = request.GET.get("q", "").strip()
    if q:
        vouchers = vouchers.filter(
            Q(voucher_no__icontains=q) |
            Q(debit_account__name__icontains=q) |
            Q(credit_account__name__icontains=q) |
            Q(paid_to_or_received_from__icontains=q) |
            Q(narration__icontains=q)
        )
        
    v_type = request.GET.get("type", "")
    if v_type:
        vouchers = vouchers.filter(voucher_type=v_type)
        
    paginator = Paginator(vouchers, 50)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    
    return render(request, "core/voucher_list.html", {
        "page_obj": page_obj,
        "q": q,
        "v_type": v_type,
        "voucher_types": Voucher.VoucherType.choices,
    })


@login_required
@permission_required('core.access_accounts', raise_exception=True)
def voucher_detail(request, pk):
    from .models import Voucher
    voucher = get_object_or_404(Voucher.objects.select_related("debit_account", "credit_account", "created_by", "cancelled_by", "edited_by"), pk=pk)
    logs = voucher.audit_logs.select_related("changed_by").all()
    return render(request, "core/voucher_detail.html", {"voucher": voucher, "logs": logs})


@login_required
@permission_required('core.access_accounts', raise_exception=True)
def voucher_edit(request, pk):
    from .models import Voucher, VoucherAuditLog
    from .forms import VoucherEditForm
    from django.db import connection, transaction
    from django.forms.models import model_to_dict
    
    voucher = get_object_or_404(Voucher, pk=pk)
    if voucher.is_cancelled:
        messages.error(request, "Cannot edit a cancelled voucher.")
        return redirect("core:voucher_detail", pk=voucher.pk)
        
    def _json_safe(d):
        """Convert model_to_dict output to a JSON-serializable dict.
        DjangoJSONEncoder handles date, datetime, Decimal, UUID, etc."""
        import json
        from django.core.serializers.json import DjangoJSONEncoder
        return json.loads(json.dumps(d, cls=DjangoJSONEncoder, default=str))

    if request.method == "POST":
        before_snapshot = _json_safe(model_to_dict(voucher))
        form = VoucherEditForm(request.POST, instance=voucher)
        if form.is_valid():

            updated = form.save(commit=False)
            updated.is_edited = True
            updated.edited_at = timezone.now()
            updated.edited_by = request.user
            updated.edit_count += 1
            updated.edit_reason = form.cleaned_data["edit_reason"]

            with transaction.atomic():
                updated.save()
                after_snapshot = _json_safe(model_to_dict(updated))
                
                # compute changes
                changes = {}
                for field in before_snapshot:
                    if field in ["edited_at", "edited_by", "edit_count", "is_edited", "edit_reason"]:
                        continue
                    if str(before_snapshot[field]) != str(after_snapshot[field]):
                        changes[field] = {"old": str(before_snapshot[field]), "new": str(after_snapshot[field])}
                
                VoucherAuditLog.objects.create(
                    voucher=updated,
                    action=VoucherAuditLog.Action.EDITED,
                    changed_by=request.user,
                    reason=updated.edit_reason,
                    before_snapshot=before_snapshot,
                    after_snapshot=after_snapshot,
                    changes=changes,
                )
            
            messages.success(request, f"Voucher {voucher.voucher_no} edited.")
            return redirect("core:voucher_detail", pk=voucher.pk)
    else:
        form = VoucherEditForm(instance=voucher)
        
    return render(request, "core/voucher_form.html", {"form": form, "voucher": voucher, "is_edit": True})


@login_required
@permission_required('core.access_accounts', raise_exception=True)
def voucher_cancel(request, pk):
    from .models import Voucher, VoucherAuditLog
    from django.db import connection, transaction
    
    voucher = get_object_or_404(Voucher, pk=pk)
    if voucher.is_cancelled:
        messages.warning(request, "Voucher is already cancelled.")
        return redirect("core:voucher_detail", pk=voucher.pk)
        
    if request.method == "POST":
        reason = request.POST.get("cancel_reason", "").strip()
        if not reason:
            messages.error(request, "Cancellation reason is required.")
        else:
            with transaction.atomic():
                voucher.is_cancelled = True
                voucher.cancelled_at = timezone.now()
                voucher.cancelled_by = request.user
                voucher.cancel_reason = reason
                voucher.save()
                
                VoucherAuditLog.objects.create(
                    voucher=voucher,
                    action=VoucherAuditLog.Action.CANCELLED,
                    changed_by=request.user,
                    reason=reason,
                )
                
            messages.success(request, f"Voucher {voucher.voucher_no} has been cancelled.")
            return redirect("core:voucher_detail", pk=voucher.pk)
            
    return render(request, "core/voucher_cancel_confirm.html", {"voucher": voucher})


@login_required
@permission_required('core.access_accounts', raise_exception=True)
def voucher_pdf(request, pk):
    from .models import Voucher
    from .pdf import build_voucher_pdf
    
    voucher = get_object_or_404(Voucher.objects.select_related("debit_account", "credit_account"), pk=pk)
    pdf_bytes = build_voucher_pdf(voucher, get_active_school_profile())
    
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="voucher_{voucher.voucher_no}.pdf"'
    return response


@login_required
@permission_required('core.access_accounts', raise_exception=True)
def cash_book(request):
    from .models import Voucher, FeeReceipt, SalaryPayment, LedgerAccount
    from django.db.models import Sum
    import datetime
    
    from_date_str = request.GET.get("from_date")
    to_date_str = request.GET.get("to_date")
    
    try:
        if from_date_str:
            from_date = datetime.datetime.strptime(from_date_str, "%Y-%m-%d").date()
        else:
            from_date = timezone.localdate()
    except ValueError:
        from_date = timezone.localdate()
        
    try:
        if to_date_str:
            to_date = datetime.datetime.strptime(to_date_str, "%Y-%m-%d").date()
        else:
            to_date = timezone.localdate()
    except ValueError:
        to_date = timezone.localdate()
        
    if from_date > to_date:
        from_date, to_date = to_date, from_date
        
    include_salary = request.GET.get("include_salary") == "1"
    
    cash_ledger = _get_cash_account()
    if not cash_ledger:
        messages.error(request, "Cash ledger not found. Please run seed_accounts.")
        return redirect("core:dashboard")
        
    # Get all transactions for the date range
    
    # 1. Fee Receipts (Cash Receipt)
    fee_receipts = FeeReceipt.objects.filter(
        receipt_date__gte=from_date,
        receipt_date__lte=to_date,
        is_cancelled=False,
        payment_mode=FeeReceipt.PaymentMode.CASH
    ).select_related("student")
    
    # 2. Vouchers (Cash Payments and Cash Receipts, Contra)
    vouchers = Voucher.objects.filter(
        voucher_date__gte=from_date,
        voucher_date__lte=to_date,
        is_cancelled=False,
    ).filter(Q(debit_account=cash_ledger) | Q(credit_account=cash_ledger)).select_related("debit_account", "credit_account")
    
    # 3. Salary Payments (Cash Payment)
    salary_payments = []
    if include_salary:
        salary_payments = SalaryPayment.objects.filter(
            payment_date__gte=from_date,
            payment_date__lte=to_date,
            payment_mode=SalaryPayment.PaymentMode.CASH,
            is_cancelled=False
        ).select_related("staff")
        
    # Calculate opening balance up to from_date - 1 day
    # Base OB
    ob = cash_ledger.opening_balance
    ob_date = cash_ledger.opening_balance_date
    
    if ob_date and ob_date <= from_date:
        # Calculate net change from ob_date to from_date - 1
        
        # 1. Fee Receipts sum
        fr_sum = FeeReceipt.objects.filter(
            receipt_date__gte=ob_date,
            receipt_date__lt=from_date,
            is_cancelled=False,
            payment_mode=FeeReceipt.PaymentMode.CASH
        ).aggregate(total=Sum("received_amount"))["total"] or Decimal("0.00")
        
        sal_sum = SalaryPayment.objects.filter(
            payment_date__gte=ob_date,
            payment_date__lt=from_date,
            payment_mode=SalaryPayment.PaymentMode.CASH,
            is_cancelled=False
        ).aggregate(total=Sum("amount_paid"))["total"] or Decimal("0.00")
        
        # 3. Voucher in/out
        v_in_sum = Voucher.objects.filter(
            voucher_date__gte=ob_date,
            voucher_date__lt=from_date,
            is_cancelled=False,
            debit_account=cash_ledger
        ).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
        
        v_out_sum = Voucher.objects.filter(
            voucher_date__gte=ob_date,
            voucher_date__lt=from_date,
            is_cancelled=False,
            credit_account=cash_ledger
        ).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
        
        ob = ob + fr_sum + v_in_sum - sal_sum - v_out_sum
        
    
    transactions = []
    
    for r in fee_receipts:
        transactions.append({
            "type": "FR",
            "ref": r,
            "ref_no": r.receipt_no,
            "date": r.receipt_date,
            "desc": f"Fee Receipt: {r.student.full_name}",
            "in_amt": r.received_amount,
            "out_amt": Decimal("0.00"),
            "sort_key": (r.receipt_date, r.created_at)
        })
        
    for v in vouchers:
        is_in = v.debit_account == cash_ledger
        party_name = v.staff.full_name if v.staff else v.paid_to_or_received_from
        head_name = v.credit_account.name if is_in else v.debit_account.name
        desc_bits = [part for part in [head_name, party_name, v.narration] if part]
        transactions.append({
            "type": "VOUCHER",
            "ref": v,
            "ref_no": v.voucher_no,
            "date": v.voucher_date,
            "desc": " - ".join(desc_bits),
            "in_amt": v.amount if is_in else Decimal("0.00"),
            "out_amt": Decimal("0.00") if is_in else v.amount,
            "sort_key": (v.voucher_date, v.created_at)
        })
        
    if include_salary:
        for s in salary_payments:
            desc_parts = [f"Salary: {s.staff.full_name}"]
            if s.remarks:
                desc_parts.append(f"({s.remarks})")
            
            transactions.append({
                "type": "SAL",
                "ref": s,
                "ref_no": f"SAL-{s.pk}",
                "date": s.payment_date,
                "desc": " ".join(desc_parts),
                "in_amt": Decimal("0.00"),
                "out_amt": s.amount_paid,
                "sort_key": (s.payment_date, s.created_at)
            })
            
    transactions.sort(key=lambda x: x["sort_key"])
    
    tx_by_date = {}
    for tx in transactions:
        tx_by_date.setdefault(tx["date"], []).append(tx)
        
    dates_to_process = sorted(list(tx_by_date.keys()))
    if from_date not in dates_to_process:
        dates_to_process.insert(0, from_date)
    if to_date not in dates_to_process:
        dates_to_process.append(to_date)
        
    # Remove duplicates but keep sorted order
    dates_to_process = sorted(list(set(dates_to_process)))
    
    daily_data = []
    current_balance = ob
    
    from itertools import zip_longest
    
    for d in dates_to_process:
        day_tx = tx_by_date.get(d, [])
        receipts = []
        payments = []
        
        # Add OB
        if current_balance > 0:
            receipts.append({"ref_no": "-", "desc": "To Balance b/d", "amount": current_balance})
        elif current_balance < 0:
            payments.append({"ref_no": "-", "desc": "By Balance b/d", "amount": abs(current_balance)})
            
        for tx in day_tx:
            if tx["in_amt"] > 0:
                receipts.append({"ref_no": tx["ref_no"], "desc": tx["desc"], "amount": tx["in_amt"]})
            elif tx["out_amt"] > 0:
                payments.append({"ref_no": tx["ref_no"], "desc": tx["desc"], "amount": tx["out_amt"]})
                
        day_total_in = sum(r["amount"] for r in receipts)
        day_total_out = sum(p["amount"] for p in payments)
        
        cb = day_total_in - day_total_out
        
        if not day_tx and cb == 0:
            continue
            
        if cb > 0:
            payments.append({"ref_no": "-", "desc": "By Balance c/d", "amount": cb})
            day_total = day_total_in
        elif cb < 0:
            receipts.append({"ref_no": "-", "desc": "To Balance c/d", "amount": abs(cb)})
            day_total = day_total_out
        else:
            day_total = day_total_in
            
        current_balance = cb
        
        zipped_rows = list(zip_longest(receipts, payments, fillvalue=None))
        
        daily_data.append({
            "date": d,
            "rows": zipped_rows,
            "total": day_total
        })
    
    return render(request, "core/cash_book.html", {
        "from_date": from_date,
        "to_date": to_date,
        "daily_data": daily_data,
        "include_salary": include_salary
    })


def salary_payment_list(request):
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    staff_id = request.GET.get("staff")
    payment_mode = request.GET.get("payment_mode")
    status = request.GET.get("status", "valid")

    qs = SalaryPayment.objects.select_related("staff").all()

    if from_date:
        qs = qs.filter(payment_date__gte=from_date)
    if to_date:
        qs = qs.filter(payment_date__lte=to_date)
    if staff_id:
        qs = qs.filter(staff_id=staff_id)
    if payment_mode:
        qs = qs.filter(payment_mode=payment_mode)

    if status == "valid":
        qs = qs.filter(is_cancelled=False)
    elif status == "cancelled":
        qs = qs.filter(is_cancelled=True)

    # Calculate totals
    if status == "cancelled":
        qs_for_totals = qs
    else:
        qs_for_totals = qs.filter(is_cancelled=False)

    total_gross = qs_for_totals.aggregate(t=Sum("basic_pay") + Sum("da") + Sum("other_allowances"))["t"] or Decimal("0.00")
    total_deductions = qs_for_totals.aggregate(t=Sum("pf_deduction") + Sum("esi_deduction") + Sum("other_deduction"))["t"] or Decimal("0.00")
    total_recovery = qs_for_totals.aggregate(t=Sum("advance_recovery"))["t"] or Decimal("0.00")

    total_net = total_gross - total_deductions - total_recovery

    paginator = Paginator(qs, 50)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "core/salary_list.html", {
        "page_obj": page_obj,
        "total_gross": total_gross,
        "total_deductions": total_deductions,
        "total_recovery": total_recovery,
        "total_net": total_net,
        "active_staff": Staff.objects.filter(is_active=True).order_by("full_name"),
    })


def salary_payment_detail(request, pk):
    payment = get_object_or_404(SalaryPayment.objects.select_related("staff", "cancelled_by", "edited_by"), pk=pk)
    logs = payment.audit_logs.select_related("changed_by").all()
    salary_msg = request.session.pop("salary_msg", None)
    return render(request, "core/salary_detail.html", {
        "payment": payment,
        "audit_logs": logs,
        "salary_msg": salary_msg,
    })


def salary_payment_edit(request, pk):
    payment = get_object_or_404(SalaryPayment, pk=pk)
    if payment.is_cancelled:
        messages.error(request, "Cannot edit a cancelled salary payment.")
        return redirect("core:salary_payment_detail", pk=pk)

    if request.method == "POST":
        form = SalaryPaymentForm(request.POST, instance=payment)
        edit_reason = request.POST.get("edit_reason", "").strip()
        if not edit_reason:
            form.add_error(None, "An edit reason is required.")
            
        if form.is_valid():
            updated_payment = form.save(commit=False)

            original_payment = SalaryPayment.objects.get(pk=pk)
            # Snapshot before
            before_snapshot = {
                "basic_pay": str(original_payment.basic_pay),
                "da": str(original_payment.da),
                "other_allowances": str(original_payment.other_allowances),
                "amount_paid": str(original_payment.amount_paid),
                "net_pay": str(original_payment.net_pay),
            }

            updated_payment.is_edited = True
            updated_payment.edited_at = timezone.now()
            updated_payment.edited_by = request.user if request.user.is_authenticated else None
            updated_payment.edit_reason = edit_reason
            updated_payment.edit_count += 1
            updated_payment.save()

            after_snapshot = {
                "basic_pay": str(updated_payment.basic_pay),
                "da": str(updated_payment.da),
                "other_allowances": str(updated_payment.other_allowances),
                "amount_paid": str(updated_payment.amount_paid),
                "net_pay": str(updated_payment.net_pay),
            }

            SalaryPaymentAuditLog.objects.create(
                payment=updated_payment,
                action=SalaryPaymentAuditLog.ActionChoices.EDITED,
                changed_by=request.user if request.user.is_authenticated else None,
                reason=edit_reason,
                before_snapshot=before_snapshot,
                after_snapshot=after_snapshot
            )
            messages.success(request, "Salary payment updated successfully.")
            return redirect("core:salary_payment_detail", pk=updated_payment.pk)
    else:
        form = SalaryPaymentForm(instance=payment)

    active_staff = Staff.objects.filter(is_active=True)
    staff_defaults = {}
    pending_advances = {}
    
    for staff in active_staff:
        staff_defaults[staff.id] = {
            "basic_pay": str(staff.basic_pay),
            "da": str(staff.da),
            "other_allowances": str(staff.other_allowances),
        }
        adv_given = Voucher.objects.filter(staff=staff, debit_account__group__name__iexact="Advance Given", is_cancelled=False).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
        adv_recovered = SalaryPayment.objects.filter(staff=staff, is_cancelled=False).exclude(pk=pk).aggregate(total=Sum("advance_recovery"))["total"] or Decimal("0.00")
        pending_advances[staff.id] = float(adv_given - adv_recovered)

    return render(
        request,
        "core/salary_edit.html",
        {
            "form": form,
            "payment": payment,
            "staff_defaults": staff_defaults,
            "pending_advances": pending_advances,
        },
    )


def salary_payment_cancel(request, pk):
    payment = get_object_or_404(SalaryPayment, pk=pk)
    if payment.is_cancelled:
        messages.error(request, "This salary payment is already cancelled.")
        return redirect("core:salary_payment_detail", pk=pk)

    if request.method == "POST":
        reason = request.POST.get("cancel_reason", "").strip()
        if not reason:
            messages.error(request, "A cancellation reason is required.")
        else:
            payment.is_cancelled = True
            payment.cancelled_at = timezone.now()
            payment.cancelled_by = request.user if request.user.is_authenticated else None
            payment.cancel_reason = reason
            payment.save(update_fields=["is_cancelled", "cancelled_at", "cancelled_by", "cancel_reason"])

            SalaryPaymentAuditLog.objects.create(
                payment=payment,
                action=SalaryPaymentAuditLog.ActionChoices.CANCELLED,
                changed_by=request.user if request.user.is_authenticated else None,
                reason=reason
            )
            messages.success(request, "Salary payment cancelled successfully.")
            return redirect("core:salary_payment_detail", pk=pk)

    return render(request, "core/salary_cancel_confirm.html", {"payment": payment})





def check_siblings(request):
    """Async endpoint to check for siblings based on guardian info."""
    from django.http import JsonResponse
    from django.db.models import Q
    from core.models import Student
    import operator
    from functools import reduce
    
    mobile = request.GET.get("mobile", "").strip()
    father = request.GET.get("father", "").strip()
    
    if not mobile and len(father) < 3:
        return JsonResponse({"siblings": []})
        
    query = Q(is_active=True)
    
    subqueries = []
    if mobile and len(mobile) >= 10:
        subqueries.append(Q(mobile_primary__icontains=mobile) | Q(mobile_secondary__icontains=mobile))
    if father and len(father) >= 3:
        subqueries.append(Q(father_name__iexact=father))
        
    if not subqueries:
        return JsonResponse({"siblings": []})
        
    query &= reduce(operator.or_, subqueries)
    
    exclude_id = request.GET.get("exclude_id")
    qs = Student.objects.filter(query)
    if exclude_id and exclude_id.isdigit():
        qs = qs.exclude(pk=exclude_id)
        
    siblings = []
    for s in qs[:5]:
        siblings.append({
            "id": s.id,
            "name": s.full_name,
            "class_name": s.current_class.name if s.current_class else "N/A",
            "section_name": s.current_section.name if s.current_section else "",
            "sid": s.legacy_sid or s.admission_no or str(s.id),
        })
        
    return JsonResponse({"siblings": siblings})


# ---------------------------------------------------------------------------
# Feeder / Attached Schools (अटैच्ड विद्यालय) Views
# ---------------------------------------------------------------------------

@login_required
def feeder_school_list(request):
    """Directory and overview of all attached/feeder schools."""
    from core.models import FeederSchool, Student, AccountGroup, LedgerAccount
    schools = FeederSchool.objects.all().prefetch_related('students', 'ledger_account')
    
    total_schools = schools.count()
    total_students = Student.objects.filter(is_active=True, feeder_school__isnull=False).count()
    
    school_data = []
    total_demand_all = Decimal("0.00")
    total_received_all = Decimal("0.00")
    total_balance_all = Decimal("0.00")
    
    for s in schools:
        s_enrolled = s.total_enrolled_students
        s_demand = s.total_demand
        s_received = s.total_received
        s_balance = s.balance_due
        
        total_demand_all += s_demand
        total_received_all += s_received
        total_balance_all += s_balance
        
        school_data.append({
            "school": s,
            "enrolled": s_enrolled,
            "demand": s_demand,
            "received": s_received,
            "balance": s_balance,
        })
        
    context = {
        "schools_data": school_data,
        "total_schools": total_schools,
        "total_students": total_students,
        "total_demand": total_demand_all,
        "total_received": total_received_all,
        "total_balance": total_balance_all,
    }
    return render(request, "core/feeder_school_list.html", context)


@login_required
def feeder_school_detail(request, pk):
    """Detailed profile, student roster, and ledger statement for a feeder school."""
    from core.models import FeederSchool, Voucher
    school = get_object_or_404(FeederSchool, pk=pk)
    students = school.students.filter(is_active=True).select_related("current_class", "current_section").order_by("current_class__display_order", "full_name")
    
    q = request.GET.get("q", "").strip()
    if q:
        students = students.filter(
            Q(full_name__icontains=q) |
            Q(father_name__icontains=q) |
            Q(admission_no__icontains=q) |
            Q(legacy_sid__icontains=q)
        )
        
    vouchers = []
    if school.ledger_account:
        vouchers = Voucher.objects.filter(
            credit_account=school.ledger_account,
            is_cancelled=False
        ).select_related("debit_account").order_by("-voucher_date", "-id")
        
    context = {
        "school": school,
        "students": students,
        "vouchers": vouchers,
        "enrolled_count": school.total_enrolled_students,
        "total_demand": school.total_demand,
        "total_received": school.total_received,
        "balance_due": school.balance_due,
        "q": q,
    }
    return render(request, "core/feeder_school_detail.html", context)


@login_required
def feeder_school_create(request):
    """Add a new attached/feeder school."""
    from core.forms import FeederSchoolForm
    from core.models import AccountGroup, LedgerAccount, FeederSchool
    if request.method == "POST":
        form = FeederSchoolForm(request.POST)
        if form.is_valid():
            school = form.save(commit=False)
            debtors_group, _ = AccountGroup.objects.get_or_create(
                name="Sundry Debtors",
                defaults={"group_type": AccountGroup.GroupType.ASSET, "display_order": 15}
            )
            ledger, _ = LedgerAccount.objects.get_or_create(
                name=f"{school.name} A/C",
                defaults={"group": debtors_group, "opening_balance": Decimal("0.00")}
            )
            school.ledger_account = ledger
            school.save()
            messages.success(request, f"Attached school '{school.name}' successfully added with Sundry Debtors ledger.")
            return redirect("core:feeder_school_detail", pk=school.pk)
    else:
        form = FeederSchoolForm()
        
    return render(request, "core/feeder_school_form.html", {"form": form, "title": "Add New Attached School (नया अटैच्ड स्कूल जोड़ें)"})


@login_required
def feeder_school_edit(request, pk):
    """Edit an existing feeder school."""
    from core.forms import FeederSchoolForm
    from core.models import FeederSchool
    school = get_object_or_404(FeederSchool, pk=pk)
    if request.method == "POST":
        form = FeederSchoolForm(request.POST, instance=school)
        if form.is_valid():
            form.save()
            messages.success(request, f"Attached school '{school.name}' updated successfully.")
            return redirect("core:feeder_school_detail", pk=school.pk)
    else:
        form = FeederSchoolForm(instance=school)
        
    return render(request, "core/feeder_school_form.html", {"form": form, "school": school, "title": f"Edit School: {school.name}"})


@login_required
def feeder_school_payment(request, pk):
    """Record payment received from a feeder school."""
    from core.forms import FeederSchoolPaymentForm
    from core.models import FeederSchool, AccountGroup, LedgerAccount, AcademicSession, Voucher, VoucherCounter
    school = get_object_or_404(FeederSchool, pk=pk)
    if not school.ledger_account:
        debtors_group, _ = AccountGroup.objects.get_or_create(
            name="Sundry Debtors",
            defaults={"group_type": AccountGroup.GroupType.ASSET, "display_order": 15}
        )
        ledger, _ = LedgerAccount.objects.get_or_create(
            name=f"{school.name} A/C",
            defaults={"group": debtors_group, "opening_balance": Decimal("0.00")}
        )
        school.ledger_account = ledger
        school.save(update_fields=["ledger_account"])

    session = AcademicSession.objects.filter(is_active=True).first()
    if not session:
        session = AcademicSession.objects.first()

    if request.method == "POST":
        form = FeederSchoolPaymentForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            p_date = cd["payment_date"]
            cash_bank_acc = cd["cash_or_bank_account"]
            amount = cd["amount"]
            mode = cd["payment_mode"]
            ref_no = cd["reference_no"]
            narration = cd["narration"] or f"Affiliation / Registration fee installment from {school.name}"
            
            v_type = Voucher.VoucherType.CASH_RECEIPT if mode == Voucher.PaymentMode.CASH else Voucher.VoucherType.BANK
            
            with transaction.atomic():
                counter, _ = VoucherCounter.objects.select_for_update().get_or_create(
                    session=session, voucher_type=v_type
                )
                counter.last_number += 1
                counter.save()
                v_num_str = f"{v_type}-{session.name[:4] if session else '2026'}-{counter.last_number:04d}"
                
                voucher = Voucher.objects.create(
                    voucher_no=v_num_str,
                    voucher_type=v_type,
                    session=session,
                    voucher_date=p_date,
                    debit_account=cash_bank_acc,
                    credit_account=school.ledger_account,
                    amount=amount,
                    paid_to_or_received_from=school.name,
                    narration=narration,
                    payment_mode=mode,
                    physical_slip_no=ref_no,
                )
                
            messages.success(request, f"Payment of Rs. {amount:,.2f} recorded successfully! Voucher #{voucher.voucher_no}")
            return redirect("core:feeder_school_detail", pk=school.pk)
    else:
        form = FeederSchoolPaymentForm()
        
    context = {
        "school": school,
        "form": form,
        "balance_due": school.balance_due,
    }
    return render(request, "core/feeder_school_payment.html", context)


@login_required
def feeder_school_statement_pdf(request, pk):
    """Generate professional A4 PDF statement for an attached school."""
    from core.models import FeederSchool, Voucher, SchoolProfile
    from core.pdf import build_feeder_school_statement_pdf
    school = get_object_or_404(FeederSchool, pk=pk)
    students = school.students.filter(is_active=True).select_related("current_class", "current_section").order_by("current_class__display_order", "full_name")
    
    vouchers = []
    if school.ledger_account:
        vouchers = list(Voucher.objects.filter(
            credit_account=school.ledger_account,
            is_cancelled=False
        ).select_related("debit_account").order_by("voucher_date", "id"))
        
    profile = SchoolProfile.objects.first()
    pdf_bytes = build_feeder_school_statement_pdf(school, students, vouchers, school_profile=profile)
    
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{school.code or "SCHOOL"}_Statement.pdf"'
    return response


@login_required
def feeder_school_statement_excel(request, pk):
    """Export Feeder School statement & roster to Excel."""
    from core.models import FeederSchool
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    school = get_object_or_404(FeederSchool, pk=pk)
    students = school.students.filter(is_active=True).select_related("current_class", "current_section").order_by("current_class__display_order", "full_name")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "School Statement"
    
    ws.merge_cells("A1:G1")
    t_cell = ws["A1"]
    t_cell.value = "THPS INTERMEDIATE COLLEGE - ATTACHED SCHOOL STATEMENT"
    t_cell.font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
    t_cell.alignment = Alignment(horizontal="center")
    
    ws.merge_cells("A2:G2")
    ws["A2"].value = f"School: {school.name} | Contact: {school.contact_person} ({school.phone}) | Location: {school.village_address}"
    ws["A2"].font = Font(name="Calibri", size=11, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")
    
    ws.append([])
    
    ws.append(["Enrolled Students", "Package Rate / Student", "Total Demand (Rs.)", "Total Received (Rs.)", "Balance Due (Rs.)"])
    ws.append([school.total_enrolled_students, float(school.package_rate_per_student), float(school.total_demand), float(school.total_received), float(school.balance_due)])
    
    ws.append([])
    ws.append(["#", "Adm No / SID", "Student Name", "Father Name", "Class", "Section", "Mobile"])
    
    for idx, s in enumerate(students, 1):
        ws.append([
            idx,
            s.admission_no or s.legacy_sid or "",
            s.full_name,
            s.father_name,
            str(s.current_class or ""),
            str(s.current_section.name if s.current_section else ""),
            s.mobile_primary or s.mobile_secondary or ""
        ])
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{school.code or "SCHOOL"}_Statement.xlsx"'
    wb.save(response)
    return response


# ---------------------------------------------------------------------------
# Attendance Register (उपस्थिति पंजिका) Views
# ---------------------------------------------------------------------------

@login_required
def attendance_register_view(request):
    """Interactive preview and filter bar for Student Monthly Attendance Register."""
    import calendar
    from datetime import date
    from core.models import SchoolClass, Section, Student, AcademicSession, SchoolProfile

    today = date.today()
    selected_month = int(request.GET.get("month", today.month))
    selected_year = int(request.GET.get("year", today.year))
    
    classes = SchoolClass.objects.all().order_by("display_order")
    selected_class_id = request.GET.get("class_id")
    
    selected_class = None
    if selected_class_id:
        selected_class = SchoolClass.objects.filter(pk=selected_class_id).first()
    if not selected_class and classes.exists():
        selected_class = classes.first()

    sections = Section.objects.filter(school_class=selected_class).order_by("name") if selected_class else Section.objects.all().order_by("name")
    selected_section_id = request.GET.get("section_id")
    selected_section = None
    if selected_section_id:
        selected_section = Section.objects.filter(pk=selected_section_id).first()

    students = []
    if selected_class:
        qs = Student.objects.filter(is_active=True, current_class=selected_class)
        if selected_section:
            qs = qs.filter(current_section=selected_section)
        students = qs.select_related("current_class", "current_section").order_by("roll_no", "full_name")

    num_days = calendar.monthrange(selected_year, selected_month)[1]
    days_list = []
    day_abbrs = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    for d in range(1, num_days + 1):
        w = calendar.weekday(selected_year, selected_month, d)
        days_list.append({
            "day": d,
            "weekday": day_abbrs[w],
            "is_sunday": (w == 6),
        })

    months_choices = [(m, calendar.month_name[m]) for m in range(1, 13)]
    years_choices = [2025, 2026, 2027]

    context = {
        "classes": classes,
        "sections": sections,
        "selected_class": selected_class,
        "selected_section": selected_section,
        "selected_month": selected_month,
        "selected_month_name": calendar.month_name[selected_month],
        "selected_year": selected_year,
        "months_choices": months_choices,
        "years_choices": years_choices,
        "days_list": days_list,
        "students": students,
        "total_students": len(students),
    }
    return render(request, "core/attendance_register.html", context)


@login_required
def attendance_register_pdf(request):
    """Stream printable A4 Landscape Monthly Attendance Register PDF."""
    import calendar
    from datetime import date
    from core.models import SchoolClass, Section, Student, AcademicSession, SchoolProfile
    from core.pdf import build_attendance_register_pdf

    today = date.today()
    month = int(request.GET.get("month", today.month))
    year = int(request.GET.get("year", today.year))
    
    class_id = request.GET.get("class_id")
    school_class = get_object_or_404(SchoolClass, pk=class_id) if class_id else SchoolClass.objects.first()
    
    section_id = request.GET.get("section_id")
    section = Section.objects.filter(pk=section_id).first() if section_id else None

    qs = Student.objects.filter(is_active=True, current_class=school_class)
    if section:
        qs = qs.filter(current_section=section)
    students = list(qs.select_related("current_class", "current_section").order_by("roll_no", "full_name"))

    session = AcademicSession.objects.filter(is_active=True).first() or AcademicSession.objects.first()
    school_profile = SchoolProfile.objects.first()

    pdf_bytes = build_attendance_register_pdf(
        students=students,
        school_class=school_class,
        section=section,
        month=month,
        year=year,
        session=session,
        school_profile=school_profile,
    )

    sec_name = f"_{section.name}" if section else ""
    filename = f"Attendance_{school_class.name}{sec_name}_{calendar.month_name[month]}_{year}.pdf".replace(" ", "_")
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# Attendance Summary Entry & Reports (Part B)
# ---------------------------------------------------------------------------

@login_required
def attendance_summary_entry(request):
    """Batch entry screen to record monthly working days and present days for a class/section."""
    import calendar
    from datetime import date
    from core.models import SchoolClass, Section, Student, AcademicSession, AttendanceSummary

    today = date.today()
    selected_month = int(request.GET.get("month", today.month))
    selected_year = int(request.GET.get("year", today.year))

    classes = SchoolClass.objects.all().order_by("display_order")
    selected_class_id = request.GET.get("class_id")
    selected_class = None
    if selected_class_id:
        selected_class = SchoolClass.objects.filter(pk=selected_class_id).first()
    if not selected_class and classes.exists():
        selected_class = classes.first()

    sections = Section.objects.filter(school_class=selected_class).order_by("name") if selected_class else Section.objects.all().order_by("name")
    selected_section_id = request.GET.get("section_id")
    selected_section = None
    if selected_section_id:
        selected_section = Section.objects.filter(pk=selected_section_id).first()

    session = AcademicSession.objects.filter(is_active=True).first() or AcademicSession.objects.first()

    students = []
    if selected_class:
        qs = Student.objects.filter(is_active=True, current_class=selected_class)
        if selected_section:
            qs = qs.filter(current_section=selected_section)
        students = list(qs.select_related("current_class", "current_section").order_by("roll_no", "full_name"))

    # Load existing summaries for these students in this session/month/year
    existing_summaries = {}
    if students:
        student_ids = [s.id for s in students]
        summaries = AttendanceSummary.objects.filter(
            student_id__in=student_ids,
            session=session,
            year=selected_year,
            month=selected_month,
        )
        for summ in summaries:
            existing_summaries[summ.student_id] = summ

    # Default working days (if previously saved by any student in batch, use that, else default to 24)
    default_working_days = 24
    if existing_summaries:
        first_summ = next(iter(existing_summaries.values()))
        default_working_days = first_summ.total_working_days

    if request.method == "POST":
        post_class_id = request.POST.get("class_id")
        post_section_id = request.POST.get("section_id")
        post_month = int(request.POST.get("month", selected_month))
        post_year = int(request.POST.get("year", selected_year))
        
        try:
            batch_working_days = int(request.POST.get("batch_working_days", default_working_days))
        except (ValueError, TypeError):
            batch_working_days = 24

        if batch_working_days < 1:
            messages.error(request, "Total Working Days must be at least 1.")
            return redirect(f"{request.path}?class_id={selected_class.id if selected_class else ''}&section_id={selected_section.id if selected_section else ''}&month={post_month}&year={post_year}")

        saved_count = 0
        validation_errors = []

        with transaction.atomic():
            for student in students:
                p_days_key = f"present_days_{student.id}"
                w_days_key = f"working_days_{student.id}"
                rem_key = f"remarks_{student.id}"

                p_days_raw = request.POST.get(p_days_key, "").strip()
                if p_days_raw == "":
                    continue  # Skip unentered rows

                try:
                    p_days = int(p_days_raw)
                except ValueError:
                    validation_errors.append(f"{student.full_name}: Invalid number of present days.")
                    continue

                try:
                    w_days = int(request.POST.get(w_days_key, batch_working_days))
                except (ValueError, TypeError):
                    w_days = batch_working_days

                if p_days < 0:
                    validation_errors.append(f"{student.full_name}: Present days cannot be negative.")
                    continue
                if p_days > w_days:
                    validation_errors.append(f"{student.full_name}: Present days ({p_days}) cannot exceed Total Working Days ({w_days}).")
                    continue

                remarks = request.POST.get(rem_key, "").strip()

                AttendanceSummary.objects.update_or_create(
                    student=student,
                    session=session,
                    year=post_year,
                    month=post_month,
                    defaults={
                        "total_working_days": w_days,
                        "days_present": p_days,
                        "remarks": remarks,
                    }
                )
                saved_count += 1

        if validation_errors:
            for err in validation_errors[:5]:
                messages.error(request, err)
            if len(validation_errors) > 5:
                messages.error(request, f"...and {len(validation_errors) - 5} more validation errors.")

        if saved_count > 0:
            messages.success(request, f"Attendance successfully saved for {saved_count} students in {selected_class.name} ({calendar.month_name[post_month]} {post_year})!")
            return redirect(f"{reverse('core:attendance_summary_report')}?class_id={selected_class.id if selected_class else ''}&section_id={selected_section.id if selected_section else ''}&month={post_month}&year={post_year}")

    # Build student row objects for template
    student_rows = []
    for idx, s in enumerate(students, 1):
        summary = existing_summaries.get(s.id)
        p_days = summary.days_present if summary else ""
        w_days = summary.total_working_days if summary else default_working_days
        rem = summary.remarks if summary else ""
        student_rows.append({
            "index": idx,
            "student": s,
            "roll_no": s.roll_no or idx,
            "admission_no": s.admission_no or s.legacy_sid or "",
            "present_days": p_days,
            "working_days": w_days,
            "absent_days": summary.days_absent if summary else "",
            "percentage": summary.attendance_percentage if summary else "",
            "remarks": rem,
            "has_record": (summary is not None),
        })

    months_choices = [(m, calendar.month_name[m]) for m in range(1, 13)]
    years_choices = [2025, 2026, 2027]

    context = {
        "classes": classes,
        "sections": sections,
        "selected_class": selected_class,
        "selected_section": selected_section,
        "selected_month": selected_month,
        "selected_month_name": calendar.month_name[selected_month],
        "selected_year": selected_year,
        "default_working_days": default_working_days,
        "months_choices": months_choices,
        "years_choices": years_choices,
        "student_rows": student_rows,
        "total_students": len(student_rows),
    }
    return render(request, "core/attendance_summary_entry.html", context)


@login_required
def attendance_summary_report(request):
    """Comprehensive Monthly Attendance Summary Report with KPIs and statistics."""
    import calendar
    from datetime import date
    from core.models import SchoolClass, Section, Student, AcademicSession, AttendanceSummary

    today = date.today()
    selected_month = int(request.GET.get("month", today.month))
    selected_year = int(request.GET.get("year", today.year))

    classes = SchoolClass.objects.all().order_by("display_order")
    selected_class_id = request.GET.get("class_id")
    selected_class = None
    if selected_class_id:
        selected_class = SchoolClass.objects.filter(pk=selected_class_id).first()
    if not selected_class and classes.exists():
        selected_class = classes.first()

    sections = Section.objects.filter(school_class=selected_class).order_by("name") if selected_class else Section.objects.all().order_by("name")
    selected_section_id = request.GET.get("section_id")
    selected_section = None
    if selected_section_id:
        selected_section = Section.objects.filter(pk=selected_section_id).first()

    session = AcademicSession.objects.filter(is_active=True).first() or AcademicSession.objects.first()

    students = []
    if selected_class:
        qs = Student.objects.filter(is_active=True, current_class=selected_class)
        if selected_section:
            qs = qs.filter(current_section=selected_section)
        students = list(qs.select_related("current_class", "current_section").order_by("roll_no", "full_name"))

    summaries_map = {}
    if students:
        student_ids = [s.id for s in students]
        summaries = AttendanceSummary.objects.filter(
            student_id__in=student_ids,
            session=session,
            year=selected_year,
            month=selected_month,
        )
        for summ in summaries:
            summaries_map[summ.student_id] = summ

    report_rows = []
    total_present_sum = 0
    total_working_sum = 0
    short_attendance_count = 0
    marked_count = 0

    for idx, s in enumerate(students, 1):
        summ = summaries_map.get(s.id)
        if summ:
            marked_count += 1
            total_present_sum += summ.days_present
            total_working_sum += summ.total_working_days
            pct = summ.attendance_percentage
            if pct < 75.0:
                short_attendance_count += 1
            badge = "danger" if pct < 60.0 else "warning" if pct < 75.0 else "success"
            status_label = "Short Attendance (<60%)" if pct < 60.0 else "Average (60-74%)" if pct < 75.0 else "Eligible (>=75%)"
        else:
            pct = None
            badge = "secondary"
            status_label = "Not Recorded"

        report_rows.append({
            "index": idx,
            "student": s,
            "roll_no": s.roll_no or idx,
            "admission_no": s.admission_no or s.legacy_sid or "",
            "working_days": summ.total_working_days if summ else "-",
            "present_days": summ.days_present if summ else "-",
            "absent_days": summ.days_absent if summ else "-",
            "percentage": pct,
            "badge": badge,
            "status_label": status_label,
            "remarks": summ.remarks if summ else "",
            "is_marked": (summ is not None),
        })

    avg_percentage = round((total_present_sum / total_working_sum * 100), 1) if total_working_sum > 0 else 0.0

    months_choices = [(m, calendar.month_name[m]) for m in range(1, 13)]
    years_choices = [2025, 2026, 2027]

    context = {
        "classes": classes,
        "sections": sections,
        "selected_class": selected_class,
        "selected_section": selected_section,
        "selected_month": selected_month,
        "selected_month_name": calendar.month_name[selected_month],
        "selected_year": selected_year,
        "months_choices": months_choices,
        "years_choices": years_choices,
        "report_rows": report_rows,
        "total_enrolled": len(students),
        "marked_count": marked_count,
        "avg_percentage": avg_percentage,
        "short_attendance_count": short_attendance_count,
    }
    return render(request, "core/attendance_summary_report.html", context)


@login_required
def attendance_summary_report_excel(request):
    """Export filtered monthly attendance report to Excel."""
    import calendar
    from datetime import date
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from core.models import SchoolClass, Section, Student, AcademicSession, AttendanceSummary

    today = date.today()
    selected_month = int(request.GET.get("month", today.month))
    selected_year = int(request.GET.get("year", today.year))

    class_id = request.GET.get("class_id")
    school_class = get_object_or_404(SchoolClass, pk=class_id) if class_id else SchoolClass.objects.first()

    section_id = request.GET.get("section_id")
    section = Section.objects.filter(pk=section_id).first() if section_id else None

    session = AcademicSession.objects.filter(is_active=True).first() or AcademicSession.objects.first()

    qs = Student.objects.filter(is_active=True, current_class=school_class)
    if section:
        qs = qs.filter(current_section=section)
    students = list(qs.select_related("current_class", "current_section").order_by("roll_no", "full_name"))

    summaries_map = {}
    if students:
        student_ids = [s.id for s in students]
        summaries = AttendanceSummary.objects.filter(
            student_id__in=student_ids,
            session=session,
            year=selected_year,
            month=selected_month,
        )
        for summ in summaries:
            summaries_map[summ.student_id] = summ

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Attendance"

    ws.merge_cells("A1:H1")
    ws["A1"] = "THPS INTERMEDIATE COLLEGE - MONTHLY ATTENDANCE REPORT"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
    ws["A1"].alignment = Alignment(horizontal="center")

    sec_name = f" - Section {section.name}" if section else ""
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Class: {school_class.name}{sec_name} | Month: {calendar.month_name[selected_month]} {selected_year} | Session: {session.name if session else '2026-27'}"
    ws["A2"].font = Font(name="Calibri", size=11, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.append([])
    headers = ["# / Roll", "Adm No", "Student Name", "Father Name", "Working Days", "Present Days", "Absent Days", "Attendance %", "Status", "Remarks"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if col_idx in [1, 2, 5, 6, 7, 8] else "left")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    row_num = 5
    for idx, s in enumerate(students, 1):
        summ = summaries_map.get(s.id)
        w_days = summ.total_working_days if summ else ""
        p_days = summ.days_present if summ else ""
        a_days = summ.days_absent if summ else ""
        pct = f"{summ.attendance_percentage}%" if summ else "Not Recorded"
        status = "Eligible (>=75%)" if (summ and summ.attendance_percentage >= 75.0) else "Average (60-74%)" if (summ and summ.attendance_percentage >= 60.0) else "Short Attendance (<60%)" if summ else "-"

        ws.append([
            s.roll_no or idx,
            s.admission_no or s.legacy_sid or "",
            s.full_name,
            s.father_name,
            w_days,
            p_days,
            a_days,
            pct,
            status,
            summ.remarks if summ else ""
        ])
        for col_idx in range(1, 11):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            if col_idx in [1, 2, 5, 6, 7, 8]:
                cell.alignment = Alignment(horizontal="center")
        row_num += 1

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    filename = f"Attendance_{school_class.name}{sec_name}_{calendar.month_name[selected_month]}_{selected_year}.xlsx".replace(" ", "_")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
