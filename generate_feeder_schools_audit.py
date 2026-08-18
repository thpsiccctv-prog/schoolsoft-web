import os
import csv
from decimal import Decimal
from collections import defaultdict
from pathlib import Path

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "schoolsoft.settings")
sqlite_path = os.environ.get("SCHOOLSOFT_SQLITE_PATH") or os.path.expandvars(r"%LOCALAPPDATA%\THPSIC-InterCollege-SchoolSoft\db.sqlite3")
os.environ["SCHOOLSOFT_SQLITE_PATH"] = sqlite_path

import django
django.setup()

from core.models import Student
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

csv.field_size_limit(2147483647)
addmission_35 = r"E:\THPSIC-INTER-COLLEGE\05-reports\access-audit-raw\school7-comp35\csv-for-analysis\ADDMISSION.csv"

# Map active students
active_students = {str(s.legacy_sid): s for s in Student.objects.filter(is_active=True).select_related('current_class', 'current_section') if s.legacy_sid}

feeder_students = defaultdict(list)

with open(addmission_35, encoding="utf-8-sig", errors="ignore") as f:
    for row in csv.DictReader(f):
        sid = str(row.get('sid') or '').strip()
        if sid in active_students:
            s = active_students[sid]
            sch = row.get('sch_name', '').strip()
            if not sch:
                if s.current_section and s.current_section.name in ['B', 'C']:
                    sch = "OTHER SECTION B/C (UNASSIGNED)"
                else:
                    sch = "REGULAR THPSIC (SECTION A/D)"
            
            feeder_students[sch].append({
                "sid": sid,
                "admission_no": s.admission_no or "",
                "student_name": s.full_name or "",
                "father_name": s.father_name or "",
                "class": str(s.current_class or ""),
                "section": str(s.current_section.name if s.current_section else ""),
                "feeder_school": sch,
            })

# Generate summary rows
summary_rows = []
default_rate = Decimal("1500.00")

for sch, students in sorted(feeder_students.items(), key=lambda x: -len(x[1])):
    is_attached = sch not in ["REGULAR THPSIC (SECTION A/D)"]
    cnt = len(students)
    rate = default_rate if is_attached else Decimal("0.00")
    total_bill = cnt * rate
    
    summary_rows.append({
        "school_name": sch,
        "is_attached": "YES (Attached Feeder School)" if is_attached else "NO (Regular)",
        "student_count": cnt,
        "package_rate": float(rate),
        "total_receivable": float(total_bill),
    })

# Save Excel
out_xlsx = Path(r"E:\THPSIC-INTER-COLLEGE\05-reports\ATTACHED_FEEDER_SCHOOLS_AUDIT.xlsx")
out_csv = Path(r"E:\THPSIC-INTER-COLLEGE\05-reports\ATTACHED_FEEDER_SCHOOLS_AUDIT.csv")

wb = openpyxl.Workbook()
ws_sum = wb.active
ws_sum.title = "Attached Schools Summary"

headers = ["Attached School Name", "Category", "Enrolled Students", "Package Rate / Student (Rs.)", "Total Billed Demand (Rs.)"]
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

ws_sum.append(headers)
for col_idx in range(1, len(headers) + 1):
    c = ws_sum.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border

for r_idx, r in enumerate(summary_rows, start=2):
    ws_sum.append(list(r.values()))
    for col_idx in range(1, len(headers) + 1):
        c = ws_sum.cell(row=r_idx, column=col_idx)
        c.border = thin_border
        if col_idx in [2, 3, 4]:
            c.alignment = Alignment(horizontal="center", vertical="center")

ws_sum.auto_filter.ref = ws_sum.dimensions

for col in ws_sum.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        val_str = str(cell.value or "")
        if len(val_str) > max_len: max_len = len(val_str)
    ws_sum.column_dimensions[col_letter].width = max(max_len + 3, 15)

# Sheet 2: Detailed Student List
ws_det = wb.create_sheet(title="Student Feeder Roster")
det_headers = ["SID", "Adm No", "Student Name", "Father Name", "Class", "Section", "Attached Feeder School"]
ws_det.append(det_headers)
for col_idx in range(1, len(det_headers) + 1):
    c = ws_det.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border

all_students_flat = [s for sch, students in feeder_students.items() for s in students]
all_students_flat.sort(key=lambda x: (x["feeder_school"], x["class"], x["student_name"]))

for r_idx, s in enumerate(all_students_flat, start=2):
    ws_det.append(list(s.values()))
    for col_idx in range(1, len(det_headers) + 1):
        c = ws_det.cell(row=r_idx, column=col_idx)
        c.border = thin_border
        if col_idx in [1, 2, 5, 6]:
            c.alignment = Alignment(horizontal="center", vertical="center")

ws_det.auto_filter.ref = ws_det.dimensions
for col in ws_det.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        val_str = str(cell.value or "")
        if len(val_str) > max_len: max_len = len(val_str)
    ws_det.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save(out_xlsx)
print(f"Generated:")
print(f"  Excel: {out_xlsx}")
