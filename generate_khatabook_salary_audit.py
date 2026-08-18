import os
import csv
from decimal import Decimal
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

salary_schedule = [
    # ASHOK SINGH (Code 14)
    {"code": 14, "name": "ASHOK SINGH", "desig": "LECTURER", "month": "April 2026", "date": "02/07/2026", "basic": 15000.0, "allowances": 1000.0, "amount": 16000.0, "remarks": "Salary of April + CL 1000"},
    {"code": 14, "name": "ASHOK SINGH", "desig": "LECTURER", "month": "May 2026", "date": "02/07/2026", "basic": 15000.0, "allowances": 1000.0, "amount": 16000.0, "remarks": "Salary of May + CL 1000"},
    {"code": 14, "name": "ASHOK SINGH", "desig": "LECTURER", "month": "July 2026", "date": "17/08/2026", "basic": 15000.0, "allowances": 1000.0, "amount": 16000.0, "remarks": "Salary of July + CL 1000"},
    
    # KM ANURADHA SINGH (Code 16)
    {"code": 16, "name": "KM ANURADHA SINGH", "desig": "LECTURER", "month": "April 2026", "date": "17/08/2026", "basic": 8000.0, "allowances": 0.0, "amount": 8000.0, "remarks": "Salary of April"},
    {"code": 16, "name": "KM ANURADHA SINGH", "desig": "LECTURER", "month": "May 2026", "date": "17/08/2026", "basic": 8000.0, "allowances": 0.0, "amount": 8000.0, "remarks": "Salary of May"},
    {"code": 16, "name": "KM ANURADHA SINGH", "desig": "LECTURER", "month": "July 2026", "date": "17/08/2026", "basic": 8000.0, "allowances": 0.0, "amount": 8000.0, "remarks": "Salary of July"},
    
    # HARIKESH YADAV (Code 12)
    {"code": 12, "name": "HARIKESH YADAV", "desig": "ENGLISH", "month": "April 2026", "date": "02/07/2026", "basic": 15467.0, "allowances": 0.0, "amount": 15467.0, "remarks": "Salary of April"},
    {"code": 12, "name": "HARIKESH YADAV", "desig": "ENGLISH", "month": "May 2026", "date": "02/07/2026", "basic": 16000.0, "allowances": 533.0, "amount": 16533.0, "remarks": "Salary of May"},
    {"code": 12, "name": "HARIKESH YADAV", "desig": "ENGLISH", "month": "July 2026", "date": "17/08/2026", "basic": 14934.0, "allowances": 0.0, "amount": 14934.0, "remarks": "Salary of July"},
    
    # JAIPRAKASH PRAJAPATI (Code 11)
    {"code": 11, "name": "JAIPRAKASH PRAJAPATI", "desig": "BIOLOGY AND CHEMISTRY TEACHER", "month": "July 2026", "date": "17/08/2026", "basic": 15000.0, "allowances": 500.0, "amount": 15500.0, "remarks": "Salary of July + 500 CL bonus"},
    
    # MANOJ KUMAR (Code 17)
    {"code": 17, "name": "MANOJ KUMAR", "desig": "ASSISTANT TEACHER", "month": "April 2026", "date": "02/07/2026", "basic": 4000.0, "allowances": 0.0, "amount": 4000.0, "remarks": "Salary of April"},
    {"code": 17, "name": "MANOJ KUMAR", "desig": "ASSISTANT TEACHER", "month": "May 2026", "date": "02/07/2026", "basic": 4000.0, "allowances": 0.0, "amount": 4000.0, "remarks": "Salary of May"},
    {"code": 17, "name": "MANOJ KUMAR", "desig": "ASSISTANT TEACHER", "month": "July 2026", "date": "17/08/2026", "basic": 4000.0, "allowances": 0.0, "amount": 4000.0, "remarks": "Salary of July (Office confirmed)"},
    
    # SARITA (Code 18)
    {"code": 18, "name": "SARITA", "desig": "PEON", "month": "April 2026", "date": "22/04/2026", "basic": 4000.0, "allowances": 0.0, "amount": 4000.0, "remarks": "Salary of April"},
    {"code": 18, "name": "SARITA", "desig": "PEON", "month": "May 2026", "date": "27/05/2026", "basic": 4000.0, "allowances": 0.0, "amount": 4000.0, "remarks": "Salary of May"},
    {"code": 18, "name": "SARITA", "desig": "PEON", "month": "July 2026", "date": "23/07/2026", "basic": 4000.0, "allowances": 0.0, "amount": 4000.0, "remarks": "Salary of July"},
    
    # SATYNARAYAN SINGH (Code 13)
    {"code": 13, "name": "SATYNARAYAN SINGH", "desig": "GATE MAN", "month": "July 2026", "date": "17/08/2026", "basic": 5000.0, "allowances": 0.0, "amount": 5000.0, "remarks": "Salary of July"},
    
    # VINOD YADAV (Code 19)
    {"code": 19, "name": "VINOD YADAV", "desig": "LECTURER", "month": "April 2026", "date": "02/07/2026", "basic": 17400.0, "allowances": 0.0, "amount": 17400.0, "remarks": "Salary of April"},
    {"code": 19, "name": "VINOD YADAV", "desig": "LECTURER", "month": "May 2026", "date": "02/07/2026", "basic": 18000.0, "allowances": 1200.0, "amount": 19200.0, "remarks": "Salary of May"},
    {"code": 19, "name": "VINOD YADAV", "desig": "LECTURER", "month": "July 2026", "date": "17/08/2026", "basic": 18000.0, "allowances": 0.0, "amount": 18000.0, "remarks": "Salary of July"},
]

out_xlsx = Path(r"E:\THPSIC-INTER-COLLEGE\05-reports\KHATABOOK_SALARY_STATEMENT_APRIL_JULY_2026.xlsx")
out_csv = Path(r"E:\THPSIC-INTER-COLLEGE\05-reports\KHATABOOK_SALARY_STATEMENT_APRIL_JULY_2026.csv")

# Save CSV
with open(out_csv, "w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=["code", "name", "desig", "month", "date", "basic", "allowances", "amount", "remarks"])
    writer.writeheader()
    writer.writerows(salary_schedule)

# Save Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Khatabook Salaries (Apr-Jul)"

headers = ["Emp Code", "Staff Name", "Designation", "Salary Month", "Payment Date", "Basic Pay (Rs.)", "Allowances/CL (Rs.)", "Amount Paid (Rs.)", "Remarks / Source"]
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

ws.append(headers)
for col_idx in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border

for r_idx, r in enumerate(salary_schedule, start=2):
    row_vals = [r["code"], r["name"], r["desig"], r["month"], r["date"], r["basic"], r["allowances"], r["amount"], r["remarks"]]
    ws.append(row_vals)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=r_idx, column=col_idx)
        c.border = thin_border
        if col_idx in [1, 4, 5]:
            c.alignment = Alignment(horizontal="center", vertical="center")

ws.auto_filter.ref = ws.dimensions

for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        val_str = str(cell.value or "")
        if len(val_str) > max_len: max_len = len(val_str)
    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save(out_xlsx)
print(f"Generated:")
print(f"  Excel: {out_xlsx}")
print(f"  CSV:   {out_csv}")
