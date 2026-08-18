import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "schoolsoft.settings")
sqlite_path = os.environ.get("SCHOOLSOFT_SQLITE_PATH") or os.path.expandvars(r"%LOCALAPPDATA%\THPSIC-InterCollege-SchoolSoft\db.sqlite3")
os.environ["SCHOOLSOFT_SQLITE_PATH"] = sqlite_path

import django
django.setup()
from core.models import FeederSchool, Student

wb = openpyxl.Workbook()

# Sheet 1: Summary
ws_sum = wb.active
ws_sum.title = "Attached Schools Summary"

title_font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
bold_font = Font(name="Calibri", size=11, bold=True)
regular_font = Font(name="Calibri", size=11)
thin_border = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)

ws_sum.merge_cells("A1:G1")
ws_sum["A1"] = "THPS INTERMEDIATE COLLEGE - ATTACHED / FEEDER SCHOOLS AUDIT REPORT"
ws_sum["A1"].font = title_font
ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 25

ws_sum.append([])
headers_sum = ["#", "Attached School Name", "Code", "Enrolled Students", "Package Rate / Stu", "Total Demand (Rs.)", "Total Paid (Rs.)", "Balance Due (Rs.)"]
ws_sum.append(headers_sum)

for col_num, h in enumerate(headers_sum, 1):
    c = ws_sum.cell(row=3, column=col_num)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center" if col_num in [1, 3, 4] else "right" if col_num >= 5 else "left")

row_idx = 4
total_stu = 0
total_dem = 0
total_paid = 0
total_bal = 0

for idx, s in enumerate(FeederSchool.objects.all().order_by("name"), 1):
    stu = s.total_enrolled_students
    dem = float(s.total_demand)
    paid = float(s.total_received)
    bal = float(s.balance_due)
    
    total_stu += stu
    total_dem += dem
    total_paid += paid
    total_bal += bal
    
    ws_sum.append([
        idx,
        s.name,
        s.code or "",
        stu,
        float(s.package_rate_per_student),
        dem,
        paid,
        bal
    ])
    for col_num in range(1, 9):
        cell = ws_sum.cell(row=row_idx, column=col_num)
        cell.font = regular_font
        cell.border = thin_border
        if col_num in [1, 3, 4]:
            cell.alignment = Alignment(horizontal="center")
        elif col_num >= 5:
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = "₹#,##0.00"
    row_idx += 1

# Grand Total Row
ws_sum.append(["", "GRAND TOTAL", "", total_stu, "", total_dem, total_paid, total_bal])
for col_num in range(1, 9):
    cell = ws_sum.cell(row=row_idx, column=col_num)
    cell.font = bold_font
    cell.border = thin_border
    if col_num == 4:
        cell.alignment = Alignment(horizontal="center")
    elif col_num >= 6:
        cell.alignment = Alignment(horizontal="right")
        cell.number_format = "₹#,##0.00"

for col in ws_sum.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws_sum.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Sheet 2: All 540 Students Roster
ws_stu = wb.create_sheet(title="All 540 Students Roster")
ws_stu.merge_cells("A1:H1")
ws_stu["A1"] = "THPS INTERMEDIATE COLLEGE - ATTACHED STUDENTS MASTER ROSTER (540 STUDENTS)"
ws_stu["A1"].font = title_font
ws_stu["A1"].alignment = Alignment(horizontal="center")
ws_stu.row_dimensions[1].height = 25

ws_stu.append([])
stu_headers = ["#", "Attached School", "Adm No / SID", "Student Name", "Father Name", "Class", "Section", "Mobile"]
ws_stu.append(stu_headers)

for col_num, h in enumerate(stu_headers, 1):
    c = ws_stu.cell(row=3, column=col_num)
    c.font = header_font
    c.fill = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
    c.alignment = Alignment(horizontal="center" if col_num in [1, 3, 6, 7] else "left")

all_students = Student.objects.filter(is_active=True, feeder_school__isnull=False).select_related(
    "feeder_school", "current_class", "current_section"
).order_by("feeder_school__name", "current_class__display_order", "full_name")

r_idx = 4
for idx, st in enumerate(all_students, 1):
    ws_stu.append([
        idx,
        st.feeder_school.name if st.feeder_school else "-",
        st.admission_no or st.legacy_sid or "",
        st.full_name,
        st.father_name,
        str(st.current_class or ""),
        str(st.current_section.name if st.current_section else ""),
        st.mobile_primary or st.mobile_secondary or ""
    ])
    for col_num in range(1, 9):
        cell = ws_stu.cell(row=r_idx, column=col_num)
        cell.font = regular_font
        cell.border = thin_border
        if col_num in [1, 3, 6, 7]:
            cell.alignment = Alignment(horizontal="center")
    r_idx += 1

for col in ws_stu.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws_stu.column_dimensions[col_letter].width = max(max_len + 4, 12)

out_path = r"E:\THPSIC-INTER-COLLEGE\05-reports\ATTACHED_FEEDER_SCHOOLS_AUDIT.xlsx"
wb.save(out_path)
print(f"Updated audit spreadsheet saved to {out_path}")
