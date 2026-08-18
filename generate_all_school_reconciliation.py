import os
import csv
from decimal import Decimal
from pathlib import Path
from collections import defaultdict, Counter

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "schoolsoft.settings")
sqlite_path = os.environ.get("SCHOOLSOFT_SQLITE_PATH") or os.path.expandvars(r"%LOCALAPPDATA%\THPSIC-InterCollege-SchoolSoft\db.sqlite3")
os.environ["SCHOOLSOFT_SQLITE_PATH"] = sqlite_path

import django
django.setup()

from core.models import Student, AcademicSession
from core.fee_engine import calculate_student_due
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

session = AcademicSession.objects.filter(is_active=True).first()

# Query ALL active students across the entire school (Classes VI to XII)
all_active_students = Student.objects.filter(is_active=True).select_related(
    'current_class', 'current_section'
).order_by('current_class__display_order', 'current_section__name', 'admission_no')

total_students = all_active_students.count()
print(f"Total Active Students across all classes (VI to XII): {total_students}")

rows = []
class_sec_summary = defaultdict(lambda: {
    "count": 0, "total_demand": Decimal('0.00'), "total_paid": Decimal('0.00'),
    "total_concession": Decimal('0.00'), "total_due": Decimal('0.00'), "zero_demand_count": 0
})

for s in all_active_students:
    res = calculate_student_due(student=s, session=session, through_month='AUG')
    
    cls_name = s.current_class.name if s.current_class else "Unknown"
    sec_name = s.current_section.name if s.current_section else "Unknown"
    class_sec_key = f"{cls_name} - {sec_name}"
    
    st_type = "New Admission (2026)" if res.is_new_student else "Promoted / Continuing"
    
    if res.due_amount == Decimal('0.00'):
        status = "Settled (Fully Paid)"
    elif res.received_amount > Decimal('0.00'):
        status = "Partially Paid"
    else:
        status = "Unpaid (Full Due)"
        
    class_sec_summary[class_sec_key]["count"] += 1
    class_sec_summary[class_sec_key]["total_demand"] += res.gross_demand
    class_sec_summary[class_sec_key]["total_paid"] += res.received_amount
    class_sec_summary[class_sec_key]["total_concession"] += res.concession_amount
    class_sec_summary[class_sec_key]["total_due"] += res.due_amount
    if res.gross_demand == Decimal('0.00'):
        class_sec_summary[class_sec_key]["zero_demand_count"] += 1
        
    rows.append({
        "sid": s.legacy_sid or "",
        "admission_no": s.admission_no or "",
        "student_name": s.full_name or "",
        "father_name": s.father_name or "",
        "class": cls_name,
        "section": sec_name,
        "student_type": st_type,
        "gross_demand": float(res.gross_demand),
        "paid_amount": float(res.received_amount),
        "concession_amount": float(res.concession_amount),
        "due_amount": float(res.due_amount),
        "status": status,
    })

# Output paths
output_csv = Path(r"E:\THPSIC-INTER-COLLEGE\05-reports\ALL_SCHOOL_FEE_RECONCILIATION_MASTER.csv")
output_xlsx = Path(r"E:\THPSIC-INTER-COLLEGE\05-reports\ALL_SCHOOL_FEE_RECONCILIATION_MASTER.xlsx")

# Write CSV
with open(output_csv, mode="w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)

# Write Multi-sheet Excel
wb = openpyxl.Workbook()

# Sheet 1: Summary by Class & Section
ws_sum = wb.active
ws_sum.title = "School Summary"

sum_headers = ["Class - Section", "Active Students", "Total Gross Demand", "Total Paid in StuFee", "Total Concessions", "Total Net Due", "Zero Demand Count"]
ws_sum.append(sum_headers)

header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

for col_idx in range(1, len(sum_headers) + 1):
    c = ws_sum.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border

for k, v in sorted(class_sec_summary.items()):
    ws_sum.append([
        k, v["count"], float(v["total_demand"]), float(v["total_paid"]),
        float(v["total_concession"]), float(v["total_due"]), v["zero_demand_count"]
    ])

# Sheet 2: All Students Detail
ws_det = wb.create_sheet(title="All Students Detail")
det_headers = ["SID", "Adm No", "Student Name", "Father Name", "Class", "Section", "Student Type", "Gross Demand (up to AUG)", "Paid in StuFee", "Concession", "Net Due Amount", "Status"]
ws_det.append(det_headers)

for col_idx in range(1, len(det_headers) + 1):
    c = ws_det.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border

for r in rows:
    ws_det.append(list(r.values()))

for ws in [ws_sum, ws_det]:
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len: max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save(output_xlsx)

print("\n--- SUMMARY OF ALL-SCHOOL RECONCILIATION ---")
total_school_demand = sum(v["total_demand"] for v in class_sec_summary.values())
total_school_paid = sum(v["total_paid"] for v in class_sec_summary.values())
total_school_due = sum(v["total_due"] for v in class_sec_summary.values())
total_zero_demands = sum(v["zero_demand_count"] for v in class_sec_summary.values())

print(f"Total Active Students: {total_students}")
print(f"Total Gross Demand (Through AUG 2026): Rs. {total_school_demand:,.2f}")
print(f"Total Paid Collected: Rs. {total_school_paid:,.2f}")
print(f"Total Net Due: Rs. {total_school_due:,.2f}")
print(f"Zero Demand Students across ENTIRE school: {total_zero_demands}")

print(f"\nGenerated Master Files:")
print(f"  Excel: {output_xlsx}")
print(f"  CSV:   {output_csv}")
