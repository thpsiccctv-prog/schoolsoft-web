import os
import csv
from decimal import Decimal
from pathlib import Path

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "schoolsoft.settings")
sqlite_path = os.environ.get("SCHOOLSOFT_SQLITE_PATH") or os.path.expandvars(r"%LOCALAPPDATA%\THPSIC-InterCollege-SchoolSoft\db.sqlite3")
os.environ["SCHOOLSOFT_SQLITE_PATH"] = sqlite_path

import django
django.setup()

from core.models import Student, StudentOpeningBalance, FeeReceipt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

csv.field_size_limit(2147483647)

# 1. Read initial FEE.csv
fee_initial = {}
with open(r"E:\THPSIC-INTER-COLLEGE\05-reports\access-audit-raw\school7-comp35\csv-for-analysis\FEE.csv", encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        sid = row.get('SID', '').strip()
        if not sid: continue
        curr = Decimal(row.get('CURR_AMT', '0') or '0')
        prv = Decimal(row.get('PRV_AMT', '0') or '0')
        tot = curr + prv
        if tot > 0:
            fee_initial[sid] = tot

# 2. Read StuFee.csv receipts
balance_paid_records = {}
balance_paid_total = {}

with open(r"E:\THPSIC-INTER-COLLEGE\05-reports\access-audit-raw\school7-comp35\csv-for-analysis\StuFee.csv", encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        sid = row.get('sid', '').strip()
        if not sid: continue
        month = (row.get('MONTH', '') or '').upper()
        due_fee = Decimal(row.get('DUE_FEE', '0') or '0')
        paid = Decimal(row.get('paid', '0') or '0')
        rcpno = row.get('rcpno', '')
        rdate = row.get('v_date', '')[:10]
        
        if 'BALANCE' in month or due_fee > 0:
            if sid not in balance_paid_records:
                balance_paid_records[sid] = []
                balance_paid_total[sid] = Decimal('0')
            balance_paid_records[sid].append(f"SF-{rcpno}: ₹{paid} ({rdate})")
            balance_paid_total[sid] += paid

rows = []
for sid, initial_due in fee_initial.items():
    s = Student.objects.filter(legacy_sid=sid).first()
    sname = s.full_name if s else ""
    fname = s.father_name if s else ""
    cls_name = str(s.current_class) if (s and s.current_class) else ""
    sec_name = str(s.current_section) if (s and s.current_section) else ""
    
    b_paid = balance_paid_total.get(sid, Decimal('0'))
    b_records = " | ".join(balance_paid_records.get(sid, []))
    rem_due = max(Decimal('0'), initial_due - b_paid)
    
    if b_paid >= initial_due:
        category = "FULLY CLEARED (Paid in Old App)"
        action = "Set Opening Balance to ₹0 (Double-Charge Fix)"
    elif b_paid > 0:
        category = "PARTIALLY CLEARED"
        action = f"Reduce Opening Balance to ₹{rem_due}"
    else:
        category = "UNPAID (Genuine Due)"
        action = "Keep Initial Opening Balance"
        
    rows.append({
        "sid": sid,
        "student_name": sname,
        "father_name": fname,
        "class": cls_name,
        "section": sec_name,
        "initial_fee_csv_balance": float(initial_due),
        "balance_fee_paid_in_stufee": float(b_paid),
        "reconciled_true_opening_due": float(rem_due),
        "category": category,
        "action_recommended": action,
        "receipt_details": b_records,
    })

rows.sort(key=lambda x: (x["category"], -x["initial_fee_csv_balance"]))

# Write Excel
output_xlsx = Path(r"E:\THPSIC-INTER-COLLEGE\05-reports\OPENING_BALANCE_RECONCILIATION_AUDIT.xlsx")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Opening Balance Reconciliation"

headers = [
    "SID", "Student Name", "Father Name", "Class", "Section",
    "Initial Snapshot (FEE.csv)", "Balance Paid (StuFee.csv)", "True Reconciled Opening Due",
    "Category", "Recommended Action", "Receipt Audit Trail"
]

header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

ws.append(headers)
for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

for r_idx, r in enumerate(rows, start=2):
    row_vals = [
        r["sid"], r["student_name"], r["father_name"], r["class"], r["section"],
        r["initial_fee_csv_balance"], r["balance_fee_paid_in_stufee"], r["reconciled_true_opening_due"],
        r["category"], r["action_recommended"], r["receipt_details"]
    ]
    ws.append(row_vals)
    
    cat = r["category"]
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=r_idx, column=col_idx)
        cell.border = thin_border
        if "FULLY CLEARED" in cat:
            cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Soft Green
        elif "PARTIALLY" in cat:
            cell.fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid") # Soft Yellow
        else:
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

ws.auto_filter.ref = ws.dimensions

for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        val_str = str(cell.value or "")
        if len(val_str) > max_len: max_len = len(val_str)
    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save(output_xlsx)
print(f"Opening Balance Reconciliation Audit Workbook generated: {output_xlsx}")
