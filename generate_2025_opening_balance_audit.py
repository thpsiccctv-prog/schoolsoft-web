import os
import csv
from decimal import Decimal
from pathlib import Path

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "schoolsoft.settings")
sqlite_path = os.environ.get("SCHOOLSOFT_SQLITE_PATH") or os.path.expandvars(r"%LOCALAPPDATA%\THPSIC-InterCollege-SchoolSoft\db.sqlite3")
os.environ["SCHOOLSOFT_SQLITE_PATH"] = sqlite_path

import django
django.setup()

from core.models import Student, AcademicSession, StudentOpeningBalance
from core.fee_engine import calculate_student_due
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

session = AcademicSession.objects.filter(is_active=True).first()

# Read 2025-2026 StuFee from Folder 33
csv.field_size_limit(2147483647)
stufee_33 = r"E:\THPSIC-INTER-COLLEGE\05-reports\access-audit-raw\school7-comp33\csv-for-analysis\StuFee.csv"

students_2025_dues = {}
with open(stufee_33, encoding="utf-8-sig", errors="ignore") as f:
    for r in csv.DictReader(f):
        sid = str(r.get('sid') or '').strip()
        if not sid: continue
        try:
            rcp_no = int(float(r.get('rcpno') or 0))
            due_amt = Decimal(str(r.get('due') or '0').strip())
        except (ValueError, TypeError):
            continue
        v_date = r.get('v_date', '')
        sname = r.get('sname', '')
        sclass = r.get('sclass', '')
        
        if sid not in students_2025_dues or rcp_no > students_2025_dues[sid]['last_rcp_no']:
            students_2025_dues[sid] = {
                'sid': sid,
                'name': sname,
                'class_2025': sclass,
                'last_rcp_no': rcp_no,
                'last_date': v_date[:10],
                'closing_due_2025': due_amt,
            }

print(f"Loaded {len(students_2025_dues)} students from 2025-2026 session.")

# Match with active students in current database
active_students = Student.objects.filter(is_active=True).select_related('current_class', 'current_section')

rows = []
matched_count = 0
total_proposed_carryover = Decimal('0.00')

for s in active_students:
    sid = str(s.legacy_sid or s.admission_no or "").strip()
    history = students_2025_dues.get(sid)
    
    closing_2025 = history['closing_due_2025'] if history else Decimal('0.00')
    last_rcp = history['last_rcp_no'] if history else "-"
    last_date = history['last_date'] if history else "-"
    class_2025 = history['class_2025'] if history else "-"
    
    if closing_2025 > Decimal('0.00'):
        matched_count += 1
        total_proposed_carryover += closing_2025
        
    res_current = calculate_student_due(student=s, session=session, through_month='AUG')
    
    # Calculate proposed new due
    proposed_gross = res_current.scheduled_fee_demand + closing_2025
    proposed_due = max(Decimal('0.00'), proposed_gross - res_current.received_amount - res_current.concession_amount)
    
    cls_curr = s.current_class.name if s.current_class else ""
    sec_curr = s.current_section.name if s.current_section else ""
    
    rows.append({
        "sid": sid,
        "admission_no": s.admission_no or "",
        "student_name": s.full_name or "",
        "father_name": s.father_name or "",
        "class_2025": class_2025,
        "current_class": cls_curr,
        "current_section": sec_curr,
        "last_rcp_2025": last_rcp,
        "last_date_2025": last_date,
        "proposed_opening_balance_2025": float(closing_2025),
        "current_2026_scheduled_demand": float(res_current.scheduled_fee_demand),
        "current_2026_paid": float(res_current.received_amount),
        "current_2026_concession": float(res_current.concession_amount),
        "current_due_without_carryover": float(res_current.due_amount),
        "proposed_new_total_due": float(proposed_due),
    })

print(f"Active Students Matched with Non-Zero 2025 Closing Due: {matched_count}")
print(f"Total Proposed Carryover Amount: Rs. {total_proposed_carryover:,.2f}")

# Sort rows: non-zero opening balances on top
rows.sort(key=lambda r: (-r["proposed_opening_balance_2025"], r["current_class"], r["student_name"]))

# Save CSV
output_csv = Path(r"E:\THPSIC-INTER-COLLEGE\05-reports\2025_2026_OPENING_BALANCE_MIGRATION_AUDIT.csv")
output_xlsx = Path(r"E:\THPSIC-INTER-COLLEGE\05-reports\2025_2026_OPENING_BALANCE_MIGRATION_AUDIT.xlsx")

with open(output_csv, mode="w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)

# Save Styled Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "2025-26 Opening Due Audit"

headers = [
    "SID", "Adm No", "Student Name", "Father Name", "Class (2025-26)", "Current Class (2026-27)",
    "Section", "Last Rcp (2025-26)", "Last Date (2025-26)", "2025-26 Closing Due (Opening Balance)",
    "Current 2026-27 Demand (up to AUG)", "Current Paid (2026-27)", "Concession",
    "Current Due (Without Carryover)", "Proposed True Total Due"
]

header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

ws.append(headers)
for col_idx in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border

for r_idx, r in enumerate(rows, start=2):
    ws.append(list(r.values()))
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=r_idx, column=col_idx)
        c.border = thin_border
        if col_idx in [1, 2, 5, 6, 7, 8, 9]:
            c.alignment = Alignment(horizontal="center", vertical="center")

ws.auto_filter.ref = ws.dimensions

for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        val_str = str(cell.value or "")
        if len(val_str) > max_len: max_len = len(val_str)
    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save(output_xlsx)
print(f"\nGenerated Files:")
print(f"  Excel: {output_xlsx}")
print(f"  CSV:   {output_csv}")
