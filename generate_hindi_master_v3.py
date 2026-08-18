import os
import sys
import csv
import re
from pathlib import Path

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "schoolsoft.settings")
sqlite_path = os.environ.get("SCHOOLSOFT_SQLITE_PATH") or os.path.expandvars(r"%LOCALAPPDATA%\THPSIC-InterCollege-SchoolSoft\db.sqlite3")
os.environ["SCHOOLSOFT_SQLITE_PATH"] = sqlite_path

import django
django.setup()

from core.models import Student
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Comprehensive 100% Offline Purvanchal / UP Board Hindi Name Dictionary
DICTIONARY = {
    # Titles / prefixes / suffixes / honorifics
    "SHRI": ("श्री", "HIGH", ""),
    "SHREE": ("श्री", "HIGH", ""),
    "SMT": ("श्रीमती", "HIGH", ""),
    "SHRIMATI": ("श्रीमती", "HIGH", ""),
    "KM": ("कु०", "HIGH", "Kumari abbreviation"),
    "KUMARI": ("कुमारी", "HIGH", ""),
    "KUMAR": ("कुमार", "HIGH", ""),
    "SINGH": ("सिंह", "HIGH", ""),
    "DEVI": ("देवी", "HIGH", ""),
    "PRASAD": ("प्रसाद", "HIGH", ""),
    "GUPTA": ("गुप्ता", "HIGH", ""),
    "GUPT": ("गुप्त", "HIGH", ""),
    "YADAV": ("यादव", "HIGH", ""),
    "VERMA": ("वर्मा", "HIGH", ""),
    "SHARMA": ("शर्मा", "HIGH", ""),
    "PATEL": ("पटेल", "HIGH", ""),
    "MISHRA": ("मिश्रा", "HIGH", ""),
    "PANDEY": ("पाण्डेय", "HIGH", ""),
    "PANDAY": ("पाण्डेय", "HIGH", ""),
    "TIWARI": ("तिवारी", "HIGH", ""),
    "SHUKLA": ("शुक्ला", "HIGH", ""),
    "SRIVASTAVA": ("श्रीवास्तव", "HIGH", ""),
    "SHRIVASTAVA": ("श्रीवास्तव", "HIGH", ""),
    "CHAUHAN": ("चौहान", "HIGH", ""),
    "MAURYA": ("मौर्य", "HIGH", ""),
    "KUSHWAHA": ("कुशवाहा", "HIGH", ""),
    "BIND": ("बिंद", "HIGH", ""),
    "GOND": ("गोंड", "HIGH", ""),
    "GAUD": ("गौड़", "HIGH", ""),
    "PASWAN": ("पासवान", "HIGH", ""),
    "PASVAN": ("पासवान", "HIGH", "Spelling variant Pasvan -> पासवान"),
    "SAHANI": ("साहनी", "HIGH", ""),
    "SAHNI": ("साहनी", "HIGH", ""),
    "SAH": ("साह", "HIGH", ""),
    "LAL": ("लाल", "HIGH", ""),
    "NATH": ("नाथ", "HIGH", ""),
    "CHAND": ("चंद", "HIGH", ""),
    "SHANKAR": ("शंकर", "HIGH", ""),
    "NARAYAN": ("नारायण", "HIGH", ""),
    "PRAKASH": ("प्रकाश", "HIGH", ""),
    "RAJ": ("राज", "HIGH", ""),
    "RAJA": ("राजा", "HIGH", ""),
    "MANI": ("मणि", "HIGH", ""),
    "KANT": ("कांत", "HIGH", ""),
    "RAM": ("राम", "HIGH", ""),
    "RAO": ("राव", "HIGH", ""),
    "PAL": ("पाल", "HIGH", ""),
    "RAI": ("राय", "HIGH", ""),
    "KHARWAR": ("खरवार", "HIGH", ""),
    "BAITHA": ("बैठा", "HIGH", ""),
    "HARIJAN": ("हरिजन", "HIGH", ""),
    "DHOBI": ("धोबी", "HIGH", ""),
    "KOIRI": ("कोइरी", "HIGH", ""),
    "RAWAT": ("रावत", "HIGH", ""),
    "GIRI": ("गिरि", "HIGH", ""),
    "GOSWAMI": ("गोस्वामी", "HIGH", ""),
    "DIXIT": ("दीक्षित", "HIGH", ""),
    "UPADHYAY": ("उपाध्याय", "HIGH", ""),
    "TRIPATHI": ("त्रिपाठी", "HIGH", ""),
    "DUBEY": ("दुबे", "HIGH", ""),
    "CHAUBEY": ("चौबे", "HIGH", ""),
    "JAISWAL": ("जायसवाल", "HIGH", ""),
    "BARNWAL": ("बर्नवाल", "HIGH", ""),
    "MADDHESHIYA": ("मद्धेशिया", "HIGH", ""),
    "MADHESHIYA": ("मद्धेशिया", "HIGH", ""),
    "KASERA": ("कसेरा", "HIGH", ""),
    "RASTOGI": ("रस्तोगी", "HIGH", ""),
    "AGARWAL": ("अग्रवाल", "HIGH", ""),
    "AGRAWAL": ("अग्रवाल", "HIGH", ""),
    "KESHARI": ("केशरी", "HIGH", ""),
    "KESHRI": ("केशरी", "HIGH", ""),
    "SONI": ("सोनी", "HIGH", ""),
    "THAKUR": ("ठाकुर", "HIGH", ""),
    "CHAUDHARY": ("चौधरी", "HIGH", ""),
    "CHOUDHARY": ("चौधरी", "HIGH", ""),
    "BHARTI": ("भारती", "HIGH", ""),
    "BHARATI": ("भारती", "HIGH", ""),
    "RAUNIYAR": ("रौनियार", "HIGH", ""),
    "RONIYAR": ("रौनियार", "HIGH", ""),
    "KANOJIYA": ("कनौजिया", "HIGH", ""),
    "KANOJIA": ("कनौजिया", "HIGH", ""),
    "KANAUJIA": ("कनौजिया", "HIGH", ""),
    "VISHWAKARMA": ("विश्वकर्मा", "HIGH", ""),
    "PRAJAPATI": ("प्रजापति", "HIGH", ""),
    "PRAJAWATI": ("प्रजापति", "HIGH", "Typo Prajawati->प्रजापति"),
    "PANCHAL": ("पांचाल", "HIGH", ""),
    "NISHAD": ("निषाद", "HIGH", ""),
    "KASAUDHAN": ("कसौधन", "HIGH", ""),
    "DWIVEDI": ("द्विवेदी", "HIGH", ""),
    "CHATURVEDI": ("चतुर्वेदी", "HIGH", ""),
    "KUMBHAR": ("कुम्हार", "HIGH", ""),
    "CHITRANSH": ("चित्रांश", "HIGH", ""),
    "BHARDWAJ": ("भारद्वाज", "HIGH", ""),
    "LATE": ("स्व०", "HIGH", "Late abbreviation"),

    # Ambiguous Gender Names -> Set to MEDIUM with clear review notes
    "KRISHNA": ("कृष्ण", "MEDIUM", "Ambiguous gender: 'कृष्ण' (male/father/Kumar) vs 'कृष्णा' (female)"),
    "RAMA": ("राम", "MEDIUM", "Ambiguous: 'राम' (male) vs 'रमा' (female)"),
    "SHIVA": ("शिव", "MEDIUM", "Ambiguous: 'शिव' (male) vs 'शिवा' (female)"),

    # Common Hindu Female Names (Clean '-i' endings & specific variants)
    "DULARI": ("दुलारी", "HIGH", ""),
    "FULKUMARI": ("फूलकुमारी", "HIGH", ""),
    "PHOOLKUMARI": ("फूलकुमारी", "HIGH", ""),
    "PHULKUMARI": ("फूलकुमारी", "HIGH", ""),
    "LALTI": ("लालती", "HIGH", ""),
    "SUGANTI": ("सुगंती", "HIGH", ""),
    "GYANTI": ("ज्ञानती", "HIGH", ""),
    "SUNAINA": ("सुनैना", "HIGH", ""),
    "SANJANA": ("संजना", "HIGH", ""),
    "SALONI": ("सलोनी", "HIGH", ""),
    "KHUSHI": ("खुशी", "HIGH", ""),
    "ANJU": ("अंजू", "HIGH", ""),
    "GUDDI": ("गुड्डी", "HIGH", ""),
    "KALAWATI": ("कलावती", "HIGH", ""),
    "PRABHAWATI": ("प्रभावती", "HIGH", ""),
    "MALATI": ("मालती", "HIGH", ""),
    "SHILA": ("शीला", "HIGH", ""),
    "REETA": ("रीता", "HIGH", ""),
    "RINA": ("रीना", "HIGH", ""),
    "PUNAM": ("पूनम", "HIGH", ""),
    "SIMA": ("सीमा", "HIGH", ""),
    "MAMATA": ("ममता", "HIGH", ""),
    "SUCHITA": ("सुचिता", "HIGH", ""),
    "KISHA": ("किशा", "HIGH", ""),
    "YASHODA": ("यशोदा", "HIGH", ""),
    "RADHIKA": ("राधिका", "HIGH", ""),
    "UMARAWATI": ("उमरावती", "HIGH", ""),
    "LUCKY": ("लकी", "HIGH", ""),
    "SUNNY": ("सन्नी", "HIGH", ""),
    "SONY": ("सोनी", "HIGH", ""),
    "SWEJAL": ("स्वेजल", "HIGH", ""),
    "DHONI": ("धोनी", "HIGH", ""),
    "KISHOR": ("किशोर", "HIGH", ""),
    "KISHORE": ("किशोर", "HIGH", ""),
    "SHRAVAN": ("श्रवण", "HIGH", ""),
    "SARAWAN": ("श्रवण", "HIGH", ""),

    # Common Hindu Given Names & Variants
    "AASHU": ("आशू", "HIGH", ""),
    "AARTI": ("आरती", "HIGH", ""),
    "ARTI": ("आरती", "HIGH", ""),
    "AANCHAL": ("आंचल", "HIGH", ""),
    "ANCHAL": ("आंचल", "HIGH", ""),
    "AANSHI": ("आंशी", "HIGH", ""),
    "AARYAN": ("आर्यन", "HIGH", ""),
    "AARYA": ("आर्या", "HIGH", ""),
    "ABHAY": ("अभय", "HIGH", ""),
    "ABHINANDAN": ("अभिनंदन", "HIGH", ""),
    "ABHISHEK": ("अभिषेक", "HIGH", ""),
    "ADARSH": ("आदर्श", "HIGH", ""),
    "ADITYA": ("आदित्य", "HIGH", ""),
    "AJAY": ("अजय", "HIGH", ""),
    "AJAI": ("अजय", "HIGH", ""),
    "AKANKSHA": ("आकांक्षा", "HIGH", ""),
    "AKASH": ("आकाश", "HIGH", ""),
    "AKHILESH": ("अखिलेश", "HIGH", ""),
    "AKRITI": ("आकृति", "HIGH", ""),
    "ALOK": ("आलोक", "HIGH", ""),
    "AMAN": ("अमन", "HIGH", ""),
    "AMAR": ("अमर", "HIGH", ""),
    "AMARJEET": ("अमरजीत", "HIGH", ""),
    "AMARNATH": ("अमरनाथ", "HIGH", ""),
    "AMRESH": ("अमरेश", "HIGH", ""),
    "AMRIT": ("अमृत", "HIGH", ""),
    "AMRITA": ("अमृता", "HIGH", ""),
    "AMIT": ("अमित", "HIGH", ""),
    "AMITA": ("अमिता", "HIGH", ""),
    "ANAMIKA": ("अनामिका", "HIGH", ""),
    "ANAND": ("आनंद", "HIGH", ""),
    "ANANYA": ("अनन्या", "HIGH", ""),
    "ANGAD": ("अंगद", "HIGH", ""),
    "ANIKET": ("अनिकेत", "HIGH", ""),
    "ANIL": ("अनिल", "HIGH", ""),
    "ANIRUDDHA": ("अनिरुद्ध", "HIGH", ""),
    "ANIRUDH": ("अनिरुद्ध", "HIGH", ""),
    "ANISHA": ("अनीशा", "HIGH", ""),
    "ANITA": ("अनीता", "HIGH", ""),
    "ANJALI": ("अंजलि", "HIGH", ""),
    "ANJANI": ("अंजनी", "HIGH", ""),
    "ANKIT": ("अंकित", "HIGH", ""),
    "ANKITA": ("अंकिता", "HIGH", ""),
    "ANKUSH": ("अंकुश", "HIGH", ""),
    "ANSH": ("अंश", "HIGH", ""),
    "ANSHIKA": ("अंशिका", "HIGH", ""),
    "ANSHU": ("अंशु", "HIGH", ""),
    "ANUP": ("अनूप", "HIGH", ""),
    "ANUPAM": ("अनुपम", "HIGH", ""),
    "ANURADHA": ("अनुराधा", "HIGH", ""),
    "ANURAG": ("अनुराग", "HIGH", ""),
    "ARCHANA": ("अर्चना", "HIGH", ""),
    "ARJUN": ("अर्जुन", "HIGH", ""),
    "ARPITA": ("अर्पिता", "HIGH", ""),
    "ARUN": ("अरुण", "HIGH", ""),
    "ARUNA": ("अरुणा", "HIGH", ""),
    "ARVIND": ("अरविंद", "HIGH", ""),
    "ARYAN": ("आर्यन", "HIGH", ""),
    "ASHISH": ("आशीष", "HIGH", ""),
    "ASHOK": ("अशोक", "HIGH", ""),
    "ASHA": ("आशा", "HIGH", ""),
    "AASHA": ("आशा", "HIGH", ""),
    "ATUL": ("अतुल", "HIGH", ""),
    "AVINASH": ("अविनाश", "HIGH", ""),
    "AYUSH": ("आयुष", "HIGH", ""),
    "AYUSHI": ("आयुषी", "HIGH", ""),
    "BABITA": ("बबीता", "HIGH", ""),
    "BABLU": ("बबलू", "HIGH", ""),
    "BABU": ("बाबू", "HIGH", ""),
    "BABURAM": ("बाबूराम", "HIGH", ""),
    "BADAL": ("बादल", "HIGH", ""),
    "BAIJNATH": ("बैजनाथ", "HIGH", ""),
    "BAJRANGI": ("बजरंगी", "HIGH", ""),
    "BALIRAM": ("बलीराम", "HIGH", ""),
    "BALWANT": ("बलवंत", "HIGH", ""),
    "BANDANA": ("वंदना", "HIGH", ""),
    "BASANT": ("बसंत", "HIGH", ""),
    "BEENA": ("बीना", "HIGH", ""),
    "BHAGWAN": ("भगवान", "HIGH", ""),
    "BHAGWATI": ("भगवती", "HIGH", ""),
    "BHARAT": ("भरत", "HIGH", ""),
    "BHOLA": ("भोला", "HIGH", ""),
    "BHUPENDRA": ("भूपेंद्र", "HIGH", ""),
    "BIJAY": ("विजय", "HIGH", ""),
    "BIKRAM": ("विक्रम", "HIGH", ""),
    "BINDESWARI": ("बिंदेश्वरी", "HIGH", ""),
    "BINDU": ("बिंदु", "HIGH", ""),
    "BIPIN": ("विपिन", "HIGH", ""),
    "BIRENDRA": ("वीरेंद्र", "HIGH", ""),
    "BRAJESH": ("ब्रजेश", "HIGH", ""),
    "BRIJESH": ("बृजेश", "HIGH", ""),
    "CHANDAN": ("चंदन", "HIGH", ""),
    "CHANDA": ("चंदा", "HIGH", ""),
    "CHANDANI": ("चांदनी", "HIGH", ""),
    "CHANDRA": ("चंद्र", "HIGH", ""),
    "CHANDRAKALA": ("चंद्रकला", "HIGH", ""),
    "CHANDRABHAN": ("चंद्रभान", "HIGH", ""),
    "CHANDRAMANI": ("चंद्रमणि", "HIGH", ""),
    "CHANDRASHEKHAR": ("चंद्रशेखर", "HIGH", ""),
    "CHANDRAWATI": ("चंद्रावती", "HIGH", ""),
    "CHHOTELAL": ("छोटेलाल", "HIGH", ""),
    "CHINTA": ("चिंता", "HIGH", ""),
    "DAMODAR": ("दामोदर", "HIGH", ""),
    "DAYA": ("दया", "HIGH", ""),
    "DAYASHANKAR": ("दयाशंकर", "HIGH", ""),
    "DYASHANKAR": ("दयाशंकर", "LOW", "Typo Dyashankar->दयाशंकर"),
    "DEENANATH": ("दीनानाथ", "HIGH", ""),
    "DEEPAK": ("दीपक", "HIGH", ""),
    "DEEPIKA": ("दीपिका", "HIGH", ""),
    "DEEPTI": ("दीप्ति", "HIGH", ""),
    "DEV": ("देव", "HIGH", ""),
    "DEVENDRA": ("देवेंद्र", "HIGH", ""),
    "DHANANJAY": ("धनंजय", "HIGH", ""),
    "DHARMENDRA": ("धर्मेंद्र", "HIGH", ""),
    "DHARMVEER": ("धर्मवीर", "HIGH", ""),
    "DHIRAJ": ("धीरज", "HIGH", ""),
    "DHIRENDRA": ("धीरेंद्र", "HIGH", ""),
    "DILIP": ("दिलीप", "HIGH", ""),
    "DINESH": ("दिनेश", "HIGH", ""),
    "DIVYA": ("दिव्या", "HIGH", ""),
    "DIVYANSH": ("दिव्यांश", "HIGH", ""),
    "DIVYANSHI": ("दिव्यांशी", "HIGH", ""),
    "DURGA": ("दुर्गा", "HIGH", ""),
    "DURGAWATI": ("दुर्गावती", "HIGH", ""),
    "DURGESH": ("दुर्गेश", "HIGH", ""),
    "DURGI": ("दुर्गी", "HIGH", ""),
    "GANESH": ("गणेश", "HIGH", ""),
    "GANGA": ("गंगा", "HIGH", ""),
    "GARIMA": ("गरिमा", "HIGH", ""),
    "GAUTAM": ("गौतम", "HIGH", ""),
    "GAYATRI": ("गायत्री", "HIGH", ""),
    "GEETA": ("गीता", "HIGH", ""),
    "GIRJESH": ("गिरजेश", "HIGH", ""),
    "GITA": ("गीता", "HIGH", ""),
    "GOBIND": ("गोविंद", "HIGH", ""),
    "GOPAL": ("गोपाल", "HIGH", ""),
    "GOVIND": ("गोविंद", "HIGH", ""),
    "GUDDU": ("गुड्डू", "HIGH", ""),
    "GUDIYA": ("गुड़िया", "HIGH", ""),
    "GULAB": ("गुलाब", "HIGH", ""),
    "GUNJAN": ("गुंजन", "HIGH", ""),
    "GYANCHAND": ("ज्ञानचंद", "HIGH", ""),
    "HARENDRA": ("हरेंद्र", "HIGH", ""),
    "HARI": ("हरि", "HIGH", ""),
    "HARIOM": ("हरिओम", "HIGH", ""),
    "HARIKESH": ("हरिकेश", "HIGH", ""),
    "HARINATH": ("हरिनाथ", "HIGH", ""),
    "HARIRAM": ("हरिराम", "HIGH", ""),
    "HARISH": ("हरीश", "HIGH", ""),
    "HARISHCHANDRA": ("हरीशचंद्र", "HIGH", ""),
    "HARSH": ("हर्ष", "HIGH", ""),
    "HARSHITA": ("हर्षिता", "HIGH", ""),
    "HEMANT": ("हेमंत", "HIGH", ""),
    "HIMANSHU": ("हिमांशु", "HIGH", ""),
    "INDAL": ("इन्दल", "HIGH", ""),
    "INDU": ("इंदु", "HIGH", ""),
    "INDRAWATI": ("इंद्रावती", "HIGH", ""),
    "INDRESH": ("इंद्रेश", "HIGH", ""),
    "JAGDISH": ("जगदीश", "HIGH", ""),
    "JAIPRAKASH": ("जयप्रकाश", "HIGH", ""),
    "JANAK": ("जनक", "HIGH", ""),
    "JANKI": ("जानकी", "HIGH", ""),
    "JAY": ("जय", "HIGH", ""),
    "JAYPRAKASH": ("जयप्रकाश", "HIGH", ""),
    "JEETENDRA": ("जितेंद्र", "HIGH", ""),
    "JITENDRA": ("जितेंद्र", "HIGH", ""),
    "JYOTI": ("ज्योति", "HIGH", ""),
    "KAILASH": ("कैलाश", "HIGH", ""),
    "KAJAL": ("काजल", "HIGH", ""),
    "KALPANA": ("कल्पना", "HIGH", ""),
    "KALYANI": ("कल्याणी", "HIGH", ""),
    "KAMAL": ("कमल", "HIGH", ""),
    "KAMALESH": ("कमलेश", "HIGH", ""),
    "KAMLESH": ("कमलेश", "HIGH", ""),
    "KAMLA": ("कमला", "HIGH", ""),
    "KAMLAWATI": ("कमलावती", "HIGH", ""),
    "KANCHAN": ("कंचन", "HIGH", ""),
    "KANHAIYA": ("कन्हैया", "HIGH", ""),
    "KAPIL": ("कपिल", "HIGH", ""),
    "KARAN": ("करण", "HIGH", ""),
    "KASHINATH": ("काशीनाथ", "HIGH", ""),
    "KAUSHAL": ("कौशल", "HIGH", ""),
    "KAUSHALYA": ("कौशल्या", "HIGH", ""),
    "KAVITA": ("कविता", "HIGH", ""),
    "KEDAR": ("केदार", "HIGH", ""),
    "KHUSHBOO": ("खुशबू", "HIGH", ""),
    "KHUSHBU": ("खुशबू", "HIGH", ""),
    "KIRAN": ("किरण", "HIGH", ""),
    "KIRTAN": ("कीर्तन", "HIGH", ""),
    "KOMAL": ("कोमल", "HIGH", ""),
    "KRISHNAWATI": ("कृष्णावती", "HIGH", ""),
    "KULDEEP": ("कुलदीप", "HIGH", ""),
    "KUSUM": ("कुसुम", "HIGH", ""),
    "LAKSHMI": ("लक्ष्मी", "HIGH", ""),
    "LALBAHADUR": ("लालबहादुर", "HIGH", ""),
    "LALBIHARI": ("लालबिहारी", "HIGH", ""),
    "LALITA": ("ललिता", "HIGH", ""),
    "LALMANI": ("लालमणि", "HIGH", ""),
    "LALTAPRASAD": ("लालताप्रसाद", "HIGH", ""),
    "LAXMI": ("लक्ष्मी", "HIGH", ""),
    "LAXMAN": ("लक्ष्मण", "HIGH", ""),
    "LILAWATI": ("लीलावती", "HIGH", ""),
    "LOKESH": ("लोकेश", "HIGH", ""),
    "MADAN": ("मदन", "HIGH", ""),
    "MADHURI": ("माधुरी", "HIGH", ""),
    "MADHU": ("मधु", "HIGH", ""),
    "MAHENDRA": ("महेंद्र", "HIGH", ""),
    "MAHESH": ("महेश", "HIGH", ""),
    "MALTI": ("मालती", "HIGH", ""),
    "MAMTA": ("ममता", "HIGH", ""),
    "MANGAL": ("मंगल", "HIGH", ""),
    "MANISH": ("मनीष", "HIGH", ""),
    "MANISHA": ("मनीषा", "HIGH", ""),
    "MANJU": ("मंजू", "HIGH", ""),
    "MANOJ": ("मनोज", "HIGH", ""),
    "MANORAMA": ("मनोरमा", "HIGH", ""),
    "MANSHI": ("मानसी", "HIGH", ""),
    "MANSI": ("मानसी", "HIGH", ""),
    "MAYA": ("माया", "HIGH", ""),
    "MEENA": ("मीना", "HIGH", ""),
    "MEERA": ("मीरा", "HIGH", ""),
    "MITHILESH": ("मिथिलेश", "HIGH", ""),
    "MITHUN": ("मिथुन", "HIGH", ""),
    "MOHAN": ("मोहन", "HIGH", ""),
    "MOHIT": ("मोहित", "HIGH", ""),
    "MOHINI": ("मोहिनी", "HIGH", ""),
    "MONI": ("मोनी", "HIGH", ""),
    "MONIKA": ("मोनिका", "HIGH", ""),
    "MONU": ("मोनू", "HIGH", ""),
    "MOTILAL": ("मोतीलाल", "HIGH", ""),
    "MUKESH": ("मुकेश", "HIGH", ""),
    "MUNNA": ("मुन्ना", "HIGH", ""),
    "MUNNI": ("मुन्नी", "HIGH", ""),
    "MURARI": ("मुरारी", "HIGH", ""),
    "MUSKAN": ("मुस्कान", "HIGH", ""),
    "NAGENDRA": ("नागेंद्र", "HIGH", ""),
    "NAINA": ("नैना", "HIGH", ""),
    "NANDKISHOR": ("नंदकिशोर", "HIGH", ""),
    "NANDLAL": ("नंदलाल", "HIGH", ""),
    "NANDINI": ("नंदिनी", "HIGH", ""),
    "NARAYAN": ("नारायण", "HIGH", ""),
    "NARENDRA": ("नरेंद्र", "HIGH", ""),
    "NARESH": ("नरेश", "HIGH", ""),
    "NAVNEET": ("नवनीत", "HIGH", ""),
    "NEELAM": ("नीलम", "HIGH", ""),
    "NEELU": ("नीलू", "HIGH", ""),
    "NEERAJ": ("नीरज", "HIGH", ""),
    "NEERU": ("नीरू", "HIGH", ""),
    "NEETU": ("नीतू", "HIGH", ""),
    "NEHA": ("नेहा", "HIGH", ""),
    "NESHA": ("नेशा", "LOW", "Urdu/Purvanchal variant - verify नेशा/निशा"),
    "NIDHI": ("निधि", "HIGH", ""),
    "NIKHIL": ("निखिल", "HIGH", ""),
    "NIKITA": ("निकिता", "HIGH", ""),
    "NIRAJ": ("नीरज", "HIGH", ""),
    "NIRALA": ("निराला", "HIGH", ""),
    "NIRANJAN": ("निरंजन", "HIGH", ""),
    "NIRMAL": ("निर्मल", "HIGH", ""),
    "NIRMALA": ("निर्मला", "HIGH", ""),
    "NISHA": ("निशा", "HIGH", ""),
    "NISHU": ("नीशू", "HIGH", ""),
    "NITESH": ("नितेश", "HIGH", ""),
    "NITIN": ("नितिन", "HIGH", ""),
    "NITISH": ("नीतीश", "HIGH", ""),
    "NITU": ("नीतू", "HIGH", ""),
    "OMPRAKASH": ("ओमप्रकाश", "HIGH", ""),
    "PANKAJ": ("पंकज", "HIGH", ""),
    "PARAS": ("पारस", "HIGH", ""),
    "PARASNATH": ("पारसनाथ", "HIGH", ""),
    "PARBHAVATI": ("प्रभावती", "HIGH", ""),
    "PARVATI": ("पार्वती", "HIGH", ""),
    "PAVAN": ("पवन", "HIGH", ""),
    "PAWAN": ("पवन", "HIGH", ""),
    "PAYAL": ("पायल", "HIGH", ""),
    "PINKI": ("पिंकी", "HIGH", ""),
    "PINKY": ("पिंकी", "HIGH", ""),
    "PINTU": ("पिंटू", "HIGH", ""),
    "PIYUSH": ("पीयूष", "HIGH", ""),
    "POOJA": ("पूजा", "HIGH", ""),
    "POONAM": ("पूनम", "HIGH", ""),
    "PRABHA": ("प्रभा", "HIGH", ""),
    "PRABHAT": ("प्रभात", "HIGH", ""),
    "PRABHAVATI": ("प्रभावती", "HIGH", ""),
    "PRABHU": ("प्रभु", "HIGH", ""),
    "PRADEEP": ("प्रदीप", "HIGH", ""),
    "PRADIP": ("प्रदीप", "HIGH", ""),
    "PRAKASH": ("प्रकाश", "HIGH", ""),
    "PRAMOD": ("प्रमोद", "HIGH", ""),
    "PRANAV": ("प्रणव", "HIGH", ""),
    "PRASHANT": ("प्रशांत", "HIGH", ""),
    "PRATAP": ("प्रताप", "HIGH", ""),
    "PRATIBHA": ("प्रतिभा", "HIGH", ""),
    "PRATIMA": ("प्रतिमा", "HIGH", ""),
    "PRAVEEN": ("प्रवीण", "HIGH", ""),
    "PREETI": ("प्रीति", "HIGH", ""),
    "PREM": ("प्रेम", "HIGH", ""),
    "PREMCHAND": ("प्रेमचंद", "HIGH", ""),
    "PRERNA": ("प्रेरणा", "HIGH", ""),
    "PRINCE": ("प्रिंस", "HIGH", ""),
    "PRITI": ("प्रीति", "HIGH", ""),
    "PRIYA": ("प्रिया", "HIGH", ""),
    "PRIYANKA": ("प्रियंका", "HIGH", ""),
    "PRIYANSHU": ("प्रियांशु", "HIGH", ""),
    "PUJA": ("पूजा", "HIGH", ""),
    "PUNITA": ("पुनीता", "HIGH", ""),
    "PUSHP": ("पुष्प", "HIGH", ""),
    "PUSHPA": ("पुष्पा", "HIGH", ""),
    "RADHA": ("राधा", "HIGH", ""),
    "RADHESHYAM": ("राधेश्याम", "HIGH", ""),
    "RAGHAV": ("राघव", "HIGH", ""),
    "RAGINI": ("रागिनी", "HIGH", ""),
    "RAGHVENDRA": ("राघवेंद्र", "HIGH", ""),
    "RAHUL": ("राहुल", "HIGH", ""),
    "RAJAN": ("राजन", "HIGH", ""),
    "RAJANI": ("रजनी", "HIGH", ""),
    "RAJENDRA": ("राजेन्द्र", "HIGH", ""),
    "RAJESH": ("राजेश", "HIGH", ""),
    "RAJIV": ("राजीव", "HIGH", ""),
    "RAJKUMAR": ("राजकुमार", "HIGH", ""),
    "RAJKUMARI": ("राजकुमारी", "HIGH", ""),
    "RAJNATH": ("राजनाथ", "HIGH", ""),
    "RAJNEESH": ("रजनीश", "HIGH", ""),
    "RAJNI": ("रजनी", "HIGH", ""),
    "RAJU": ("राजू", "HIGH", ""),
    "RAKESH": ("राकेश", "HIGH", ""),
    "RAKHI": ("राखी", "HIGH", ""),
    "RAMAKANT": ("रमाकांत", "HIGH", ""),
    "RAMAN": ("रमन", "HIGH", ""),
    "RAMANAND": ("रामानंद", "HIGH", ""),
    "RAMASHANKAR": ("रमाशंकर", "HIGH", ""),
    "RAMAVTAR": ("रामावतार", "HIGH", ""),
    "RAMAWATI": ("रामावती", "HIGH", ""),
    "RAMAYAN": ("रामायण", "HIGH", ""),
    "RAMBACHAN": ("रामबचन", "HIGH", ""),
    "RAMBAHADUR": ("रामबहादुर", "HIGH", ""),
    "RAMBALI": ("रामबली", "HIGH", ""),
    "RAMBHA": ("रंभा", "HIGH", ""),
    "RAMBHAWAN": ("रामभवन", "HIGH", ""),
    "RAMBILAS": ("रामबिलास", "HIGH", ""),
    "RAMBRIKSH": ("रामबृक्ष", "HIGH", ""),
    "RAMCHANDRA": ("रामचंद्र", "HIGH", ""),
    "RAMCHET": ("रामचेत", "HIGH", ""),
    "RAMDARASH": ("रामदरश", "HIGH", ""),
    "RAMDHANI": ("रामधनी", "HIGH", ""),
    "RAMDHYAN": ("रामध्यान", "HIGH", ""),
    "RAMESH": ("रमेश", "HIGH", ""),
    "RAMESHWAR": ("रामेश्वर", "HIGH", ""),
    "RAMGOPAL": ("रामगोपाल", "HIGH", ""),
    "RAMJANAM": ("रामजनम", "HIGH", ""),
    "RAMJI": ("रामजी", "HIGH", ""),
    "RAMKESH": ("रामकेश", "HIGH", ""),
    "RAMKRIPAL": ("रामकृपाल", "HIGH", ""),
    "RAMLAL": ("रामलाल", "HIGH", ""),
    "RAMLAKHAN": ("रामलखन", "HIGH", ""),
    "RAMMANI": ("राममणि", "HIGH", ""),
    "RAMNARESH": ("रामनरेश", "HIGH", ""),
    "RAMNATH": ("रामनाथ", "HIGH", ""),
    "RAMNIWAS": ("रामनिवास", "HIGH", ""),
    "RAMPATI": ("रामपति", "HIGH", ""),
    "RAMPRASAD": ("रामप्रसाद", "HIGH", ""),
    "RAMPRATAP": ("रामप्रताप", "HIGH", ""),
    "RAMPREET": ("रामप्रीत", "HIGH", ""),
    "RAMPYARE": ("रामप्यारे", "HIGH", ""),
    "RAMSAKAL": ("रामसकल", "HIGH", ""),
    "RAMSAMUJH": ("रामसमुझ", "HIGH", ""),
    "RAMSEWAK": ("रामसेवक", "HIGH", ""),
    "RAMSHANKAR": ("रामशंकर", "HIGH", ""),
    "RAMSHARAN": ("रामशरण", "HIGH", ""),
    "RAMSHEEL": ("रामशील", "HIGH", ""),
    "RAMSINGH": ("रामसिंह", "HIGH", ""),
    "RAMSUNDAR": ("रामसुंदर", "HIGH", ""),
    "RAMSURAT": ("रामसूरत", "HIGH", ""),
    "RAMTAHAL": ("रामटहल", "HIGH", ""),
    "RAMU": ("रामू", "HIGH", ""),
    "RAMVILAS": ("रामविलास", "HIGH", ""),
    "RAMVRIKSH": ("रामवृक्ष", "HIGH", ""),
    "RANI": ("रानी", "HIGH", ""),
    "RANJANA": ("रंजना", "HIGH", ""),
    "RANJEET": ("रणजीत", "HIGH", ""),
    "RANJIT": ("रणजीत", "HIGH", ""),
    "RANJU": ("रंजू", "HIGH", ""),
    "RATNESH": ("रत्नेश", "HIGH", ""),
    "RAVI": ("रवि", "HIGH", ""),
    "RAVINDRA": ("रवींद्र", "HIGH", ""),
    "RAVISH": ("रविश", "HIGH", ""),
    "REENA": ("रीना", "HIGH", ""),
    "REETU": ("रीतू", "HIGH", ""),
    "REKHA": ("रेखा", "HIGH", ""),
    "RENU": ("रेनू", "HIGH", ""),
    "RINKI": ("रिंकी", "HIGH", ""),
    "RINKU": ("रिंकू", "HIGH", ""),
    "RISHABH": ("ऋषभ", "HIGH", ""),
    "RISHI": ("ऋषि", "HIGH", ""),
    "RITA": ("रीता", "HIGH", ""),
    "RITESH": ("रितेश", "HIGH", ""),
    "RITIK": ("ऋतिक", "HIGH", ""),
    "RITU": ("रितु", "HIGH", ""),
    "RIYA": ("रिया", "HIGH", ""),
    "ROHIT": ("रोहित", "HIGH", ""),
    "ROSHAN": ("रोशन", "HIGH", ""),
    "ROSHANI": ("रोशनी", "HIGH", ""),
    "RUBI": ("रूबी", "HIGH", ""),
    "RUBY": ("रूबी", "HIGH", ""),
    "RUCHI": ("रुचि", "HIGH", ""),
    "RUPA": ("रूपा", "HIGH", ""),
    "RUPALI": ("रूपाली", "HIGH", ""),
    "RUPESH": ("रूपेश", "HIGH", ""),
    "SACHIN": ("सचिन", "HIGH", ""),
    "SADHANA": ("साधना", "HIGH", ""),
    "SAGAR": ("सागर", "HIGH", ""),
    "SAKSHI": ("साक्षी", "HIGH", ""),
    "SANDEEP": ("संदीप", "HIGH", ""),
    "SANDHYA": ("संध्या", "HIGH", ""),
    "SANDIP": ("संदीप", "HIGH", ""),
    "SANDESH": ("संदेश", "HIGH", ""),
    "SANGEETA": ("संगीता", "HIGH", ""),
    "SANGITA": ("संगीता", "HIGH", ""),
    "SANJAY": ("संजय", "HIGH", ""),
    "SANJEEV": ("संजीव", "HIGH", ""),
    "SANJU": ("संजू", "HIGH", ""),
    "SANTLAL": ("संतलाल", "HIGH", ""),
    "SANTOSH": ("संतोष", "HIGH", ""),
    "SARITA": ("सरिता", "HIGH", ""),
    "SAROJ": ("सरोज", "HIGH", ""),
    "SAROJINI": ("सरोजिनी", "HIGH", ""),
    "SARVESH": ("सर्वेश", "HIGH", ""),
    "SATENDRA": ("सत्येंद्र", "HIGH", ""),
    "SATISH": ("सतीश", "HIGH", ""),
    "SATYA": ("सत्य", "HIGH", ""),
    "SATYAM": ("सत्यम", "HIGH", ""),
    "SATYANARAYAN": ("सत्यनारायण", "HIGH", ""),
    "SATYAPRAKASH": ("सत्यप्रकाश", "HIGH", ""),
    "SATYENDRA": ("सत्येंद्र", "HIGH", ""),
    "SAURABH": ("सौरभ", "HIGH", ""),
    "SAVITA": ("सविता", "HIGH", ""),
    "SAVITRI": ("सावित्री", "HIGH", ""),
    "SEEMA": ("सीमा", "HIGH", ""),
    "SHAILESH": ("शैलेश", "HIGH", ""),
    "SHALINI": ("शालिनी", "HIGH", ""),
    "SHAMBHU": ("शंभू", "HIGH", ""),
    "SHANKAR": ("शंकर", "HIGH", ""),
    "SHANI": ("शनि", "HIGH", ""),
    "SHANTI": ("शांति", "HIGH", ""),
    "SHARDA": ("शारदा", "HIGH", ""),
    "SHASHANK": ("शशांक", "HIGH", ""),
    "SHASHI": ("शशि", "HIGH", ""),
    "SHEELA": ("शीला", "HIGH", ""),
    "SHEETAL": ("शीतल", "HIGH", ""),
    "SHIKHA": ("शिखा", "HIGH", ""),
    "SHIV": ("शिव", "HIGH", ""),
    "SHIVAJI": ("शिवाजी", "HIGH", ""),
    "SHIVAM": ("शिवम", "HIGH", ""),
    "SHIVANGI": ("शिवांगी", "HIGH", ""),
    "SHIVANI": ("शिवानी", "HIGH", ""),
    "SHIVKUMAR": ("शिवकुमार", "HIGH", ""),
    "SHIVNATH": ("शिवनाथ", "HIGH", ""),
    "SHIVPUJAN": ("शिवपूजन", "HIGH", ""),
    "SHIVRAM": ("शिवराम", "HIGH", ""),
    "SHOBHA": ("शोभा", "HIGH", ""),
    "SHRADHA": ("श्रद्धा", "HIGH", ""),
    "SHREYA": ("श्रेया", "HIGH", ""),
    "SHRIKANT": ("श्रीकांत", "HIGH", ""),
    "SHRINATH": ("श्रीनाथ", "HIGH", ""),
    "SHRIRAM": ("श्रीराम", "HIGH", ""),
    "SHRUTI": ("श्रुति", "HIGH", ""),
    "SHUBHAM": ("शुभम", "HIGH", ""),
    "SHWETA": ("श्वेता", "HIGH", ""),
    "SHYAM": ("श्याम", "HIGH", ""),
    "SHYAMSUNDAR": ("श्यामसुंदर", "HIGH", ""),
    "SIDDHARTH": ("सिद्धार्थ", "HIGH", ""),
    "SITA": ("सीता", "HIGH", ""),
    "SITARAM": ("सीताराम", "HIGH", ""),
    "SMRITI": ("स्मृति", "HIGH", ""),
    "SNEHA": ("स्नेहा", "HIGH", ""),
    "SOBHA": ("शोभा", "HIGH", ""),
    "SOHAN": ("सोहन", "HIGH", ""),
    "SOMNATH": ("सोमनाथ", "HIGH", ""),
    "SONAL": ("सोनल", "HIGH", ""),
    "SONALI": ("सोनाली", "HIGH", ""),
    "SONAM": ("सोनम", "HIGH", ""),
    "SOANAM": ("सोनम", "HIGH", "Typo Soanam->सोनम"),
    "SONI": ("सोनी", "HIGH", ""),
    "SONIA": ("सोनिया", "HIGH", ""),
    "SONIYA": ("सोनिया", "HIGH", ""),
    "SONU": ("सोनू", "HIGH", ""),
    "SUBASH": ("सुभाष", "HIGH", "Subash->सुभाष"),
    "SUBHASH": ("सुभाष", "HIGH", ""),
    "SUDHA": ("सुधा", "HIGH", ""),
    "SUDHIR": ("सुधीर", "HIGH", ""),
    "SUJEET": ("सुजीत", "HIGH", ""),
    "SUJIT": ("सुजीत", "HIGH", ""),
    "SUMAN": ("सुमन", "HIGH", ""),
    "SUMIT": ("सुमित", "HIGH", ""),
    "SUMITRA": ("सुमित्रा", "HIGH", ""),
    "SUNDAR": ("सुंदर", "HIGH", ""),
    "SUNIL": ("सुनील", "HIGH", ""),
    "SUNITA": ("सुनीता", "HIGH", ""),
    "SURAJ": ("सूरज", "HIGH", ""),
    "SURAJBHAN": ("सूरजभान", "HIGH", ""),
    "SURENDRA": ("सुरेंद्र", "HIGH", ""),
    "SURESH": ("सुरेश", "HIGH", ""),
    "SURYA": ("सूर्य", "HIGH", ""),
    "SUSHIL": ("सुशील", "HIGH", ""),
    "SUSHILA": ("सुशीला", "HIGH", ""),
    "SUSHMA": ("सुषमा", "HIGH", ""),
    "SWATI": ("स्वाति", "HIGH", ""),
    "TANNU": ("तन्नू", "HIGH", ""),
    "TANU": ("तनु", "HIGH", ""),
    "TANUJ": ("तनुज", "HIGH", ""),
    "TANYA": ("तान्या", "HIGH", ""),
    "TARA": ("तारा", "HIGH", ""),
    "TARUN": ("तरुण", "HIGH", ""),
    "TEJPRATAP": ("तेजप्रताप", "HIGH", ""),
    "TRIBHUVAN": ("त्रिभुवन", "HIGH", ""),
    "TRILOKI": ("त्रिलोकी", "HIGH", ""),
    "TRIPTI": ("तृप्ति", "HIGH", ""),
    "TULSI": ("तुलसी", "HIGH", ""),
    "TUNTUN": ("टुनटुन", "HIGH", ""),
    "UDAY": ("उदय", "HIGH", ""),
    "UDAYBHAN": ("उदयभान", "HIGH", ""),
    "UDAYRAJ": ("उदयराज", "HIGH", ""),
    "UJJAWAL": ("उज्ज्वल", "HIGH", ""),
    "UJJWAL": ("उज्ज्वल", "HIGH", ""),
    "UMA": ("उमा", "HIGH", ""),
    "UMASHANKAR": ("उमाशंकर", "HIGH", ""),
    "UMESH": ("उमेश", "HIGH", ""),
    "UPENDRA": ("उपेंद्र", "HIGH", ""),
    "URMILA": ("उर्मिला", "HIGH", ""),
    "USHA": ("उषा", "HIGH", ""),
    "UTKARSH": ("उत्कर्ष", "HIGH", ""),
    "VAIBHAV": ("वैभव", "HIGH", ""),
    "VAISHALI": ("वैशाली", "HIGH", ""),
    "VAISHNAVI": ("वैष्णवी", "HIGH", ""),
    "VANDANA": ("वंदना", "HIGH", ""),
    "VARSHA": ("वर्षा", "HIGH", ""),
    "VARUN": ("वरुण", "HIGH", ""),
    "VASUDEV": ("वासुदेव", "HIGH", ""),
    "VED": ("वेद", "HIGH", ""),
    "VEDPRAKASH": ("वेदप्रकाश", "HIGH", ""),
    "VIDYA": ("विद्या", "HIGH", ""),
    "VIDYAWATI": ("विद्यावती", "HIGH", ""),
    "VIJAY": ("विजय", "HIGH", ""),
    "VIJAYLAXMI": ("विजयलक्ष्मी", "HIGH", ""),
    "VIKAS": ("विकास", "HIGH", ""),
    "VIKASH": ("विकास", "HIGH", ""),
    "VIKRAM": ("विक्रम", "HIGH", ""),
    "VIMAL": ("विमल", "HIGH", ""),
    "VIMLA": ("विमला", "HIGH", ""),
    "VIMLESH": ("विमलेश", "HIGH", ""),
    "VINAY": ("विनय", "HIGH", ""),
    "VINOD": ("विनोद", "HIGH", ""),
    "VIPIN": ("विपिन", "HIGH", ""),
    "VIRENDRA": ("वीरेंद्र", "HIGH", ""),
    "VISHAL": ("विशाल", "HIGH", ""),
    "VISHNU": ("विष्णु", "HIGH", ""),
    "VIVEK": ("विवेक", "HIGH", ""),
    "YOGENDRA": ("योगेंद्र", "HIGH", ""),
    "YOGESH": ("योगेश", "HIGH", ""),

    # Muslim / Urdu Names -> LOW confidence, flagged for school review
    "ABDUL": ("अब्दुल", "LOW", "Muslim/Urdu name"),
    "AFTAB": ("आफ़ताब", "LOW", "Muslim/Urdu name"),
    "AHAMAD": ("अहमद", "LOW", "Muslim/Urdu name"),
    "AHMAD": ("अहमद", "LOW", "Muslim/Urdu name"),
    "AHMED": ("अहमद", "LOW", "Muslim/Urdu name"),
    "AHSAN": ("अहसान", "LOW", "Muslim/Urdu name"),
    "AKBAR": ("अकबर", "LOW", "Muslim/Urdu name"),
    "AKHTAR": ("अख्तर", "LOW", "Muslim/Urdu name"),
    "AKIL": ("अकील", "LOW", "Muslim/Urdu name"),
    "AKRAM": ("अकरम", "LOW", "Muslim/Urdu name"),
    "ALAM": ("आलम", "LOW", "Muslim/Urdu name"),
    "ALEY": ("अले", "LOW", "Muslim/Urdu name"),
    "ALI": ("अली", "LOW", "Muslim/Urdu name"),
    "ALTAMASH": ("अल्तमश", "LOW", "Muslim/Urdu name"),
    "AMEER": ("अमीर", "LOW", "Muslim/Urdu name"),
    "AMIR": ("आमिर", "LOW", "Muslim/Urdu name"),
    "AMJAD": ("अमजद", "LOW", "Muslim/Urdu name"),
    "ANIS": ("अनीस", "LOW", "Muslim/Urdu name"),
    "ANSARI": ("अंसारी", "LOW", "Muslim/Urdu community"),
    "ANWAR": ("अनवर", "LOW", "Muslim/Urdu name"),
    "ARBAZ": ("अरबाज़", "LOW", "Muslim/Urdu name"),
    "ARIF": ("आरिफ", "LOW", "Muslim/Urdu name"),
    "ARMAN": ("अरमान", "LOW", "Muslim/Urdu name"),
    "ARMAAN": ("अरमान", "LOW", "Muslim/Urdu name"),
    "ASGAR": ("असगर", "LOW", "Muslim/Urdu name"),
    "ASHRAF": ("अशरफ", "LOW", "Muslim/Urdu name"),
    "ASIF": ("आसिफ", "LOW", "Muslim/Urdu name"),
    "ASLAM": ("असलम", "LOW", "Muslim/Urdu name"),
    "AYESHA": ("आयशा", "LOW", "Muslim/Urdu name"),
    "AZAD": ("आज़ाद", "LOW", "Muslim/Urdu name"),
    "AZHAR": ("अज़हर", "LOW", "Muslim/Urdu name"),
    "AZIM": ("अज़ीम", "LOW", "Muslim/Urdu name"),
    "BANO": ("बानो", "LOW", "Muslim/Urdu name"),
    "BASIR": ("बशीर", "LOW", "Muslim/Urdu name"),
    "BEGUM": ("बेगम", "LOW", "Muslim/Urdu name"),
    "BILKISH": ("बिल्कीस", "LOW", "Muslim/Urdu name"),
    "DANISH": ("दानिश", "LOW", "Muslim/Urdu name"),
    "DILSHAD": ("दिलशाद", "LOW", "Muslim/Urdu name"),
    "EHSAN": ("एहसान", "LOW", "Muslim/Urdu name"),
    "FAHIM": ("फहीम", "LOW", "Muslim/Urdu name"),
    "FAIJ": ("फैज़", "LOW", "Muslim/Urdu name"),
    "FAIJAN": ("फैज़ान", "LOW", "Muslim/Urdu name"),
    "FAISAL": ("फैसल", "LOW", "Muslim/Urdu name"),
    "FAIZ": ("फैज़", "LOW", "Muslim/Urdu name"),
    "FAIZAN": ("फैज़ान", "LOW", "Muslim/Urdu name"),
    "FARHA": ("फ़रहा", "LOW", "Muslim/Urdu name"),
    "FARHAN": ("फरहान", "LOW", "Muslim/Urdu name"),
    "FARHEEN": ("फ़रहीन", "LOW", "Muslim/Urdu name"),
    "FARIDA": ("फरीदा", "LOW", "Muslim/Urdu name"),
    "FARUKH": ("फारुख", "LOW", "Muslim/Urdu name"),
    "FATIMA": ("फातिमा", "LOW", "Muslim/Urdu name"),
    "FIROJ": ("फिरोज़", "LOW", "Muslim/Urdu name"),
    "FIROZ": ("फिरोज़", "LOW", "Muslim/Urdu name"),
    "GADDI": ("गद्दी", "LOW", "Community name Gaddi"),
    "GULAM": ("गुलाम", "LOW", "Muslim/Urdu name"),
    "GULNAZ": ("गुलनाज़", "LOW", "Muslim/Urdu name"),
    "GULSHAN": ("गुलशन", "LOW", "Muslim/Urdu name"),
    "HABIB": ("हबीब", "LOW", "Muslim/Urdu name"),
    "HAFIZ": ("हाफिज", "LOW", "Muslim/Urdu name"),
    "HAIDAR": ("हैदर", "LOW", "Muslim/Urdu name"),
    "HAMID": ("हामिद", "LOW", "Muslim/Urdu name"),
    "HANIF": ("हनीफ", "LOW", "Muslim/Urdu name"),
    "HARUN": ("हारुन", "LOW", "Muslim/Urdu name"),
    "HASAN": ("हसन", "LOW", "Muslim/Urdu name"),
    "HASEEN": ("हसीन", "LOW", "Muslim/Urdu name"),
    "HASHIM": ("हाशिम", "LOW", "Muslim/Urdu name"),
    "HASIB": ("हसीब", "LOW", "Muslim/Urdu name"),
    "HASINA": ("हसीना", "LOW", "Muslim/Urdu name"),
    "HASRAT": ("हसरत", "LOW", "Muslim/Urdu name"),
    "HUSAIN": ("हुसैन", "LOW", "Muslim/Urdu name"),
    "HUSSAIN": ("हुसैन", "LOW", "Muslim/Urdu name"),
    "IDRISH": ("इदरीश", "LOW", "Muslim/Urdu name"),
    "IJHAR": ("इज़हार", "LOW", "Muslim/Urdu name"),
    "IKRAM": ("इकराम", "LOW", "Muslim/Urdu name"),
    "ILTAF": ("इल्ताफ", "LOW", "Muslim/Urdu name"),
    "IMAM": ("इमाम", "LOW", "Muslim/Urdu name"),
    "IMRAN": ("इमरान", "LOW", "Muslim/Urdu name"),
    "INAMUL": ("इनामुल", "LOW", "Muslim/Urdu name"),
    "INTAKHAB": ("इंतखाब", "LOW", "Muslim/Urdu name"),
    "IQBAL": ("इकबाल", "LOW", "Muslim/Urdu name"),
    "IRFAN": ("इरफ़ान", "LOW", "Muslim/Urdu name"),
    "IRSHAD": ("इरशाद", "LOW", "Muslim/Urdu name"),
    "ISHTIYAQ": ("इश्तियाक", "LOW", "Muslim/Urdu name"),
    "ISMAIL": ("इस्माइल", "LOW", "Muslim/Urdu name"),
    "JABBAR": ("जब्बार", "LOW", "Muslim/Urdu name"),
    "JAFER": ("जाफर", "LOW", "Muslim/Urdu name"),
    "JAHAN": ("जहाँ", "LOW", "Muslim/Urdu name"),
    "JAHID": ("ज़ाहिद", "LOW", "Muslim/Urdu name"),
    "JAHIR": ("ज़ाहिर", "LOW", "Muslim/Urdu name"),
    "JAHARINA": ("जहरीना", "LOW", "Muslim/Urdu name"),
    "JAKIR": ("ज़ाकिर", "LOW", "Muslim/Urdu name"),
    "JALAL": ("जलाल", "LOW", "Muslim/Urdu name"),
    "JAMAL": ("जमाल", "LOW", "Muslim/Urdu name"),
    "JAMALUDDIN": ("जमालुद्दीन", "LOW", "Muslim/Urdu name"),
    "JAMILA": ("जमीला", "LOW", "Muslim/Urdu name"),
    "JAMIL": ("जमील", "LOW", "Muslim/Urdu name"),
    "JARINA": ("ज़रीना", "LOW", "Muslim/Urdu name"),
    "JAVED": ("जावेद", "LOW", "Muslim/Urdu name"),
    "JISAN": ("जीशान", "LOW", "Muslim/Urdu name"),
    "JUBAIDA": ("ज़ुबैदा", "LOW", "Muslim/Urdu name"),
    "JUBER": ("जुबैर", "LOW", "Muslim/Urdu name"),
    "JUMMAN": ("जुम्मन", "LOW", "Muslim/Urdu name"),
    "KABIR": ("कबीर", "LOW", "Muslim/Urdu name"),
    "KALAM": ("कलाम", "LOW", "Muslim/Urdu name"),
    "KALIM": ("कलीम", "LOW", "Muslim/Urdu name"),
    "KAMARUDDIN": ("कमरुद्दीन", "LOW", "Muslim/Urdu name"),
    "KAMRAN": ("कामरान", "LOW", "Muslim/Urdu name"),
    "KARIM": ("करीम", "LOW", "Muslim/Urdu name"),
    "KASIM": ("कासिम", "LOW", "Muslim/Urdu name"),
    "KHAIRUL": ("खैरुल", "LOW", "Muslim/Urdu name"),
    "KHAN": ("खान", "LOW", "Muslim/Urdu surname"),
    "KHATOON": ("खातून", "LOW", "Muslim/Urdu name"),
    "KHATUN": ("खातून", "LOW", "Muslim/Urdu name"),
    "KHURSHEED": ("खुर्शीद", "LOW", "Muslim/Urdu name"),
    "KISHMAT": ("किस्मत", "LOW", "Muslim/Urdu name"),
    "KUDDUS": ("कुद्दुस", "LOW", "Muslim/Urdu name"),
    "MAHBOOB": ("महबूब", "LOW", "Muslim/Urdu name"),
    "MAHFOOZ": ("महफूज़", "LOW", "Muslim/Urdu name"),
    "MAHMUD": ("महमूद", "LOW", "Muslim/Urdu name"),
    "MAINUDDIN": ("मैनुद्दीन", "LOW", "Muslim/Urdu name"),
    "MAIRUN": ("मैरुन", "LOW", "Muslim/Urdu name"),
    "MAJID": ("माजिद", "LOW", "Muslim/Urdu name"),
    "MANSOOR": ("मंसूर", "LOW", "Muslim/Urdu name"),
    "MARIYAM": ("मरियम", "LOW", "Muslim/Urdu name"),
    "MASOOM": ("मासूम", "LOW", "Muslim/Urdu name"),
    "MD": ("मोहम्मद", "LOW", "Abbreviation Md -> मोहम्मद"),
    "MERAJ": ("मेराज", "LOW", "Muslim/Urdu name"),
    "MOHD": ("मोहम्मद", "LOW", "Abbreviation Mohd -> मोहम्मद"),
    "MOHAMMAD": ("मोहम्मद", "LOW", "Muslim/Urdu name"),
    "MOHAMMED": ("मोहम्मद", "LOW", "Muslim/Urdu name"),
    "MUBARAK": ("मुबारक", "LOW", "Muslim/Urdu name"),
    "MUBIN": ("मुबीन", "LOW", "Muslim/Urdu name"),
    "MUMTAZ": ("मुमताज़", "LOW", "Muslim/Urdu name"),
    "MUNIR": ("मुनीर", "LOW", "Muslim/Urdu name"),
    "MUSHTAQ": ("मुश्ताक", "LOW", "Muslim/Urdu name"),
    "MUSTAFA": ("मुस्तफ़ा", "LOW", "Muslim/Urdu name"),
    "MUSTAK": ("मुश्ताक", "LOW", "Muslim/Urdu name"),
    "NABI": ("नबी", "LOW", "Muslim/Urdu name"),
    "NADEEM": ("नदीम", "LOW", "Muslim/Urdu name"),
    "NAEEM": ("नईम", "LOW", "Muslim/Urdu name"),
    "NAFEES": ("नफीस", "LOW", "Muslim/Urdu name"),
    "NAIM": ("नईम", "LOW", "Muslim/Urdu name"),
    "NAJIR": ("नाज़िर", "LOW", "Muslim/Urdu name"),
    "NAJAMA": ("नज़मा", "LOW", "Muslim/Urdu name"),
    "NAJMA": ("नज़मा", "LOW", "Muslim/Urdu name"),
    "NAJIYA": ("नाज़िया", "LOW", "Muslim/Urdu name"),
    "NARGIS": ("नर्गिस", "LOW", "Muslim/Urdu name"),
    "NASEEM": ("नसीम", "LOW", "Muslim/Urdu name"),
    "NASIM": ("नसीम", "LOW", "Muslim/Urdu name"),
    "NASIR": ("नासिर", "LOW", "Muslim/Urdu name"),
    "NASRIN": ("नसरीन", "LOW", "Muslim/Urdu name"),
    "NAUSHAD": ("नौशाद", "LOW", "Muslim/Urdu name"),
    "NAWAJ": ("नवाज़", "LOW", "Muslim/Urdu name"),
    "NAWAB": ("नवाब", "LOW", "Muslim/Urdu name"),
    "NAZIM": ("नाज़िम", "LOW", "Muslim/Urdu name"),
    "NAZIR": ("नाज़िर", "LOW", "Muslim/Urdu name"),
    "NAZMA": ("नज़मा", "LOW", "Muslim/Urdu name"),
    "NEHAL": ("नेहाल", "LOW", "Muslim/Urdu name"),
    "NEJAM": ("निज़ाम", "LOW", "Muslim/Urdu name"),
    "NIAZ": ("नियाज़", "LOW", "Muslim/Urdu name"),
    "NIDA": ("निदा", "LOW", "Muslim/Urdu name"),
    "NIKAHAT": ("निकहत", "LOW", "Muslim/Urdu name"),
    "NIZAM": ("निज़ाम", "LOW", "Muslim/Urdu name"),
    "NIZAMUDDIN": ("निज़ामुद्दीन", "LOW", "Muslim/Urdu name"),
    "NOOR": ("नूर", "LOW", "Muslim/Urdu name"),
    "NOORJAHAN": ("नूरजहाँ", "LOW", "Muslim/Urdu name"),
    "PARVEEN": ("परवीन", "LOW", "Muslim/Urdu name"),
    "PARVEZ": ("परवेज़", "LOW", "Muslim/Urdu name"),
    "QASIM": ("कासिम", "LOW", "Muslim/Urdu name"),
    "QURESHI": ("कुरैशी", "LOW", "Muslim/Urdu surname"),
    "RABIA": ("राबिया", "LOW", "Muslim/Urdu name"),
    "RAEES": ("रईस", "LOW", "Muslim/Urdu name"),
    "RAFI": ("रफी", "LOW", "Muslim/Urdu name"),
    "RAFIQ": ("रफीक", "LOW", "Muslim/Urdu name"),
    "RAHAT": ("राहत", "LOW", "Muslim/Urdu name"),
    "RAHMAT": ("रहमत", "LOW", "Muslim/Urdu name"),
    "RAISA": ("रईसा", "LOW", "Muslim/Urdu name"),
    "RAJIA": ("रज़िया", "LOW", "Muslim/Urdu name"),
    "RAZIYA": ("रज़िया", "LOW", "Muslim/Urdu name"),
    "RAZIA": ("रज़िया", "LOW", "Muslim/Urdu name"),
    "RAZA": ("रज़ा", "LOW", "Muslim/Urdu name"),
    "REHANA": ("रेहाना", "LOW", "Muslim/Urdu name"),
    "REHAN": ("रेहान", "LOW", "Muslim/Urdu name"),
    "REYAZ": ("रियाज़", "LOW", "Muslim/Urdu name"),
    "RIYAZ": ("रियाज़", "LOW", "Muslim/Urdu name"),
    "RIZWAN": ("रिज़वान", "LOW", "Muslim/Urdu name"),
    "RIJWAN": ("रिजवान", "LOW", "Muslim/Urdu name"),
    "ROKSAR": ("रुखसार", "LOW", "Muslim/Urdu name"),
    "RUKHSAR": ("रुखसार", "LOW", "Muslim/Urdu name"),
    "ROSHAN": ("रोशन", "LOW", "Muslim/Urdu name"),
    "ROJI": ("रोजी", "LOW", "Muslim/Urdu name"),
    "RUBINA": ("रुबीना", "LOW", "Muslim/Urdu name"),
    "RUKHSANA": ("रुखसाना", "LOW", "Muslim/Urdu name"),
    "SABANA": ("शबाना", "LOW", "Muslim/Urdu name"),
    "SHABANA": ("शबाना", "LOW", "Muslim/Urdu name"),
    "SABIR": ("साबिर", "LOW", "Muslim/Urdu name"),
    "SABRUN": ("सबरुन", "LOW", "Muslim/Urdu name"),
    "SADAM": ("सद्दाम", "LOW", "Muslim/Urdu name"),
    "SADDAM": ("सद्दाम", "LOW", "Muslim/Urdu name"),
    "SADIQ": ("सादिक", "LOW", "Muslim/Urdu name"),
    "SAEED": ("सईद", "LOW", "Muslim/Urdu name"),
    "SAFIK": ("शफीक", "LOW", "Muslim/Urdu name"),
    "SAFIQ": ("शफीक", "LOW", "Muslim/Urdu name"),
    "SAHBAJ": ("शाहबाज़", "LOW", "Muslim/Urdu name"),
    "SHAHBAZ": ("शाहबाज़", "LOW", "Muslim/Urdu name"),
    "SAHID": ("शाहिद", "LOW", "Muslim/Urdu name"),
    "SHAHID": ("शाहिद", "LOW", "Muslim/Urdu name"),
    "SAHIL": ("साहिल", "LOW", "Muslim/Urdu name"),
    "SHAHIL": ("साहिल", "LOW", "Muslim/Urdu name"),
    "SAHINA": ("शाहीन", "LOW", "Muslim/Urdu name"),
    "SHAHINA": ("शाहीन", "LOW", "Muslim/Urdu name"),
    "SAHISTA": ("शाहिस्ता", "LOW", "Muslim/Urdu name"),
    "SHAHISTA": ("शाहिस्ता", "LOW", "Muslim/Urdu name"),
    "SAIDA": ("सईदा", "LOW", "Muslim/Urdu name"),
    "SAIF": ("सैफ", "LOW", "Muslim/Urdu name"),
    "SAJID": ("साजिद", "LOW", "Muslim/Urdu name"),
    "SAJIDA": ("साजिदा", "LOW", "Muslim/Urdu name"),
    "SAKIL": ("शकील", "LOW", "Muslim/Urdu name"),
    "SHAKEEL": ("शकील", "LOW", "Muslim/Urdu name"),
    "SALAM": ("सलाम", "LOW", "Muslim/Urdu name"),
    "SALAMAN": ("सलमान", "LOW", "Muslim/Urdu name"),
    "SALMAN": ("सलमान", "LOW", "Muslim/Urdu name"),
    "SALIM": ("सलीम", "LOW", "Muslim/Urdu name"),
    "SALMA": ("सलमा", "LOW", "Muslim/Urdu name"),
    "SAMEER": ("समीर", "LOW", "Muslim/Urdu name"),
    "SAMINA": ("समीना", "LOW", "Muslim/Urdu name"),
    "SAMIR": ("समीर", "LOW", "Muslim/Urdu name"),
    "SAMIULLAH": ("समीउल्लाह", "LOW", "Muslim/Urdu name"),
    "SAMSHER": ("शमशेर", "LOW", "Muslim/Urdu name"),
    "SHAMSHER": ("शमशेर", "LOW", "Muslim/Urdu name"),
    "SANABANO": ("सना बानो", "LOW", "Muslim/Urdu name"),
    "SANA": ("सना", "LOW", "Muslim/Urdu name"),
    "SANAZ": ("शहनाज़", "LOW", "Muslim/Urdu name"),
    "SHAHNAZ": ("शहनाज़", "LOW", "Muslim/Urdu name"),
    "SANJIDA": ("संजीदा", "LOW", "Muslim/Urdu name"),
    "SARFARAJ": ("सरफ़राज़", "LOW", "Muslim/Urdu name"),
    "SARFARAZ": ("सरफ़राज़", "LOW", "Muslim/Urdu name"),
    "SARIK": ("शारिक", "LOW", "Muslim/Urdu name"),
    "SHARIQ": ("शारिक", "LOW", "Muslim/Urdu name"),
    "SARIQ": ("सारिक", "LOW", "Muslim/Urdu name"),
    "SARTAJ": ("सरताज", "LOW", "Muslim/Urdu name"),
    "SARVAR": ("सरवर", "LOW", "Muslim/Urdu name"),
    "SAUKAT": ("शौकत", "LOW", "Muslim/Urdu name"),
    "SHAUKAT": ("शौकत", "LOW", "Muslim/Urdu name"),
    "SAYEED": ("सईद", "LOW", "Muslim/Urdu name"),
    "SHABBO": ("शब्बो", "LOW", "Muslim/Urdu name"),
    "SHABIR": ("शबीर", "LOW", "Muslim/Urdu name"),
    "SHABNAM": ("शबनम", "LOW", "Muslim/Urdu name"),
    "SHADAT": ("सहादत", "LOW", "Muslim/Urdu name (Shadat/Saadat)"),
    "SAADAT": ("सआदत", "LOW", "Muslim/Urdu name"),
    "SHAH": ("शाह", "LOW", "Muslim/Urdu name"),
    "SHAREEF": ("शरीफ", "LOW", "Muslim/Urdu name"),
    "SHAHIN": ("शाहीन", "LOW", "Muslim/Urdu name"),
    "SHAHJAHAN": ("शाहजहाँ", "LOW", "Muslim/Urdu name"),
    "SHAHNAWAZ": ("शाहनवाज़", "LOW", "Muslim/Urdu name"),
    "SHAHROOKH": ("शाहरुख", "LOW", "Muslim/Urdu name"),
    "SHAMEEM": ("शमीम", "LOW", "Muslim/Urdu name"),
    "SHAMIM": ("शमीम", "LOW", "Muslim/Urdu name"),
    "SHEKH": ("शेख", "LOW", "Muslim/Urdu surname"),
    "SHEIKH": ("शेख", "LOW", "Muslim/Urdu surname"),
    "SHER": ("शेर", "LOW", "Muslim/Urdu name"),
    "SHOAIB": ("शोएब", "LOW", "Muslim/Urdu name"),
    "SIKANDAR": ("सिकंदर", "LOW", "Muslim/Urdu name"),
    "SIRAJ": ("सिराज", "LOW", "Muslim/Urdu name"),
    "SIRAJUDDIN": ("सिराजुद्दीन", "LOW", "Muslim/Urdu name"),
    "SOHRAB": ("सोहराब", "LOW", "Muslim/Urdu name"),
    "SUHAIL": ("सुहैल", "LOW", "Muslim/Urdu name"),
    "SULTAN": ("सुल्तान", "LOW", "Muslim/Urdu name"),
    "TABASSOOM": ("तबस्सुम", "LOW", "Muslim/Urdu name"),
    "TABASSUM": ("तबस्सुम", "LOW", "Muslim/Urdu name"),
    "TABREJ": ("तबरेज़", "LOW", "Muslim/Urdu name"),
    "TABREZ": ("तबरेज़", "LOW", "Muslim/Urdu name"),
    "TAHIR": ("ताहिर", "LOW", "Muslim/Urdu name"),
    "TAJ": ("ताज", "LOW", "Muslim/Urdu name"),
    "TALIB": ("तालिब", "LOW", "Muslim/Urdu name"),
    "TARIQ": ("तारिक", "LOW", "Muslim/Urdu name"),
    "TASLEEM": ("तसलीम", "LOW", "Muslim/Urdu name"),
    "TASLIM": ("तसलीम", "LOW", "Muslim/Urdu name"),
    "TASLIMA": ("तस्लीमा", "LOW", "Muslim/Urdu name"),
    "TAUFIQ": ("तौफीक", "LOW", "Muslim/Urdu name"),
    "TOUFIQ": ("तौफीक", "LOW", "Muslim/Urdu name"),
    "USMAN": ("उस्मान", "LOW", "Muslim/Urdu name"),
    "VAHID": ("वाहिद", "LOW", "Muslim/Urdu name"),
    "WAHID": ("वाहिद", "LOW", "Muslim/Urdu name"),
    "WASEEM": ("वसीम", "LOW", "Muslim/Urdu name"),
    "WASIM": ("वसीम", "LOW", "Muslim/Urdu name"),
    "YASIR": ("यासिर", "LOW", "Muslim/Urdu name"),
    "YASMEEN": ("यासमीन", "LOW", "Muslim/Urdu name"),
    "YASMIN": ("यासमीन", "LOW", "Muslim/Urdu name"),
    "YUSUF": ("यूसुफ", "LOW", "Muslim/Urdu name"),
    "ZAFAR": ("ज़फ़र", "LOW", "Muslim/Urdu name"),
    "ZAHID": ("ज़ाहिद", "LOW", "Muslim/Urdu name"),
    "ZAHIDA": ("ज़ाहिदा", "LOW", "Muslim/Urdu name"),
    "ZAHIR": ("ज़ाहिर", "LOW", "Muslim/Urdu name"),
    "ZAKIR": ("ज़ाकिर", "LOW", "Muslim/Urdu name"),
    "ZAMEER": ("ज़मीर", "LOW", "Muslim/Urdu name"),
    "ZARINA": ("ज़रीना", "LOW", "Muslim/Urdu name"),
    "ZAREENA": ("ज़रीना", "LOW", "Muslim/Urdu name"),
    "ZEENAT": ("ज़ीनत", "LOW", "Muslim/Urdu name"),
    "ZIA": ("ज़िया", "LOW", "Muslim/Urdu name"),
    "ZOYA": ("ज़ोया", "LOW", "Muslim/Urdu name"),
    "ZULFIQAR": ("ज़ुल्फ़िक़ार", "LOW", "Muslim/Urdu name"),
}

# Improved Offline Phonetic Transliterator
CONSONANTS = [
    ("KSH", "क्ष"), ("GYA", "ज्ञा"), ("GY", "ज्ञ"), ("TRA", "त्रा"), ("TR", "त्र"),
    ("SHH", "ष"), ("SH", "श"), ("CHH", "छ"), ("CH", "च"),
    ("KH", "ख"), ("GH", "घ"), ("JH", "झ"), ("TH", "थ"),
    ("DH", "ध"), ("BH", "भ"), ("PH", "फ"),
    ("K", "क"), ("G", "ग"), ("J", "ज"), ("T", "त"),
    ("D", "द"), ("N", "न"), ("P", "प"), ("B", "ब"),
    ("M", "म"), ("Y", "य"), ("R", "र"), ("L", "ल"),
    ("V", "व"), ("W", "व"), ("S", "स"), ("H", "ह"),
    ("F", "फ"), ("Z", "ज़"),
]

MATRAS = [
    ("AA", "ा"), ("EE", "ी"), ("OO", "ू"), ("AI", "ै"), ("AU", "ौ"),
    ("A", "ा"), ("I", "ी"), ("U", "ु"), ("E", "े"), ("O", "ो"),
]

VOWELS_START = [
    ("AA", "आ"), ("EE", "ई"), ("OO", "ऊ"), ("AI", "ऐ"), ("AU", "औ"),
    ("A", "अ"), ("I", "इ"), ("U", "उ"), ("E", "ए"), ("O", "ओ"),
]

def rule_based_transliterate(w):
    w = w.upper()
    res = ""
    i = 0
    n = len(w)
    
    while i < n:
        if i == 0:
            matched_vowel = False
            for v_str, v_hindi in VOWELS_START:
                if w.startswith(v_str, i):
                    res += v_hindi
                    i += len(v_str)
                    matched_vowel = True
                    break
            if matched_vowel:
                continue
        
        matched_c = False
        for c_str, c_hindi in CONSONANTS:
            if w.startswith(c_str, i):
                c_len = len(c_str)
                i += c_len
                matched_c = True
                
                matched_m = False
                for m_str, m_hindi in MATRAS:
                    if w.startswith(m_str, i):
                        res += c_hindi
                        if m_str == "A":
                            if i + 1 == n:
                                res += "ा"
                            else:
                                pass
                        elif m_str == "I":
                            # End of word 'i' -> 'ी', middle 'i' -> 'ि'
                            if i + 1 == n:
                                res += "ी"
                            else:
                                res += "ि"
                        else:
                            res += m_hindi
                        i += len(m_str)
                        matched_m = True
                        break
                
                if not matched_m:
                    if i < n and w[i] not in "AEIOU":
                        res += c_hindi + "्"
                    else:
                        res += c_hindi
                break
        
        if not matched_c:
            matched_v = False
            for v_str, v_hindi in VOWELS_START:
                if w.startswith(v_str, i):
                    res += v_hindi
                    i += len(v_str)
                    matched_v = True
                    break
            if not matched_v:
                i += 1
                
    return res

def transliterate_word(word, context_type="student"):
    if not word:
        return "", "HIGH", ""
    w_upper = word.upper().strip()
    
    # 1. Exact match in dictionary
    if w_upper in DICTIONARY:
        h, conf, note = DICTIONARY[w_upper]
        
        # Context-aware adjustments for ambiguous names
        if w_upper == "KRISHNA":
            if context_type == "father":
                h = "कृष्ण"
            elif context_type == "mother":
                h = "कृष्णा"
            else:
                h = "कृष्ण" # default male
        elif w_upper == "RAMA":
            if context_type == "father":
                h = "राम"
            elif context_type == "mother":
                h = "रमा"
        elif w_upper == "SHIVA":
            if context_type == "father":
                h = "शिव"
            elif context_type == "mother":
                h = "शिवा"
                
        return h, conf, note
        
    # 2. Compound matching (e.g. NANDKISHOR, RAMAKANT)
    for length in range(len(w_upper) - 2, 2, -1):
        pre = w_upper[:length]
        suf = w_upper[length:]
        if pre in DICTIONARY and suf in DICTIONARY:
            h_pre, c_pre, n_pre = DICTIONARY[pre]
            h_suf, c_suf, n_suf = DICTIONARY[suf]
            combined_conf = "LOW" if ("LOW" in (c_pre, c_suf)) else "MEDIUM"
            return h_pre + h_suf, combined_conf, f"Compound ({pre}+{suf})"
            
    # 3. Rule based fallback
    h_rule = rule_based_transliterate(w_upper)
    if h_rule:
        return h_rule, "LOW", f"Rule-based transliteration of '{word}' (verify spelling)"
    return word, "LOW", f"Unmatched word '{word}'"

def transliterate_name(full_name_eng, context_type="student"):
    if not full_name_eng or not full_name_eng.strip():
        return "", "HIGH", ""
    
    cleaned = re.sub(r"[^a-zA-Z\s\.]", " ", full_name_eng).strip()
    words = [w for w in cleaned.split() if w]
    
    if not words:
        return "", "HIGH", ""
    
    hindi_words = []
    confidences = []
    notes_list = []
    
    words_upper = [w.upper() for w in words]
    if "KUMARI" in words_upper or "DEVI" in words_upper or context_type == "mother":
        effective_context = "mother"
    elif "KUMAR" in words_upper or "PRASAD" in words_upper or "SINGH" in words_upper or "YADAV" in words_upper or context_type == "father":
        effective_context = "father"
    else:
        effective_context = context_type

    for word in words:
        h_word, conf, note = transliterate_word(word, context_type=effective_context)
        hindi_words.append(h_word)
        confidences.append(conf)
        if note:
            notes_list.append(note)
    
    if "LOW" in confidences:
        final_conf = "LOW"
    elif "MEDIUM" in confidences:
        final_conf = "MEDIUM"
    else:
        final_conf = "HIGH"
        
    final_hindi = " ".join(hindi_words)
    final_note = "; ".join(notes_list)
    return final_hindi, final_conf, final_note

def generate_master_files():
    output_csv_path = Path(r"E:\THPSIC-INTER-COLLEGE\05-reports\HINDI_NAME_TRANSLITERATION_MASTER.csv")
    output_xlsx_path = Path(r"E:\THPSIC-INTER-COLLEGE\05-reports\HINDI_NAME_TRANSLITERATION_MASTER.xlsx")
    
    students = Student.objects.filter(is_active=True).order_by("current_class__name", "current_section__name", "admission_no")
    
    rows = []
    high_count = 0
    med_count = 0
    low_count = 0
    
    for s in students:
        s_name = s.full_name or ""
        f_name = s.father_name or ""
        m_name = s.mother_name or ""
        
        s_hindi, s_conf, s_notes = transliterate_name(s_name, context_type="student")
        f_hindi, f_conf, f_notes = transliterate_name(f_name, context_type="father")
        m_hindi, m_conf, m_notes = transliterate_name(m_name, context_type="mother")
        
        # Overall flag logic:
        # - LOW if ANY name has LOW
        # - MEDIUM if ANY name has MEDIUM (and none LOW)
        # - HIGH ONLY if all three are HIGH
        if "LOW" in (s_conf, f_conf, m_conf):
            overall_flag = "LOW"
            low_count += 1
        elif "MEDIUM" in (s_conf, f_conf, m_conf):
            overall_flag = "MEDIUM"
            med_count += 1
        else:
            overall_flag = "HIGH"
            high_count += 1
            
        all_notes = []
        if s_notes:
            all_notes.append(f"Student: {s_notes}")
        if f_notes:
            all_notes.append(f"Father: {f_notes}")
        if m_notes:
            all_notes.append(f"Mother: {m_notes}")
            
        cls_name = s.current_class.name if s.current_class else ""
        sec_name = s.current_section.name if s.current_section else ""
        
        row = {
            "legacy_sid": s.legacy_sid or "",
            "admission_no": s.admission_no or "",
            "class": cls_name,
            "section": sec_name,
            "full_name": s_name,
            "full_name_hindi_suggested": s_hindi,
            "full_name_confidence": s_conf,
            "father_name": f_name,
            "father_name_hindi_suggested": f_hindi,
            "father_name_confidence": f_conf,
            "mother_name": m_name,
            "mother_name_hindi_suggested": m_hindi,
            "mother_name_confidence": m_conf,
            "overall_review_flag": overall_flag,
            "notes": " | ".join(all_notes),
        }
        rows.append(row)
        
    output_csv_path.parent.mkdir(parents=True, exist_ok=True)
    
    fieldnames = [
        "legacy_sid",
        "admission_no",
        "class",
        "section",
        "full_name",
        "full_name_hindi_suggested",
        "full_name_confidence",
        "father_name",
        "father_name_hindi_suggested",
        "father_name_confidence",
        "mother_name",
        "mother_name_hindi_suggested",
        "mother_name_confidence",
        "overall_review_flag",
        "notes"
    ]
    
    # 1. Write CSV with UTF-8 BOM
    try:
        with output_csv_path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
    except PermissionError:
        print("Note: CSV file is currently open. Will write XLSX.")
        
    # 2. Write formatted XLSX workbook using openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hindi Transliteration Master"
    
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_font = Font(name="Calibri", size=11)
    hindi_font = Font(name="Nirmala UI", size=11)
    
    flag_high_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Soft Green
    flag_high_font = Font(name="Calibri", size=11, bold=True, color="166534")
    
    flag_med_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid") # Soft Yellow
    flag_med_font = Font(name="Calibri", size=11, bold=True, color="854D0E")
    
    flag_low_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Soft Red
    flag_low_font = Font(name="Calibri", size=11, bold=True, color="991B1B")
    
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    ws.append(fieldnames)
    
    for col_idx in range(1, len(fieldnames) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for r_idx, row_dict in enumerate(rows, start=2):
        row_vals = [row_dict[col] for col in fieldnames]
        ws.append(row_vals)
        
        for col_idx, col_name in enumerate(fieldnames, start=1):
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.border = thin_border
            
            if "hindi" in col_name:
                cell.font = hindi_font
            else:
                cell.font = data_font
                
            if col_name == "overall_review_flag":
                val = cell.value
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if val == "HIGH":
                    cell.fill = flag_high_fill
                    cell.font = flag_high_font
                elif val == "MEDIUM":
                    cell.fill = flag_med_fill
                    cell.font = flag_med_font
                elif val == "LOW":
                    cell.fill = flag_low_fill
                    cell.font = flag_low_font
            elif col_name in ("legacy_sid", "admission_no", "class", "section"):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
    ws.auto_filter.ref = ws.dimensions
    
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    try:
        wb.save(output_xlsx_path)
    except PermissionError:
        print("Note: XLSX file is currently open in Excel. Please close it.")
    
    print(f"100% Offline Master Generation Complete!")
    print(f"Master CSV: {output_csv_path}")
    print(f"Master XLSX: {output_xlsx_path}")
    print(f"Total Active Students Processed: {len(rows)}")
    print(f"Overall Review Flags:")
    print(f"  - HIGH Confidence (All 3 names High): {high_count} ({high_count/len(rows)*100:.1f}%)")
    print(f"  - MEDIUM Confidence: {med_count} ({med_count/len(rows)*100:.1f}%)")
    print(f"  - LOW Confidence (Flagged for Office Review): {low_count} ({low_count/len(rows)*100:.1f}%)")

if __name__ == "__main__":
    generate_master_files()
