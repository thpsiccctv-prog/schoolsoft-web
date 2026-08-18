attendance_summary_views_code = '''

# ---------------------------------------------------------------------------
# Attendance Summary Entry & Reports (Part B)
# ---------------------------------------------------------------------------

@login_required
def attendance_summary_entry(request):
    """Batch entry screen to record monthly working days and present days for a class/section."""
    import calendar
    from datetime import date
    from core.models import SchoolClass, Section, Student, AcademicSession, AttendanceSummary

    today = date.today()
    selected_month = int(request.GET.get("month", today.month))
    selected_year = int(request.GET.get("year", today.year))

    classes = SchoolClass.objects.all().order_by("display_order")
    selected_class_id = request.GET.get("class_id")
    selected_class = None
    if selected_class_id:
        selected_class = SchoolClass.objects.filter(pk=selected_class_id).first()
    if not selected_class and classes.exists():
        selected_class = classes.first()

    sections = Section.objects.filter(school_class=selected_class).order_by("name") if selected_class else Section.objects.all().order_by("name")
    selected_section_id = request.GET.get("section_id")
    selected_section = None
    if selected_section_id:
        selected_section = Section.objects.filter(pk=selected_section_id).first()

    session = AcademicSession.objects.filter(is_active=True).first() or AcademicSession.objects.first()

    students = []
    if selected_class:
        qs = Student.objects.filter(is_active=True, current_class=selected_class)
        if selected_section:
            qs = qs.filter(current_section=selected_section)
        students = list(qs.select_related("current_class", "current_section").order_by("roll_no", "full_name"))

    # Load existing summaries for these students in this session/month/year
    existing_summaries = {}
    if students:
        student_ids = [s.id for s in students]
        summaries = AttendanceSummary.objects.filter(
            student_id__in=student_ids,
            session=session,
            year=selected_year,
            month=selected_month,
        )
        for summ in summaries:
            existing_summaries[summ.student_id] = summ

    # Default working days (if previously saved by any student in batch, use that, else default to 24)
    default_working_days = 24
    if existing_summaries:
        first_summ = next(iter(existing_summaries.values()))
        default_working_days = first_summ.total_working_days

    if request.method == "POST":
        post_class_id = request.POST.get("class_id")
        post_section_id = request.POST.get("section_id")
        post_month = int(request.POST.get("month", selected_month))
        post_year = int(request.POST.get("year", selected_year))
        
        try:
            batch_working_days = int(request.POST.get("batch_working_days", default_working_days))
        except (ValueError, TypeError):
            batch_working_days = 24

        if batch_working_days < 1:
            messages.error(request, "Total Working Days must be at least 1.")
            return redirect(f"{request.path}?class_id={selected_class.id if selected_class else ''}&section_id={selected_section.id if selected_section else ''}&month={post_month}&year={post_year}")

        saved_count = 0
        validation_errors = []

        with transaction.atomic():
            for student in students:
                p_days_key = f"present_days_{student.id}"
                w_days_key = f"working_days_{student.id}"
                rem_key = f"remarks_{student.id}"

                p_days_raw = request.POST.get(p_days_key, "").strip()
                if p_days_raw == "":
                    continue  # Skip unentered rows

                try:
                    p_days = int(p_days_raw)
                except ValueError:
                    validation_errors.append(f"{student.full_name}: Invalid number of present days.")
                    continue

                try:
                    w_days = int(request.POST.get(w_days_key, batch_working_days))
                except (ValueError, TypeError):
                    w_days = batch_working_days

                if p_days < 0:
                    validation_errors.append(f"{student.full_name}: Present days cannot be negative.")
                    continue
                if p_days > w_days:
                    validation_errors.append(f"{student.full_name}: Present days ({p_days}) cannot exceed Total Working Days ({w_days}).")
                    continue

                remarks = request.POST.get(rem_key, "").strip()

                AttendanceSummary.objects.update_or_create(
                    student=student,
                    session=session,
                    year=post_year,
                    month=post_month,
                    defaults={
                        "total_working_days": w_days,
                        "days_present": p_days,
                        "remarks": remarks,
                    }
                )
                saved_count += 1

        if validation_errors:
            for err in validation_errors[:5]:
                messages.error(request, err)
            if len(validation_errors) > 5:
                messages.error(request, f"...and {len(validation_errors) - 5} more validation errors.")

        if saved_count > 0:
            messages.success(request, f"Attendance successfully saved for {saved_count} students in {selected_class.name} ({calendar.month_name[post_month]} {post_year})!")
            return redirect(f"{reverse('core:attendance_summary_report')}?class_id={selected_class.id if selected_class else ''}&section_id={selected_section.id if selected_section else ''}&month={post_month}&year={post_year}")

    # Build student row objects for template
    student_rows = []
    for idx, s in enumerate(students, 1):
        summary = existing_summaries.get(s.id)
        p_days = summary.days_present if summary else ""
        w_days = summary.total_working_days if summary else default_working_days
        rem = summary.remarks if summary else ""
        student_rows.append({
            "index": idx,
            "student": s,
            "roll_no": s.roll_no or idx,
            "admission_no": s.admission_no or s.legacy_sid or "",
            "present_days": p_days,
            "working_days": w_days,
            "absent_days": summary.days_absent if summary else "",
            "percentage": summary.attendance_percentage if summary else "",
            "remarks": rem,
            "has_record": (summary is not None),
        })

    months_choices = [(m, calendar.month_name[m]) for m in range(1, 13)]
    years_choices = [2025, 2026, 2027]

    context = {
        "classes": classes,
        "sections": sections,
        "selected_class": selected_class,
        "selected_section": selected_section,
        "selected_month": selected_month,
        "selected_month_name": calendar.month_name[selected_month],
        "selected_year": selected_year,
        "default_working_days": default_working_days,
        "months_choices": months_choices,
        "years_choices": years_choices,
        "student_rows": student_rows,
        "total_students": len(student_rows),
    }
    return render(request, "core/attendance_summary_entry.html", context)


@login_required
def attendance_summary_report(request):
    """Comprehensive Monthly Attendance Summary Report with KPIs and statistics."""
    import calendar
    from datetime import date
    from core.models import SchoolClass, Section, Student, AcademicSession, AttendanceSummary

    today = date.today()
    selected_month = int(request.GET.get("month", today.month))
    selected_year = int(request.GET.get("year", today.year))

    classes = SchoolClass.objects.all().order_by("display_order")
    selected_class_id = request.GET.get("class_id")
    selected_class = None
    if selected_class_id:
        selected_class = SchoolClass.objects.filter(pk=selected_class_id).first()
    if not selected_class and classes.exists():
        selected_class = classes.first()

    sections = Section.objects.filter(school_class=selected_class).order_by("name") if selected_class else Section.objects.all().order_by("name")
    selected_section_id = request.GET.get("section_id")
    selected_section = None
    if selected_section_id:
        selected_section = Section.objects.filter(pk=selected_section_id).first()

    session = AcademicSession.objects.filter(is_active=True).first() or AcademicSession.objects.first()

    students = []
    if selected_class:
        qs = Student.objects.filter(is_active=True, current_class=selected_class)
        if selected_section:
            qs = qs.filter(current_section=selected_section)
        students = list(qs.select_related("current_class", "current_section").order_by("roll_no", "full_name"))

    summaries_map = {}
    if students:
        student_ids = [s.id for s in students]
        summaries = AttendanceSummary.objects.filter(
            student_id__in=student_ids,
            session=session,
            year=selected_year,
            month=selected_month,
        )
        for summ in summaries:
            summaries_map[summ.student_id] = summ

    report_rows = []
    total_present_sum = 0
    total_working_sum = 0
    short_attendance_count = 0
    marked_count = 0

    for idx, s in enumerate(students, 1):
        summ = summaries_map.get(s.id)
        if summ:
            marked_count += 1
            total_present_sum += summ.days_present
            total_working_sum += summ.total_working_days
            pct = summ.attendance_percentage
            if pct < 75.0:
                short_attendance_count += 1
            badge = "danger" if pct < 60.0 else "warning" if pct < 75.0 else "success"
            status_label = "Short Attendance (<60%)" if pct < 60.0 else "Average (60-74%)" if pct < 75.0 else "Eligible (>=75%)"
        else:
            pct = None
            badge = "secondary"
            status_label = "Not Recorded"

        report_rows.append({
            "index": idx,
            "student": s,
            "roll_no": s.roll_no or idx,
            "admission_no": s.admission_no or s.legacy_sid or "",
            "working_days": summ.total_working_days if summ else "-",
            "present_days": summ.days_present if summ else "-",
            "absent_days": summ.days_absent if summ else "-",
            "percentage": pct,
            "badge": badge,
            "status_label": status_label,
            "remarks": summ.remarks if summ else "",
            "is_marked": (summ is not None),
        })

    avg_percentage = round((total_present_sum / total_working_sum * 100), 1) if total_working_sum > 0 else 0.0

    months_choices = [(m, calendar.month_name[m]) for m in range(1, 13)]
    years_choices = [2025, 2026, 2027]

    context = {
        "classes": classes,
        "sections": sections,
        "selected_class": selected_class,
        "selected_section": selected_section,
        "selected_month": selected_month,
        "selected_month_name": calendar.month_name[selected_month],
        "selected_year": selected_year,
        "months_choices": months_choices,
        "years_choices": years_choices,
        "report_rows": report_rows,
        "total_enrolled": len(students),
        "marked_count": marked_count,
        "avg_percentage": avg_percentage,
        "short_attendance_count": short_attendance_count,
    }
    return render(request, "core/attendance_summary_report.html", context)


@login_required
def attendance_summary_report_excel(request):
    """Export filtered monthly attendance report to Excel."""
    import calendar
    from datetime import date
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from core.models import SchoolClass, Section, Student, AcademicSession, AttendanceSummary

    today = date.today()
    selected_month = int(request.GET.get("month", today.month))
    selected_year = int(request.GET.get("year", today.year))

    class_id = request.GET.get("class_id")
    school_class = get_object_or_404(SchoolClass, pk=class_id) if class_id else SchoolClass.objects.first()

    section_id = request.GET.get("section_id")
    section = Section.objects.filter(pk=section_id).first() if section_id else None

    session = AcademicSession.objects.filter(is_active=True).first() or AcademicSession.objects.first()

    qs = Student.objects.filter(is_active=True, current_class=school_class)
    if section:
        qs = qs.filter(current_section=section)
    students = list(qs.select_related("current_class", "current_section").order_by("roll_no", "full_name"))

    summaries_map = {}
    if students:
        student_ids = [s.id for s in students]
        summaries = AttendanceSummary.objects.filter(
            student_id__in=student_ids,
            session=session,
            year=selected_year,
            month=selected_month,
        )
        for summ in summaries:
            summaries_map[summ.student_id] = summ

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Attendance"

    ws.merge_cells("A1:H1")
    ws["A1"] = "THPS INTERMEDIATE COLLEGE - MONTHLY ATTENDANCE REPORT"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
    ws["A1"].alignment = Alignment(horizontal="center")

    sec_name = f" - Section {section.name}" if section else ""
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Class: {school_class.name}{sec_name} | Month: {calendar.month_name[selected_month]} {selected_year} | Session: {session.name if session else '2026-27'}"
    ws["A2"].font = Font(name="Calibri", size=11, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.append([])
    headers = ["# / Roll", "Adm No", "Student Name", "Father Name", "Working Days", "Present Days", "Absent Days", "Attendance %", "Status", "Remarks"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if col_idx in [1, 2, 5, 6, 7, 8] else "left")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    row_num = 5
    for idx, s in enumerate(students, 1):
        summ = summaries_map.get(s.id)
        w_days = summ.total_working_days if summ else ""
        p_days = summ.days_present if summ else ""
        a_days = summ.days_absent if summ else ""
        pct = f"{summ.attendance_percentage}%" if summ else "Not Recorded"
        status = "Eligible (>=75%)" if (summ and summ.attendance_percentage >= 75.0) else "Average (60-74%)" if (summ and summ.attendance_percentage >= 60.0) else "Short Attendance (<60%)" if summ else "-"

        ws.append([
            s.roll_no or idx,
            s.admission_no or s.legacy_sid or "",
            s.full_name,
            s.father_name,
            w_days,
            p_days,
            a_days,
            pct,
            status,
            summ.remarks if summ else ""
        ])
        for col_idx in range(1, 11):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            if col_idx in [1, 2, 5, 6, 7, 8]:
                cell.alignment = Alignment(horizontal="center")
        row_num += 1

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    filename = f"Attendance_{school_class.name}{sec_name}_{calendar.month_name[selected_month]}_{selected_year}.xlsx".replace(" ", "_")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
'''

views_path = r'E:\THPSIC-INTER-COLLEGE\01-source-code\schoolsoft_web\core\views.py'
with open(views_path, 'r', encoding='utf-8') as f:
    content = f.read()

if "def attendance_summary_entry" not in content:
    with open(views_path, 'a', encoding='utf-8') as f:
        f.write(attendance_summary_views_code)
    print("attendance_summary_entry views appended to core/views.py successfully!")
else:
    print("Views already present in core/views.py")
