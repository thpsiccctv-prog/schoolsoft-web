import os

views_path = r'E:\THPSIC-INTER-COLLEGE\01-source-code\schoolsoft_web\core\views.py'
with open(views_path, 'r', encoding='utf-8') as f:
    content = f.read()

feeder_views_code = '''

# ---------------------------------------------------------------------------
# Feeder / Attached Schools (अटैच्ड विद्यालय) Views
# ---------------------------------------------------------------------------

@login_required
def feeder_school_list(request):
    """Directory and overview of all attached/feeder schools."""
    from core.models import FeederSchool, Student, AccountGroup, LedgerAccount
    schools = FeederSchool.objects.all().prefetch_related('students', 'ledger_account')
    
    total_schools = schools.count()
    total_students = Student.objects.filter(is_active=True, feeder_school__isnull=False).count()
    
    school_data = []
    total_demand_all = Decimal("0.00")
    total_received_all = Decimal("0.00")
    total_balance_all = Decimal("0.00")
    
    for s in schools:
        s_enrolled = s.total_enrolled_students
        s_demand = s.total_demand
        s_received = s.total_received
        s_balance = s.balance_due
        
        total_demand_all += s_demand
        total_received_all += s_received
        total_balance_all += s_balance
        
        school_data.append({
            "school": s,
            "enrolled": s_enrolled,
            "demand": s_demand,
            "received": s_received,
            "balance": s_balance,
        })
        
    context = {
        "schools_data": school_data,
        "total_schools": total_schools,
        "total_students": total_students,
        "total_demand": total_demand_all,
        "total_received": total_received_all,
        "total_balance": total_balance_all,
    }
    return render(request, "core/feeder_school_list.html", context)


@login_required
def feeder_school_detail(request, pk):
    """Detailed profile, student roster, and ledger statement for a feeder school."""
    from core.models import FeederSchool, Voucher
    school = get_object_or_404(FeederSchool, pk=pk)
    students = school.students.filter(is_active=True).select_related("current_class", "current_section").order_by("current_class__display_order", "full_name")
    
    q = request.GET.get("q", "").strip()
    if q:
        students = students.filter(
            Q(full_name__icontains=q) |
            Q(father_name__icontains=q) |
            Q(admission_no__icontains=q) |
            Q(legacy_sid__icontains=q)
        )
        
    vouchers = []
    if school.ledger_account:
        vouchers = Voucher.objects.filter(
            credit_account=school.ledger_account,
            is_cancelled=False
        ).select_related("debit_account").order_by("-voucher_date", "-id")
        
    context = {
        "school": school,
        "students": students,
        "vouchers": vouchers,
        "enrolled_count": school.total_enrolled_students,
        "total_demand": school.total_demand,
        "total_received": school.total_received,
        "balance_due": school.balance_due,
        "q": q,
    }
    return render(request, "core/feeder_school_detail.html", context)


@login_required
def feeder_school_create(request):
    """Add a new attached/feeder school."""
    from core.forms import FeederSchoolForm
    from core.models import AccountGroup, LedgerAccount, FeederSchool
    if request.method == "POST":
        form = FeederSchoolForm(request.POST)
        if form.is_valid():
            school = form.save(commit=False)
            debtors_group, _ = AccountGroup.objects.get_or_create(
                name="Sundry Debtors",
                defaults={"group_type": AccountGroup.GroupType.ASSET, "display_order": 15}
            )
            ledger, _ = LedgerAccount.objects.get_or_create(
                name=f"{school.name} A/C",
                defaults={"group": debtors_group, "opening_balance": Decimal("0.00")}
            )
            school.ledger_account = ledger
            school.save()
            messages.success(request, f"Attached school '{school.name}' successfully added with Sundry Debtors ledger.")
            return redirect("core:feeder_school_detail", pk=school.pk)
    else:
        form = FeederSchoolForm()
        
    return render(request, "core/feeder_school_form.html", {"form": form, "title": "Add New Attached School (नया अटैच्ड स्कूल जोड़ें)"})


@login_required
def feeder_school_edit(request, pk):
    """Edit an existing feeder school."""
    from core.forms import FeederSchoolForm
    from core.models import FeederSchool
    school = get_object_or_404(FeederSchool, pk=pk)
    if request.method == "POST":
        form = FeederSchoolForm(request.POST, instance=school)
        if form.is_valid():
            form.save()
            messages.success(request, f"Attached school '{school.name}' updated successfully.")
            return redirect("core:feeder_school_detail", pk=school.pk)
    else:
        form = FeederSchoolForm(instance=school)
        
    return render(request, "core/feeder_school_form.html", {"form": form, "school": school, "title": f"Edit School: {school.name}"})


@login_required
def feeder_school_payment(request, pk):
    """Record payment received from a feeder school."""
    from core.forms import FeederSchoolPaymentForm
    from core.models import FeederSchool, AccountGroup, LedgerAccount, AcademicSession, Voucher, VoucherCounter
    school = get_object_or_404(FeederSchool, pk=pk)
    if not school.ledger_account:
        debtors_group, _ = AccountGroup.objects.get_or_create(
            name="Sundry Debtors",
            defaults={"group_type": AccountGroup.GroupType.ASSET, "display_order": 15}
        )
        ledger, _ = LedgerAccount.objects.get_or_create(
            name=f"{school.name} A/C",
            defaults={"group": debtors_group, "opening_balance": Decimal("0.00")}
        )
        school.ledger_account = ledger
        school.save(update_fields=["ledger_account"])

    session = AcademicSession.objects.filter(is_current=True).first()
    if not session:
        session = AcademicSession.objects.first()

    if request.method == "POST":
        form = FeederSchoolPaymentForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            p_date = cd["payment_date"]
            cash_bank_acc = cd["cash_or_bank_account"]
            amount = cd["amount"]
            mode = cd["payment_mode"]
            ref_no = cd["reference_no"]
            narration = cd["narration"] or f"Affiliation / Registration fee installment from {school.name}"
            
            v_type = Voucher.VoucherType.CASH_RECEIPT if mode == Voucher.PaymentMode.CASH else Voucher.VoucherType.BANK
            
            with transaction.atomic():
                counter, _ = VoucherCounter.objects.select_for_update().get_or_create(
                    session=session, voucher_type=v_type
                )
                counter.last_number += 1
                counter.save()
                v_num_str = f"{v_type}-{session.name[:4] if session else '2026'}-{counter.last_number:04d}"
                
                voucher = Voucher.objects.create(
                    voucher_no=v_num_str,
                    voucher_type=v_type,
                    session=session,
                    voucher_date=p_date,
                    debit_account=cash_bank_acc,
                    credit_account=school.ledger_account,
                    amount=amount,
                    paid_to_or_received_from=school.name,
                    narration=narration,
                    payment_mode=mode,
                    physical_slip_no=ref_no,
                )
                
            messages.success(request, f"Payment of Rs. {amount:,.2f} recorded successfully! Voucher #{voucher.voucher_no}")
            return redirect("core:feeder_school_detail", pk=school.pk)
    else:
        form = FeederSchoolPaymentForm()
        
    context = {
        "school": school,
        "form": form,
        "balance_due": school.balance_due,
    }
    return render(request, "core/feeder_school_payment.html", context)


@login_required
def feeder_school_statement_pdf(request, pk):
    """Generate professional A4 PDF statement for an attached school."""
    from core.models import FeederSchool, Voucher, SchoolProfile
    from core.pdf import build_feeder_school_statement_pdf
    school = get_object_or_404(FeederSchool, pk=pk)
    students = school.students.filter(is_active=True).select_related("current_class", "current_section").order_by("current_class__display_order", "full_name")
    
    vouchers = []
    if school.ledger_account:
        vouchers = list(Voucher.objects.filter(
            credit_account=school.ledger_account,
            is_cancelled=False
        ).select_related("debit_account").order_by("voucher_date", "id"))
        
    profile = SchoolProfile.objects.first()
    pdf_bytes = build_feeder_school_statement_pdf(school, students, vouchers, school_profile=profile)
    
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{school.code or "SCHOOL"}_Statement.pdf"'
    return response


@login_required
def feeder_school_statement_excel(request, pk):
    """Export Feeder School statement & roster to Excel."""
    from core.models import FeederSchool
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    school = get_object_or_404(FeederSchool, pk=pk)
    students = school.students.filter(is_active=True).select_related("current_class", "current_section").order_by("current_class__display_order", "full_name")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "School Statement"
    
    ws.merge_cells("A1:G1")
    t_cell = ws["A1"]
    t_cell.value = "THPS INTERMEDIATE COLLEGE - ATTACHED SCHOOL STATEMENT"
    t_cell.font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
    t_cell.alignment = Alignment(horizontal="center")
    
    ws.merge_cells("A2:G2")
    ws["A2"].value = f"School: {school.name} | Contact: {school.contact_person} ({school.phone}) | Location: {school.village_address}"
    ws["A2"].font = Font(name="Calibri", size=11, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")
    
    ws.append([])
    
    ws.append(["Enrolled Students", "Package Rate / Student", "Total Demand (Rs.)", "Total Received (Rs.)", "Balance Due (Rs.)"])
    ws.append([school.total_enrolled_students, float(school.package_rate_per_student), float(school.total_demand), float(school.total_received), float(school.balance_due)])
    
    ws.append([])
    ws.append(["#", "Adm No / SID", "Student Name", "Father Name", "Class", "Section", "Mobile"])
    
    for idx, s in enumerate(students, 1):
        ws.append([
            idx,
            s.admission_no or s.legacy_sid or "",
            s.full_name,
            s.father_name,
            str(s.current_class or ""),
            str(s.current_section.name if s.current_section else ""),
            s.mobile_primary or s.mobile_secondary or ""
        ])
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{school.code or "SCHOOL"}_Statement.xlsx"'
    wb.save(response)
    return response
'''

if "def feeder_school_list" not in content:
    with open(views_path, 'a', encoding='utf-8') as f:
        f.write(feeder_views_code)
    print("Feeder school views appended to core/views.py successfully!")
else:
    print("Feeder views already present in core/views.py")
