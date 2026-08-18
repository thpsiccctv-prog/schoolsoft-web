import os
import csv
from decimal import Decimal
from pathlib import Path

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "schoolsoft.settings")
sqlite_path = os.environ.get("SCHOOLSOFT_SQLITE_PATH") or os.path.expandvars(r"%LOCALAPPDATA%\THPSIC-InterCollege-SchoolSoft\db.sqlite3")
os.environ["SCHOOLSOFT_SQLITE_PATH"] = sqlite_path

import django
django.setup()

from core.models import Student, FeeStructure, FeeHead, SchoolClass, AcademicSession
from core.fee_engine import calculate_student_due, _academic_month_cutoff, ACADEMIC_MONTHS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

session = AcademicSession.objects.filter(is_active=True).first()

# Filter all active students in Class IX and Class XI
students = Student.objects.filter(
    is_active=True,
    current_class__name__regex=r'^(IX|XI)'
).select_related('current_class', 'current_section').order_by(
    'current_class__display_order', 'current_section__name', 'admission_no'
)

print(f"Total Active Students in Class IX and XI: {students.count()}")

rows = []
total_promoted = 0
total_new = 0

for s in students:
    res = calculate_student_due(student=s, session=session, through_month='AUG')
    
    is_new = res.is_new_student
    if is_new:
        total_new += 1
        st_type = "New Admission (2026)"
    else:
        total_promoted += 1
        st_type = "Promoted / Continuing (Pre-2026)"
        
    cls_str = s.current_class.name if s.current_class else ""
    sec_str = s.current_section.name if s.current_section else ""
    
    # Old Demand (without ₹2000 starter fee for promoted):
    # If promoted, Old Demand was res.gross_demand - 2000; if new, it was res.gross_demand
    old_demand = res.gross_demand if is_new else max(Decimal('0.00'), res.gross_demand - Decimal('2000.00'))
    old_due = max(Decimal('0.00'), old_demand - res.received_amount - res.concession_amount)
    
    new_demand = res.gross_demand
    new_due = res.due_amount
    
    if new_due == Decimal('0.00'):
        status = "Settled (Fully Paid)"
    elif res.received_amount > Decimal('0.00'):
        status = "Partially Paid"
    else:
        status = "Unpaid (Full Due)"
        
    rows.append({
        "sid": s.legacy_sid or "",
        "admission_no": s.admission_no or "",
        "student_name": s.full_name or "",
        "father_name": s.father_name or "",
        "class": cls_str,
        "section": sec_str,
        "student_type": st_type,
        "starter_fee_applied": 2000.00,
        "old_gross_demand": float(old_demand),
        "new_gross_demand": float(new_demand),
        "paid_amount": float(res.received_amount),
        "concession_amount": float(res.concession_amount),
        "old_due_amount": float(old_due),
        "new_due_amount": float(new_due),
        "status": status,
    })

print(f"  Promoted / Continuing Students: {total_promoted}")
print(f"  New Admission Students: {total_new}")

# Save CSV
output_csv = Path(r"E:\THPSIC-INTER-COLLEGE\05-reports\CLASS_IX_XI_STARTER_FEE_AUDIT.csv")
output_xlsx = Path(r"E:\THPSIC-INTER-COLLEGE\05-reports\CLASS_IX_XI_STARTER_FEE_AUDIT.xlsx")

fieldnames = list(rows[0].keys())
with open(output_csv, mode="w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)

# Save Styled Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Class IX & XI Starter Fee Audit"

headers = [
    "SID", "Adm No", "Student Name", "Father Name", "Class", "Section",
    "Student Type", "Starter Fee (Board Reg+Kit)", "Old Gross Demand", "New Gross Demand",
    "Paid Amount", "Concession", "Old Due", "New Due (Office Reconciled)", "Status"
]

header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

ws.append(headers)
for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

for r_idx, r in enumerate(rows, start=2):
    row_vals = list(r.values())
    ws.append(row_vals)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=r_idx, column=col_idx)
        cell.border = thin_border
        if col_idx in [1, 2, 5, 6]:
            cell.alignment = Alignment(horizontal="center", vertical="center")

ws.auto_filter.ref = ws.dimensions

for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        val_str = str(cell.value or "")
        if len(val_str) > max_len: max_len = len(val_str)
    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save(output_xlsx)
print(f"\nGenerated Audit Files:")
print(f"  CSV:  {output_csv}")
print(f"  XLSX: {output_xlsx}")
